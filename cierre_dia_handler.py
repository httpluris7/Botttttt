"""
HANDLER DE CIERRE DE DÍA v1.1 (Telegram)
=========================================
Interfaz de Telegram para el cierre de día.

CAMBIOS v1.1:
- FIX Bug #5: Usar números directamente, no len()
- FIX Bug #6: Advertir si ya existe cierre de hoy + opción actualizar

Menú:
📅 Cierre de día
├── 🔄 Cerrar día actual
├── 📊 Ver resumen del día
├── 📂 Ver día anterior
└── ❌ Cancelar
"""

import logging
from datetime import datetime, timedelta
from typing import Optional

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes,
    ConversationHandler,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters
)

from cierre_dia import CierreDia

logger = logging.getLogger(__name__)

# Estados
CIERRE_MENU = 400
CIERRE_CONFIRMAR = 401
CIERRE_HISTORICO = 402
CIERRE_EXISTE = 403  # NUEVO: Estado para cuando ya existe archivo


class CierreDiaHandler:
    """Handler de Telegram para cierre de día"""
    
    def __init__(self, cierre: CierreDia, es_admin_func, teclado_admin):
        self.cierre = cierre
        self.es_admin = es_admin_func
        self.teclado_admin = teclado_admin
        logger.info("[CIERRE_HANDLER] Handler de cierre de día v1.1 inicializado")
    
    def get_conversation_handler(self):
        """Devuelve el ConversationHandler"""
        return ConversationHandler(
            entry_points=[
                MessageHandler(filters.Regex("^📅 Cierre de día$"), self.inicio),
                CommandHandler("cierre", self.inicio),
            ],
            states={
                CIERRE_MENU: [
                    CallbackQueryHandler(self.mostrar_resumen, pattern="^cierre_resumen$"),
                    CallbackQueryHandler(self.confirmar_cierre, pattern="^cierre_ejecutar$"),
                    CallbackQueryHandler(self.listar_historico, pattern="^cierre_historico$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cierre_cancelar$"),
                ],
                CIERRE_CONFIRMAR: [
                    CallbackQueryHandler(self.ejecutar_cierre, pattern="^cierre_confirmar_si$"),
                    CallbackQueryHandler(self.volver_menu, pattern="^cierre_volver$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cierre_cancelar$"),
                ],
                CIERRE_EXISTE: [
                    CallbackQueryHandler(self.actualizar_cierre_existente, pattern="^cierre_actualizar$"),
                    CallbackQueryHandler(self.volver_menu, pattern="^cierre_volver$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cierre_cancelar$"),
                ],
                CIERRE_HISTORICO: [
                    CallbackQueryHandler(self.ver_excel_historico, pattern="^cierre_ver_"),
                    CallbackQueryHandler(self.volver_menu, pattern="^cierre_volver$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cierre_cancelar$"),
                ],
            },
            fallbacks=[
                CommandHandler("cancelar", self.cancelar),
                MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
            ],
        )
    
    # ============================================================
    # INICIO
    # ============================================================
    
    async def inicio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra el menú de cierre de día"""
        user = update.effective_user
        
        if not self.es_admin(user.id):
            await update.message.reply_text(
                "❌ Solo para responsables.",
                reply_markup=self.teclado_admin
            )
            return ConversationHandler.END
        
        verificacion = self.cierre.verificar_cierre_seguro()
        excel_activo = self.cierre.obtener_excel_activo()
        
        texto = (
            f"📅 CIERRE DE DÍA\n\n"
            f"📁 Excel activo: {excel_activo}\n\n"
            f"📊 Estado del día:\n"
            f"✅ Conductores terminaron: {verificacion['conductores_terminaron']}\n"
            f"🚛 Conductores disponibles: {verificacion.get('conductores_disponibles', 0)}\n"
            f"⏳ Viajes pendientes: {verificacion['viajes_pendientes']}\n"
            f"🏁 Viajes completados: {verificacion['viajes_completados']}\n"
        )
        
        if verificacion['advertencia']:
            texto += f"\n⚠️ {verificacion['advertencia']}\n"
        
        texto += "\n¿Qué quieres hacer?"
        
        keyboard = [
            [InlineKeyboardButton("📊 Ver resumen detallado", callback_data="cierre_resumen")],
            [InlineKeyboardButton("🔄 Cerrar día actual", callback_data="cierre_ejecutar")],
            [InlineKeyboardButton("📂 Ver días anteriores", callback_data="cierre_historico")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")]
        ]
        
        await update.message.reply_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return CIERRE_MENU
    
    # ============================================================
    # RESUMEN DETALLADO - BUGFIX #5
    # ============================================================
    
    async def mostrar_resumen(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra resumen detallado del día - BUGFIX #5: usar números directamente"""
        query = update.callback_query
        await query.answer()
        
        analisis = self.cierre.analizar_excel_actual()
        
        texto = "📊 RESUMEN DETALLADO\n\n"
        
        # BUGFIX #5: Usar números directamente, no len()
        num_terminaron = analisis.get('conductores_terminaron', 0)
        texto += f"✅ Conductores que terminaron: {num_terminaron}\n"
        
        num_disponibles = analisis.get('conductores_disponibles', 0)
        texto += f"🚛 Conductores disponibles: {num_disponibles}\n"
        
        num_pendientes = analisis.get('viajes_pendientes', 0)
        texto += f"⏳ Viajes pendientes: {num_pendientes}\n"
        
        num_completados = analisis.get('viajes_completados', 0)
        texto += f"🏁 Viajes completados: {num_completados}\n"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Cerrar día actual", callback_data="cierre_ejecutar")],
            [InlineKeyboardButton("⬅️ Volver", callback_data="cierre_volver")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")]
        ]
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return CIERRE_MENU
    
    # ============================================================
    # CONFIRMAR CIERRE - BUGFIX #6
    # ============================================================
    
    async def confirmar_cierre(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide confirmación para el cierre - BUGFIX #6: avisar si ya existe"""
        query = update.callback_query
        await query.answer()
        
        verificacion = self.cierre.verificar_cierre_seguro()
        nombre_nuevo = self.cierre.generar_nombre_excel()
        
        # BUGFIX #6: Verificar si ya existe el archivo de hoy
        if verificacion.get('excel_hoy_existe'):
            texto = (
                f"⚠️ YA EXISTE CIERRE DE HOY\n\n"
                f"📁 Archivo: {nombre_nuevo}\n\n"
                f"¿Qué deseas hacer?\n\n"
                f"• *Actualizar*: Sobrescribir el archivo existente con los cambios actuales\n"
                f"• *Cancelar*: No hacer nada"
            )
            
            keyboard = [
                [InlineKeyboardButton("🔄 Actualizar cierre", callback_data="cierre_actualizar")],
                [InlineKeyboardButton("⬅️ Volver", callback_data="cierre_volver")],
                [InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")]
            ]
            
            await query.edit_message_text(
                texto,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            
            return CIERRE_EXISTE
        
        # No existe, continuar normal
        texto = (
            f"⚠️ CONFIRMAR CIERRE DE DÍA\n\n"
            f"Se creará: {nombre_nuevo}\n\n"
            f"Se exportarán:\n"
            f"• {verificacion['conductores_terminaron']} conductores (ubicación actualizada)\n"
            f"• {verificacion['viajes_pendientes']} viajes pendientes\n\n"
        )
        
        if verificacion['advertencia']:
            texto += f"⚠️ ADVERTENCIA: {verificacion['advertencia']}\n\n"
        
        texto += "¿Confirmar cierre?"
        
        keyboard = [
            [InlineKeyboardButton("✅ Sí, cerrar día", callback_data="cierre_confirmar_si")],
            [InlineKeyboardButton("⬅️ Volver", callback_data="cierre_volver")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")]
        ]
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return CIERRE_CONFIRMAR
    
    async def actualizar_cierre_existente(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """BUGFIX #6: Actualiza el cierre existente"""
        query = update.callback_query
        await query.answer()
        
        await query.edit_message_text(
            "🔄 Actualizando cierre de día...\n\n"
            "⏳ Procesando cambios..."
        )
        
        # Ejecutar cierre (sobrescribirá el archivo existente)
        resultado = self.cierre.ejecutar_cierre()
        
        if resultado['exito']:
            texto = (
                f"✅ CIERRE ACTUALIZADO\n\n"
                f"📁 Excel: {resultado['excel_nuevo']}\n\n"
                f"📊 Exportados:\n"
                f"• {resultado['conductores_exportados']} conductores\n"
                f"• {resultado['viajes_pendientes']} viajes pendientes\n"
                f"• {resultado['viajes_completados']} viajes archivados\n\n"
                f"☁️ Drive actualizado: {'✅' if resultado.get('drive_subido') else '❌'}"
            )
        else:
            texto = (
                f"❌ ERROR EN ACTUALIZACIÓN\n\n"
                f"Errores:\n"
            )
            for error in resultado['errores']:
                texto += f"• {error}\n"
        
        await query.edit_message_text(texto)
        
        await query.message.reply_text(
            "¿Qué más necesitas?",
            reply_markup=self.teclado_admin
        )
        
        return ConversationHandler.END
    
    async def ejecutar_cierre(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ejecuta el cierre de día"""
        query = update.callback_query
        await query.answer()
        
        await query.edit_message_text(
            "🔄 Ejecutando cierre de día...\n\n"
            "⏳ Analizando Excel..."
        )
        
        resultado = self.cierre.ejecutar_cierre()
        
        if resultado['exito']:
            texto = (
                f"✅ CIERRE COMPLETADO\n\n"
                f"📁 Excel original: {resultado['excel_original']} (sin modificar)\n"
                f"📁 Excel CREADO: {resultado['excel_nuevo']}\n\n"
                f"📊 Exportados:\n"
                f"• {resultado['conductores_exportados']} conductores\n"
                f"• {resultado['viajes_pendientes']} viajes pendientes\n"
                f"• {resultado['viajes_completados']} viajes archivados\n\n"
                f"☁️ Drive actualizado: {'✅' if resultado.get('drive_subido') else '❌'}"
            )
        else:
            texto = (
                f"❌ ERROR EN CIERRE\n\n"
                f"Errores:\n"
            )
            for error in resultado['errores']:
                texto += f"• {error}\n"
        
        await query.edit_message_text(texto)
        
        await query.message.reply_text(
            "¿Qué más necesitas?",
            reply_markup=self.teclado_admin
        )
        
        return ConversationHandler.END
    
    # ============================================================
    # HISTÓRICO
    # ============================================================
    
    async def listar_historico(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Lista los Excels históricos"""
        query = update.callback_query
        await query.answer()
        
        excels = self.cierre.listar_excels_historicos(limite=7)
        
        if not excels:
            texto = "📂 HISTÓRICO\n\nNo hay Excels históricos disponibles."
            keyboard = [
                [InlineKeyboardButton("⬅️ Volver", callback_data="cierre_volver")],
                [InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")]
            ]
        else:
            texto = "📂 HISTÓRICO DE DÍAS\n\nSelecciona un día para ver detalles:\n\n"
            
            keyboard = []
            for excel in excels:
                fecha = excel['fecha_modificacion'].strftime("%d/%m/%Y")
                nombre = excel['nombre']
                keyboard.append([
                    InlineKeyboardButton(
                        f"📄 {nombre} ({fecha})", 
                        callback_data=f"cierre_ver_{nombre}"
                    )
                ])
            
            keyboard.append([InlineKeyboardButton("⬅️ Volver", callback_data="cierre_volver")])
            keyboard.append([InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")])
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return CIERRE_HISTORICO
    
    async def ver_excel_historico(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra info de un Excel histórico"""
        query = update.callback_query
        await query.answer()
        
        nombre_excel = query.data.replace("cierre_ver_", "")
        
        excels = self.cierre.listar_excels_historicos(limite=30)
        excel_info = next((e for e in excels if e['nombre'] == nombre_excel), None)
        
        if not excel_info:
            texto = f"❌ No se encontró el archivo: {nombre_excel}"
        else:
            from openpyxl import load_workbook
            try:
                wb = load_workbook(excel_info['ruta'])
                ws = wb.active
                total_filas = ws.max_row - 1
                wb.close()
                
                texto = (
                    f"📄 {nombre_excel}\n\n"
                    f"📅 Fecha: {excel_info['fecha_modificacion'].strftime('%d/%m/%Y %H:%M')}\n"
                    f"📏 Tamaño: {excel_info['tamaño'] / 1024:.1f} KB\n"
                    f"📊 Filas de datos: {total_filas}\n"
                )
            except Exception as e:
                texto = f"❌ Error leyendo archivo: {e}"
        
        keyboard = [
            [InlineKeyboardButton("⬅️ Volver a lista", callback_data="cierre_historico")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")]
        ]
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return CIERRE_HISTORICO
    
    # ============================================================
    # NAVEGACIÓN
    # ============================================================
    
    async def volver_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve al menú principal de cierre"""
        query = update.callback_query
        await query.answer()
        
        verificacion = self.cierre.verificar_cierre_seguro()
        excel_activo = self.cierre.obtener_excel_activo()
        
        texto = (
            f"📅 CIERRE DE DÍA\n\n"
            f"📁 Excel activo: {excel_activo}\n\n"
            f"📊 Estado del día:\n"
            f"✅ Conductores terminaron: {verificacion['conductores_terminaron']}\n"
            f"🚛 Conductores disponibles: {verificacion.get('conductores_disponibles', 0)}\n"
            f"⏳ Viajes pendientes: {verificacion['viajes_pendientes']}\n"
            f"🏁 Viajes completados: {verificacion['viajes_completados']}\n"
        )
        
        if verificacion['advertencia']:
            texto += f"\n⚠️ {verificacion['advertencia']}\n"
        
        texto += "\n¿Qué quieres hacer?"
        
        keyboard = [
            [InlineKeyboardButton("📊 Ver resumen detallado", callback_data="cierre_resumen")],
            [InlineKeyboardButton("🔄 Cerrar día actual", callback_data="cierre_ejecutar")],
            [InlineKeyboardButton("📂 Ver días anteriores", callback_data="cierre_historico")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="cierre_cancelar")]
        ]
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return CIERRE_MENU
    
    async def cancelar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela (mensaje)"""
        context.user_data.clear()
        await update.message.reply_text(
            "❌ Operación cancelada.",
            reply_markup=self.teclado_admin
        )
        return ConversationHandler.END
    
    async def cancelar_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela (callback)"""
        query = update.callback_query
        await query.answer()
        context.user_data.clear()
        
        await query.edit_message_text("❌ Operación cancelada.")
        await query.message.reply_text(
            "¿Qué más necesitas?",
            reply_markup=self.teclado_admin
        )
        return ConversationHandler.END


# ============================================================
# FUNCIÓN PARA INTEGRAR EN BOT
# ============================================================

def crear_cierre_handler(cierre: CierreDia, es_admin_func, teclado_admin):
    """
    Crea el handler de cierre de día.
    """
    return CierreDiaHandler(cierre, es_admin_func, teclado_admin)
