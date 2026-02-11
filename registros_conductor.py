"""
REGISTROS DE CONDUCTOR v1.1
============================
Permite a los conductores registrar horas de llegada/salida
en puntos de carga y descarga.

CAMBIOS v1.1:
- FIX Bug #3: Mensaje claro cuando viaje estÃ¡ COMPLETADO
- FIX Bug #4: No bloquear descarga despuÃ©s de carga

Flujo:
1. Conductor pulsa "ğŸ“ Registros"
2. Selecciona "ğŸ“¥ Carga" o "ğŸ“¤ Descarga"
3. Selecciona "ğŸš› Llegada" o "ğŸ Salida"
4. Se registra la hora actual en el Excel y se sube a Drive

Columnas Excel:
- Col O (15): HORA LLEGADA carga
- Col P (16): HORA SALIDA carga
- Col R (18): HORA LLEGADA descarga
- Col S (19): HORA SALIDA descarga
"""

import sqlite3
import logging
from datetime import datetime
from typing import Dict, Optional
from pathlib import Path

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes,
    ConversationHandler,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters
)

logger = logging.getLogger(__name__)

# ============================================================
# ESTADOS DEL CONVERSATION HANDLER
# ============================================================
REG_TIPO = 200          # Seleccionar Carga o Descarga
REG_ACCION = 201        # Seleccionar Llegada o Salida
REG_CONFIRMAR = 202     # Confirmar registro

# ============================================================
# COLUMNAS EXCEL (Ã­ndices openpyxl, 1-indexed)
# ============================================================
COLUMNAS_EXCEL = {
    'carga_llegada': 15,      # Col O
    'carga_salida': 16,       # Col P
    'descarga_llegada': 18,   # Col R
    'descarga_salida': 19,    # Col S
}

# Columnas de datos del conductor (izquierda)
COLUMNAS_CONDUCTOR = {
    'ubicacion': 2,           # Col B - UBICACIÃ“N
    'hora_llegada': 3,        # Col C - H.LL
    'hora_salida': 4,         # Col D - H.SA
}


class RegistrosConductor:
    """
    Gestiona los registros de llegada/salida de conductores.
    """
    
    def __init__(self, excel_path: str, db_path: str, subir_drive_func=None):
        self.excel_path = excel_path
        self.db_path = db_path
        self.subir_drive = subir_drive_func
        logger.info("[REGISTROS] MÃ³dulo de registros v1.1 inicializado")
    
    def get_conversation_handler(self):
        """Devuelve el ConversationHandler para registros"""
        return ConversationHandler(
            entry_points=[
                MessageHandler(filters.Regex("^ğŸ“ Registros$"), self.inicio),
            ],
            states={
                REG_TIPO: [
                    CallbackQueryHandler(self.seleccionar_tipo, pattern="^reg_tipo_"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^reg_cancelar$"),
                ],
                REG_ACCION: [
                    CallbackQueryHandler(self.seleccionar_accion, pattern="^reg_accion_"),
                    CallbackQueryHandler(self.volver_tipo, pattern="^reg_volver_tipo$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^reg_cancelar$"),
                ],
                REG_CONFIRMAR: [
                    CallbackQueryHandler(self.confirmar_registro, pattern="^reg_confirmar_"),
                    CallbackQueryHandler(self.volver_accion, pattern="^reg_volver_accion$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^reg_cancelar$"),
                ],
            },
            fallbacks=[
                CommandHandler("cancelar", self.cancelar),
                MessageHandler(filters.Regex("^âŒ Cancelar$"), self.cancelar),
            ],
        )
    
    # ============================================================
    # OBTENER VIAJE ACTIVO DEL CONDUCTOR
    # ============================================================
    
    def _obtener_viaje_conductor(self, telegram_id: int) -> Optional[Dict]:
        """
        Obtiene el viaje activo del conductor.
        BUGFIX #3: TambiÃ©n devuelve viajes completados para mostrar mensaje correcto.
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Primero obtener el nombre del conductor
            cursor.execute("""
                SELECT nombre FROM conductores_empresa 
                WHERE telegram_id = ?
            """, (telegram_id,))
            row = cursor.fetchone()
            
            if not row:
                conn.close()
                return None
            
            nombre = row['nombre']
            
            # Buscar viaje activo (pendiente, en_ruta, asignado)
            cursor.execute("""
                SELECT * FROM viajes_empresa 
                WHERE conductor_asignado LIKE ? 
                AND estado IN ('pendiente', 'en_ruta', 'asignado')
                ORDER BY id DESC
                LIMIT 1
            """, (f"%{nombre}%",))
            
            viaje = cursor.fetchone()
            
            if viaje:
                conn.close()
                return dict(viaje)
            
            # BUGFIX #3: Si no hay viaje activo, buscar si tiene viaje completado
            cursor.execute("""
                SELECT * FROM viajes_empresa 
                WHERE conductor_asignado LIKE ? 
                AND estado = 'completado'
                ORDER BY id DESC
                LIMIT 1
            """, (f"%{nombre}%",))
            
            viaje_completado = cursor.fetchone()
            conn.close()
            
            if viaje_completado:
                # Devolver viaje con marca especial de completado
                viaje_dict = dict(viaje_completado)
                viaje_dict['_completado'] = True
                return viaje_dict
            
            return None
            
        except Exception as e:
            logger.error(f"[REGISTROS] Error obteniendo viaje: {e}")
            return None
    
    # ============================================================
    # ACTUALIZAR EXCEL
    # ============================================================
    
    def _actualizar_hora_excel(self, fila_excel: int, tipo: str, accion: str) -> bool:
        """
        Actualiza la hora en el Excel.
        """
        try:
            from openpyxl import load_workbook
            
            if not Path(self.excel_path).exists():
                logger.error(f"[REGISTROS] Excel no encontrado: {self.excel_path}")
                return False
            
            clave = f"{tipo}_{accion}"
            columna = COLUMNAS_EXCEL.get(clave)
            
            if not columna:
                logger.error(f"[REGISTROS] Columna no encontrada para: {clave}")
                return False
            
            hora_actual = datetime.now().strftime("%H:%M")
            
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            fila_openpyxl = fila_excel + 1
            
            if fila_openpyxl > ws.max_row:
                logger.error(f"[REGISTROS] Fila {fila_openpyxl} fuera de rango")
                return False
            
            celda = ws.cell(row=fila_openpyxl, column=columna)
            valor_anterior = celda.value
            celda.value = hora_actual
            
            wb.save(self.excel_path)
            wb.close()
            
            logger.info(f"[REGISTROS] âœ… Excel actualizado: Fila {fila_openpyxl}, "
                       f"Col {columna} ({clave}) = '{hora_actual}' (antes: '{valor_anterior}')")
            
            # Subir a Drive
            if self.subir_drive:
                try:
                    self.subir_drive()
                    logger.info("[REGISTROS] âœ… Subido a Drive")
                except Exception as e:
                    logger.error(f"[REGISTROS] Error subiendo a Drive: {e}")
            
            return True
            
        except Exception as e:
            logger.error(f"[REGISTROS] Error actualizando Excel: {e}")
            return False
    
    def _actualizar_ubicacion_conductor(self, fila_excel: int, lugar_descarga: str, 
                                         hora_llegada: str = None) -> bool:
        """
        Actualiza la ubicaciÃ³n del conductor cuando completa una descarga.
        """
        try:
            from openpyxl import load_workbook
            
            if not Path(self.excel_path).exists():
                logger.error(f"[REGISTROS] Excel no encontrado: {self.excel_path}")
                return False
            
            hora_actual = datetime.now().strftime("%H:%M")
            
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            fila_openpyxl = fila_excel + 1
            
            if fila_openpyxl > ws.max_row:
                logger.error(f"[REGISTROS] Fila {fila_openpyxl} fuera de rango")
                return False
            
            # Col B - UBICACIÃ“N
            celda_ubicacion = ws.cell(row=fila_openpyxl, column=COLUMNAS_CONDUCTOR['ubicacion'])
            ubicacion_anterior = celda_ubicacion.value
            celda_ubicacion.value = lugar_descarga
            
            # Col C - H.LL (hora llegada)
            celda_hll = ws.cell(row=fila_openpyxl, column=COLUMNAS_CONDUCTOR['hora_llegada'])
            if hora_llegada:
                celda_hll.value = hora_llegada
            else:
                hora_llegada_descarga = ws.cell(row=fila_openpyxl, column=COLUMNAS_EXCEL['descarga_llegada']).value
                if hora_llegada_descarga:
                    celda_hll.value = hora_llegada_descarga
            
            # Col D - H.SA (hora salida = hora actual)
            celda_hsa = ws.cell(row=fila_openpyxl, column=COLUMNAS_CONDUCTOR['hora_salida'])
            celda_hsa.value = hora_actual
            
            wb.save(self.excel_path)
            wb.close()
            
            logger.info(f"[REGISTROS] ğŸ“ UbicaciÃ³n actualizada: Fila {fila_openpyxl}, "
                       f"'{ubicacion_anterior}' â†’ '{lugar_descarga}' "
                       f"(H.LL: {celda_hll.value}, H.SA: {hora_actual})")
            
            # Actualizar tambiÃ©n en BD
            self._actualizar_ubicacion_bd(fila_excel, lugar_descarga)
            
            return True
            
        except Exception as e:
            logger.error(f"[REGISTROS] Error actualizando ubicaciÃ³n: {e}")
            return False
    
    def _actualizar_ubicacion_bd(self, fila_excel: int, nueva_ubicacion: str):
        """Actualiza la ubicaciÃ³n del conductor en la BD"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT conductor_asignado FROM viajes_empresa 
                WHERE fila_excel = ?
            """, (fila_excel,))
            row = cursor.fetchone()
            
            if row and row[0]:
                nombre_conductor = row[0]
                cursor.execute("""
                    UPDATE conductores_empresa 
                    SET ubicacion = ?
                    WHERE nombre LIKE ?
                """, (nueva_ubicacion, f"%{nombre_conductor}%"))
                conn.commit()
                logger.info(f"[REGISTROS] BD actualizada: {nombre_conductor} â†’ {nueva_ubicacion}")
            
            conn.close()
            
        except Exception as e:
            logger.error(f"[REGISTROS] Error actualizando BD: {e}")
    
    def _verificar_registro_existente(self, fila_excel: int, tipo: str, accion: str) -> Optional[str]:
        """
        Verifica si ya existe un registro en la celda.
        """
        try:
            from openpyxl import load_workbook
            
            if not Path(self.excel_path).exists():
                return None
            
            clave = f"{tipo}_{accion}"
            columna = COLUMNAS_EXCEL.get(clave)
            
            if not columna:
                return None
            
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            fila_openpyxl = fila_excel + 1
            
            if fila_openpyxl > ws.max_row:
                wb.close()
                return None
            
            celda = ws.cell(row=fila_openpyxl, column=columna)
            valor = celda.value
            wb.close()
            
            if valor and str(valor).strip():
                return str(valor).strip()
            
            return None
            
        except Exception as e:
            logger.error(f"[REGISTROS] Error verificando registro: {e}")
            return None
    
    def _marcar_viaje_completado(self, viaje_id: int):
        """Marca un viaje como completado en la BD"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE viajes_empresa 
                SET estado = 'completado'
                WHERE id = ?
            """, (viaje_id,))
            conn.commit()
            conn.close()
            logger.info(f"[REGISTROS] Viaje {viaje_id} marcado como completado")
        except Exception as e:
            logger.error(f"[REGISTROS] Error marcando viaje completado: {e}")
    
    # ============================================================
    # HANDLERS DEL CONVERSATION
    # ============================================================
    
    async def inicio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Inicia el flujo de registros"""
        telegram_id = update.effective_user.id
        
        # Verificar que tiene viaje
        viaje = self._obtener_viaje_conductor(telegram_id)
        
        if not viaje:
            from teclados import teclado_conductor
            await update.message.reply_text(
                "âŒ No tienes ningÃºn viaje asignado actualmente.\n\n"
                "Cuando tengas un viaje, podrÃ¡s registrar tus llegadas y salidas.",
                reply_markup=teclado_conductor
            )
            return ConversationHandler.END
        
        # BUGFIX #3: Viaje completado - mensaje diferente
        if viaje.get('_completado'):
            from teclados import teclado_conductor
            cliente = viaje.get('cliente', 'N/A')
            lugar_descarga = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
            await update.message.reply_text(
                f"âœ… *VIAJE COMPLETADO*\n\n"
                f"Tu Ãºltimo viaje ({cliente}) a {lugar_descarga} ya estÃ¡ completado.\n\n"
                f"Espera a que te asignen un nuevo viaje.",
                parse_mode="Markdown",
                reply_markup=teclado_conductor
            )
            return ConversationHandler.END
        
        # Guardar viaje en contexto
        context.user_data['viaje'] = viaje
        
        # Mostrar info del viaje
        lugar_carga = viaje.get('lugar_carga', 'N/A')
        lugar_descarga = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
        cliente = viaje.get('cliente', 'N/A')
        fila_excel = viaje.get('fila_excel')
        
        # Verificar registros existentes
        estado_registros = ""
        if fila_excel:
            carga_llegada = self._verificar_registro_existente(fila_excel, 'carga', 'llegada')
            carga_salida = self._verificar_registro_existente(fila_excel, 'carga', 'salida')
            descarga_llegada = self._verificar_registro_existente(fila_excel, 'descarga', 'llegada')
            descarga_salida = self._verificar_registro_existente(fila_excel, 'descarga', 'salida')
            
            estado_registros = "\nğŸ“‹ *Estado registros:*\n"
            estado_registros += f"ğŸ“¥ Carga: "
            estado_registros += f"{'âœ…' if carga_llegada else 'â¬œ'} Llegada {carga_llegada or ''} "
            estado_registros += f"{'âœ…' if carga_salida else 'â¬œ'} Salida {carga_salida or ''}\n"
            estado_registros += f"ğŸ“¤ Descarga: "
            estado_registros += f"{'âœ…' if descarga_llegada else 'â¬œ'} Llegada {descarga_llegada or ''} "
            estado_registros += f"{'âœ…' if descarga_salida else 'â¬œ'} Salida {descarga_salida or ''}\n"
        
        texto = (
            f"ğŸ“ *REGISTRAR HORA*\n\n"
            f"ğŸ¢ Cliente: *{cliente}*\n"
            f"ğŸ“¥ Carga: {lugar_carga}\n"
            f"ğŸ“¤ Descarga: {lugar_descarga}\n"
            f"{estado_registros}\n"
            f"Â¿QuÃ© quieres registrar?"
        )
        
        keyboard = [
            [
                InlineKeyboardButton("ğŸ“¥ Carga", callback_data="reg_tipo_carga"),
                InlineKeyboardButton("ğŸ“¤ Descarga", callback_data="reg_tipo_descarga"),
            ],
            [InlineKeyboardButton("âŒ Cancelar", callback_data="reg_cancelar")]
        ]
        
        await update.message.reply_text(
            texto,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return REG_TIPO
    
    async def seleccionar_tipo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cuando selecciona Carga o Descarga"""
        query = update.callback_query
        await query.answer()
        
        tipo = query.data.replace("reg_tipo_", "")
        context.user_data['tipo_registro'] = tipo
        
        viaje = context.user_data.get('viaje', {})
        
        if tipo == 'carga':
            lugar = viaje.get('lugar_carga', 'N/A')
            emoji = "ğŸ“¥"
        else:
            lugar = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
            emoji = "ğŸ“¤"
        
        texto = (
            f"{emoji} *{tipo.upper()}*\n\n"
            f"ğŸ“ Lugar: *{lugar}*\n\n"
            f"Â¿QuÃ© momento quieres registrar?"
        )
        
        keyboard = [
            [
                InlineKeyboardButton("ğŸš› Llegada", callback_data="reg_accion_llegada"),
                InlineKeyboardButton("ğŸ Salida", callback_data="reg_accion_salida"),
            ],
            [
                InlineKeyboardButton("â¬…ï¸ Volver", callback_data="reg_volver_tipo"),
                InlineKeyboardButton("âŒ Cancelar", callback_data="reg_cancelar"),
            ]
        ]
        
        await query.edit_message_text(
            texto,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return REG_ACCION
    
    async def seleccionar_accion(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cuando selecciona Llegada o Salida"""
        query = update.callback_query
        await query.answer()
        
        accion = query.data.replace("reg_accion_", "")
        tipo = context.user_data.get('tipo_registro', 'carga')
        viaje = context.user_data.get('viaje', {})
        fila_excel = viaje.get('fila_excel')
        
        # VERIFICAR SI YA EXISTE REGISTRO
        if fila_excel:
            hora_existente = self._verificar_registro_existente(fila_excel, tipo, accion)
            
            if hora_existente:
                emoji_tipo = "ğŸ“¥" if tipo == 'carga' else "ğŸ“¤"
                emoji_accion = "ğŸš›" if accion == 'llegada' else "ğŸ"
                
                if tipo == 'carga':
                    lugar = viaje.get('lugar_carga', 'N/A')
                else:
                    lugar = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
                
                texto = (
                    f"âš ï¸ *YA REGISTRADO*\n\n"
                    f"{emoji_tipo} {tipo.capitalize()}: *{lugar}*\n"
                    f"{emoji_accion} {accion.capitalize()}: *{hora_existente}*\n\n"
                    f"_Este registro ya fue realizado._\n"
                    f"_No se puede modificar._"
                )
                
                keyboard = [
                    [
                        InlineKeyboardButton("â¬…ï¸ Volver", callback_data="reg_volver_tipo"),
                        InlineKeyboardButton("âŒ Cancelar", callback_data="reg_cancelar"),
                    ]
                ]
                
                await query.edit_message_text(
                    texto,
                    parse_mode="Markdown",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
                
                return REG_TIPO  # Volver a selecciÃ³n de tipo
        
        # No existe registro, continuar normalmente
        context.user_data['accion_registro'] = accion
        
        hora_actual = datetime.now().strftime("%H:%M")
        
        if tipo == 'carga':
            lugar = viaje.get('lugar_carga', 'N/A')
            emoji_tipo = "ğŸ“¥"
        else:
            lugar = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
            emoji_tipo = "ğŸ“¤"
        
        emoji_accion = "ğŸš›" if accion == 'llegada' else "ğŸ"
        
        texto = (
            f"â° *CONFIRMAR REGISTRO*\n\n"
            f"{emoji_tipo} {tipo.capitalize()}: *{lugar}*\n"
            f"{emoji_accion} AcciÃ³n: *{accion.upper()}*\n"
            f"ğŸ• Hora: *{hora_actual}*\n\n"
            f"Â¿Confirmar registro?"
        )
        
        keyboard = [
            [InlineKeyboardButton(f"âœ… Registrar {accion}", callback_data=f"reg_confirmar_si")],
            [
                InlineKeyboardButton("â¬…ï¸ Volver", callback_data="reg_volver_accion"),
                InlineKeyboardButton("âŒ Cancelar", callback_data="reg_cancelar"),
            ]
        ]
        
        await query.edit_message_text(
            texto,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return REG_CONFIRMAR
    
    async def confirmar_registro(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Confirma y guarda el registro"""
        query = update.callback_query
        await query.answer()
        
        tipo = context.user_data.get('tipo_registro', 'carga')
        accion = context.user_data.get('accion_registro', 'llegada')
        viaje = context.user_data.get('viaje', {})
        fila_excel = viaje.get('fila_excel')
        
        hora_actual = datetime.now().strftime("%H:%M")
        
        if not fila_excel:
            await query.edit_message_text(
                "âŒ Error: No se encontrÃ³ la fila del viaje en el Excel.\n"
                "Contacta con el administrador."
            )
            context.user_data.clear()
            return ConversationHandler.END
        
        # Actualizar Excel (hora de carga/descarga)
        exito = self._actualizar_hora_excel(fila_excel, tipo, accion)
        
        # Si es SALIDA de DESCARGA â†’ actualizar ubicaciÃ³n del conductor y marcar completado
        ubicacion_actualizada = False
        if tipo == 'descarga' and accion == 'salida' and exito:
            lugar_descarga = viaje.get('lugar_entrega', viaje.get('lugar_descarga', ''))
            if lugar_descarga:
                ubicacion_actualizada = self._actualizar_ubicacion_conductor(
                    fila_excel, 
                    lugar_descarga
                )
                # Marcar viaje como completado
                viaje_id = viaje.get('id')
                if viaje_id:
                    self._marcar_viaje_completado(viaje_id)
                
                # Subir a Drive despuÃ©s de actualizar ubicaciÃ³n
                if self.subir_drive and ubicacion_actualizada:
                    try:
                        self.subir_drive()
                    except Exception as e:
                        logger.error(f"[REGISTROS] Error subiendo a Drive: {e}")
        
        if tipo == 'carga':
            lugar = viaje.get('lugar_carga', 'N/A')
            emoji_tipo = "ğŸ“¥"
        else:
            lugar = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
            emoji_tipo = "ğŸ“¤"
        
        emoji_accion = "ğŸš›" if accion == 'llegada' else "ğŸ"
        
        if exito:
            texto = (
                f"âœ… *REGISTRO GUARDADO*\n\n"
                f"{emoji_tipo} {tipo.capitalize()}: *{lugar}*\n"
                f"{emoji_accion} {accion.capitalize()}: *{hora_actual}*\n\n"
            )
            
            # AÃ±adir info de ubicaciÃ³n si se actualizÃ³
            if ubicacion_actualizada:
                texto += f"ğŸ“ _Tu ubicaciÃ³n se ha actualizado a:_ *{lugar}*\n\n"
            
            texto += "â˜ï¸ _Sincronizado con Drive_"
        else:
            texto = (
                f"âš ï¸ *ERROR AL GUARDAR*\n\n"
                f"No se pudo registrar la hora.\n"
                f"IntÃ©ntalo de nuevo o contacta con el administrador."
            )
        
        from teclados import teclado_conductor
        
        await query.edit_message_text(texto, parse_mode="Markdown")
        await query.message.reply_text(
            "Â¿QuÃ© mÃ¡s necesitas?",
            reply_markup=teclado_conductor
        )
        
        context.user_data.clear()
        return ConversationHandler.END
    
    # ============================================================
    # NAVEGACIÃ“N
    # ============================================================
    
    async def volver_tipo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve a selecciÃ³n de tipo"""
        query = update.callback_query
        await query.answer()
        
        viaje = context.user_data.get('viaje', {})
        lugar_carga = viaje.get('lugar_carga', 'N/A')
        lugar_descarga = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
        cliente = viaje.get('cliente', 'N/A')
        fila_excel = viaje.get('fila_excel')
        
        # Verificar registros existentes
        estado_registros = ""
        if fila_excel:
            carga_llegada = self._verificar_registro_existente(fila_excel, 'carga', 'llegada')
            carga_salida = self._verificar_registro_existente(fila_excel, 'carga', 'salida')
            descarga_llegada = self._verificar_registro_existente(fila_excel, 'descarga', 'llegada')
            descarga_salida = self._verificar_registro_existente(fila_excel, 'descarga', 'salida')
            
            estado_registros = "\nğŸ“‹ *Estado registros:*\n"
            estado_registros += f"ğŸ“¥ Carga: "
            estado_registros += f"{'âœ…' if carga_llegada else 'â¬œ'} Llegada {carga_llegada or ''} "
            estado_registros += f"{'âœ…' if carga_salida else 'â¬œ'} Salida {carga_salida or ''}\n"
            estado_registros += f"ğŸ“¤ Descarga: "
            estado_registros += f"{'âœ…' if descarga_llegada else 'â¬œ'} Llegada {descarga_llegada or ''} "
            estado_registros += f"{'âœ…' if descarga_salida else 'â¬œ'} Salida {descarga_salida or ''}\n"
        
        texto = (
            f"ğŸ“ *REGISTRAR HORA*\n\n"
            f"ğŸ¢ Cliente: *{cliente}*\n"
            f"ğŸ“¥ Carga: {lugar_carga}\n"
            f"ğŸ“¤ Descarga: {lugar_descarga}\n"
            f"{estado_registros}\n"
            f"Â¿QuÃ© quieres registrar?"
        )
        
        keyboard = [
            [
                InlineKeyboardButton("ğŸ“¥ Carga", callback_data="reg_tipo_carga"),
                InlineKeyboardButton("ğŸ“¤ Descarga", callback_data="reg_tipo_descarga"),
            ],
            [InlineKeyboardButton("âŒ Cancelar", callback_data="reg_cancelar")]
        ]
        
        await query.edit_message_text(
            texto,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return REG_TIPO
    
    async def volver_accion(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve a selecciÃ³n de acciÃ³n"""
        query = update.callback_query
        await query.answer()
        
        tipo = context.user_data.get('tipo_registro', 'carga')
        viaje = context.user_data.get('viaje', {})
        
        if tipo == 'carga':
            lugar = viaje.get('lugar_carga', 'N/A')
            emoji = "ğŸ“¥"
        else:
            lugar = viaje.get('lugar_entrega', viaje.get('lugar_descarga', 'N/A'))
            emoji = "ğŸ“¤"
        
        texto = (
            f"{emoji} *{tipo.upper()}*\n\n"
            f"ğŸ“ Lugar: *{lugar}*\n\n"
            f"Â¿QuÃ© momento quieres registrar?"
        )
        
        keyboard = [
            [
                InlineKeyboardButton("ğŸš› Llegada", callback_data="reg_accion_llegada"),
                InlineKeyboardButton("ğŸ Salida", callback_data="reg_accion_salida"),
            ],
            [
                InlineKeyboardButton("â¬…ï¸ Volver", callback_data="reg_volver_tipo"),
                InlineKeyboardButton("âŒ Cancelar", callback_data="reg_cancelar"),
            ]
        ]
        
        await query.edit_message_text(
            texto,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return REG_ACCION
    
    async def cancelar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela el registro (mensaje)"""
        context.user_data.clear()
        
        from teclados import teclado_conductor
        
        await update.message.reply_text(
            "âŒ Registro cancelado.",
            reply_markup=teclado_conductor
        )
        
        return ConversationHandler.END
    
    async def cancelar_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela el registro (callback)"""
        query = update.callback_query
        await query.answer()
        
        context.user_data.clear()
        
        from teclados import teclado_conductor
        
        await query.edit_message_text("âŒ Registro cancelado.")
        await query.message.reply_text(
            "Â¿QuÃ© mÃ¡s necesitas?",
            reply_markup=teclado_conductor
        )
        
        return ConversationHandler.END


# ============================================================
# FUNCIÃ“N PARA INTEGRAR EN BOT_TRANSPORTE.PY
# ============================================================

def crear_registros_conductor(excel_path: str, db_path: str, subir_drive_func=None):
    """
    Crea una instancia del mÃ³dulo de registros.
    """
    return RegistrosConductor(
        excel_path=excel_path,
        db_path=db_path,
        subir_drive_func=subir_drive_func
    )
