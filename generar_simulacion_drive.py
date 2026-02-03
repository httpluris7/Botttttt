"""
GENERADOR DE SIMULACIÓN - CON GOOGLE DRIVE
============================================
Descarga Excel de Drive → Añade datos → Sube a Drive

USO:
    python generar_simulacion_drive.py

REQUISITOS:
    - token.json (credenciales de Drive)
    - credentials.json
    - .env con DRIVE_EXCEL_EMPRESA_ID
"""

import openpyxl
import openpyxl.cell.cell
from openpyxl.comments import Comment
import random
import json
import os
import io
from pathlib import Path
from dotenv import load_dotenv

# Google Drive
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

load_dotenv()

# ============================================================
# CONFIGURACIÓN
# ============================================================

EXCEL_LOCAL = "PRUEBO_SIMULACION.xlsx"  # Archivo temporal local
DRIVE_FILE_ID = os.getenv("DRIVE_EXCEL_EMPRESA_ID")
SCOPES = ['https://www.googleapis.com/auth/drive']

# Datos de simulación
UBICACIONES_BASE = ["AZAGRA", "TUDELA", "CALAHORRA", "SAN ADRIAN", "LOGROÑO", "PAMPLONA", "ALFARO", "ARNEDO", "ESTELLA", "TAFALLA"]
CLIENTES = ["HERO", "NESTLE", "CAMPOFRIO", "CONSUM", "GRUPO AN", "EROSKI", "MERCADONA", "DIA", "CARREFOUR", "DANONE", "PASCUAL", "FLORETTE", "BONDUELLE", "COVIRAN", "ALDI", "LIDL"]
LUGARES_CARGA = ["CALAHORRA", "TUDELA", "ALFARO", "SAN ADRIAN", "AZAGRA", "PERALTA", "MENDAVIA", "LODOSA", "ARNEDO", "AUTOL", "QUEL"]
LUGARES_DESCARGA = ["MADRID", "BARCELONA", "VALENCIA", "SEVILLA", "ZARAGOZA", "BILBAO", "MALAGA", "MURCIA", "ALICANTE", "VALLADOLID", "VIGO", "GIJON", "VITORIA", "GRANADA", "OVIEDO", "SANTANDER"]
MERCANCIAS = ["REFRIGERADO +2º", "REFRIGERADO +5º", "CONGELADO -18º", "CONGELADO -25º", "SECO"]
NOMBRES = ["ANTONIO", "MANUEL", "JOSE", "FRANCISCO", "DAVID", "JUAN", "CARLOS", "JESUS", "JAVIER", "DANIEL", "MIGUEL", "RAFAEL", "PEDRO", "PABLO", "ANGEL", "SERGIO", "FERNANDO", "JORGE", "LUIS", "ALBERTO", "ALEJANDRO", "DIEGO", "ADRIAN", "RAUL", "IVAN", "RUBEN", "OSCAR", "RAMON", "VICENTE", "ENRIQUE"]
APELLIDOS = ["GARCIA", "MARTINEZ", "LOPEZ", "SANCHEZ", "GONZALEZ", "RODRIGUEZ", "FERNANDEZ", "PEREZ", "GOMEZ", "MARTIN", "JIMENEZ", "RUIZ", "HERNANDEZ", "DIAZ", "MORENO", "ALVAREZ", "MUÑOZ", "ROMERO", "ALONSO", "GUTIERREZ", "NAVARRO", "TORRES", "DOMINGUEZ", "VAZQUEZ", "RAMOS", "GIL", "SERRANO", "BLANCO", "MOLINA", "MORALES"]

# ============================================================
# GOOGLE DRIVE
# ============================================================

def inicializar_drive():
    """Inicializa conexión con Google Drive"""
    creds = None
    
    if Path('token.json').exists():
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not Path('credentials.json').exists():
                print("❌ No existe credentials.json")
                return None
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    
    return build('drive', 'v3', credentials=creds)

def descargar_excel(drive_service, file_id, local_path):
    """Descarga Excel de Drive"""
    print(f"📥 Descargando Excel de Drive...")
    
    request = drive_service.files().get_media(fileId=file_id)
    
    with open(local_path, 'wb') as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
    
    print(f"   ✅ Descargado: {local_path}")
    return True

def subir_excel(drive_service, file_id, local_path):
    """Sube Excel a Drive"""
    print(f"📤 Subiendo Excel a Drive...")
    
    media = MediaFileUpload(
        local_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        resumable=True
    )
    
    drive_service.files().update(
        fileId=file_id,
        media_body=media
    ).execute()
    
    print(f"   ✅ Subido a Drive")
    return True

# ============================================================
# GENERADORES DE DATOS
# ============================================================

def generar_telefono():
    prefijos = ["666", "677", "688", "699", "622", "633", "644", "655"]
    return f"{random.choice(prefijos)}{random.randint(100000, 999999)}"

def generar_matricula_tractora():
    letras = "BCDFGHJKLMNPRSTVWXYZ"
    return f"{random.randint(1000, 9999)}{random.choice(letras)}{random.choice(letras)}{random.choice(letras)}"

def generar_matricula_remolque():
    letras = "BCDFGHJKLMNPRSTVWXYZ"
    return f"R{random.randint(1000, 9999)}{random.choice(letras)}{random.choice(letras)}{random.choice(letras)}"

def calcular_km(destino):
    distancias = {
        "MADRID": 320, "BARCELONA": 400, "VALENCIA": 450, "SEVILLA": 700,
        "ZARAGOZA": 150, "BILBAO": 150, "MALAGA": 750, "MURCIA": 550,
        "ALICANTE": 500, "VALLADOLID": 200, "VIGO": 500, "GIJON": 350,
        "VITORIA": 100, "GRANADA": 650, "OVIEDO": 350, "SANTANDER": 200
    }
    return distancias.get(destino, 300) + random.randint(-50, 50)

def calcular_precio(km, mercancia):
    if "CONGELADO" in mercancia:
        precio_km = random.uniform(1.4, 1.7)
    elif "REFRIGERADO" in mercancia:
        precio_km = random.uniform(1.2, 1.5)
    else:
        precio_km = random.uniform(1.0, 1.3)
    return round((km * precio_km) / 25) * 25

def generar_camioneros(cantidad=40):
    """Genera lista de camioneros únicos"""
    camioneros = []
    nombres_usados = set()
    
    for _ in range(cantidad):
        nombre = f"{random.choice(NOMBRES)} {random.choice(APELLIDOS)}"
        intentos = 0
        while nombre in nombres_usados and intentos < 100:
            nombre = f"{random.choice(NOMBRES)} {random.choice(APELLIDOS)}"
            intentos += 1
        nombres_usados.add(nombre)
        
        camioneros.append({
            'nombre': nombre,
            'telefono': generar_telefono(),
            'tractora': generar_matricula_tractora(),
            'remolque': generar_matricula_remolque(),
            'ubicacion': random.choice(UBICACIONES_BASE),
        })
    
    return camioneros

def generar_viajes(cantidad, camioneros):
    """Genera lista de viajes"""
    viajes = []
    
    for _ in range(cantidad):
        lugar_carga = random.choice(LUGARES_CARGA)
        lugar_descarga = random.choice(LUGARES_DESCARGA)
        mercancia = random.choice(MERCANCIAS)
        km = calcular_km(lugar_descarga)
        precio = calcular_precio(km, mercancia)
        
        # 60% asignados
        if random.random() < 0.6:
            cam = random.choice(camioneros)
            transportista = cam['nombre']
            tractora = cam['tractora']
            remolque = cam['remolque']
            ubicacion = cam['ubicacion']
            telefono = cam['telefono']
        else:
            transportista = ""
            tractora = ""
            remolque = ""
            ubicacion = ""
            telefono = ""
        
        # 20% con carga adicional
        carga_adicional = random.choice([l for l in LUGARES_CARGA if l != lugar_carga]) if random.random() < 0.2 else None
        
        # 15% con descarga adicional
        descarga_adicional = random.choice([l for l in LUGARES_DESCARGA if l != lugar_descarga]) if random.random() < 0.15 else None
        
        viajes.append({
            'ubicacion': ubicacion,
            'transportista': transportista,
            'telefono': telefono,
            'tractora': tractora,
            'remolque': remolque,
            'cliente': random.choice(CLIENTES),
            'num_pedido': random.randint(10000, 99999) if random.random() > 0.3 else None,
            'ref_cliente': f"REF-{random.randint(1000, 9999)}" if random.random() > 0.5 else None,
            'intercambio': random.choice(["SI", "NO", "NO", "NO"]),
            'lugar_carga': lugar_carga,
            'carga_adicional': carga_adicional,
            'lugar_descarga': lugar_descarga,
            'descarga_adicional': descarga_adicional,
            'mercancia': mercancia,
            'precio': precio,
            'km': km,
        })
    
    return viajes

# ============================================================
# ESCRIBIR EN EXCEL
# ============================================================

def escribir_excel(excel_path, camioneros, viajes):
    """Escribe los datos en el Excel"""
    print(f"\n📝 Modificando Excel...")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Encontrar primera fila vacía (empezando en 3)
    fila_inicio = 3
    while fila_inicio < 200:
        cliente = ws.cell(row=fila_inicio, column=9).value
        transportista = ws.cell(row=fila_inicio, column=5).value
        
        # Verificar si es celda combinada (header)
        cell = ws.cell(row=fila_inicio, column=9)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            fila_inicio += 1
            continue
            
        if not cliente and not transportista:
            break
        fila_inicio += 1
    
    print(f"   Comenzando en fila {fila_inicio}")
    
    fila = fila_inicio
    escritos = 0
    
    for v in viajes:
        try:
            # Verificar que no sea celda combinada
            cell_test = ws.cell(row=fila, column=9)
            if isinstance(cell_test, openpyxl.cell.cell.MergedCell):
                print(f"   ⚠️ Fila {fila} tiene celdas combinadas, saltando...")
                fila += 1
                continue
            
            # Columna 2: UBICACIÓN
            try:
                ws.cell(row=fila, column=2, value=v['ubicacion'])
            except:
                pass
            
            # Columna 5: TRANSPORTISTA + comentario con teléfono
            if v['transportista']:
                ws.cell(row=fila, column=5, value=v['transportista'])
                if v['telefono']:
                    comentario = Comment(f"Tel. empresa: {v['telefono']}", "Bot")
                    ws.cell(row=fila, column=5).comment = comentario
            
            # Columna 7: TRACTORA
            ws.cell(row=fila, column=7, value=v['tractora'])
            
            # Columna 8: REMOLQUE
            ws.cell(row=fila, column=8, value=v['remolque'])
            
            # Columna 9: CLIENTE
            ws.cell(row=fila, column=9, value=v['cliente'])
            
            # Columna 10: Nº PEDIDO
            if v['num_pedido']:
                ws.cell(row=fila, column=10, value=v['num_pedido'])
            
            # Columna 11: REF CLIENTE
            if v['ref_cliente']:
                ws.cell(row=fila, column=11, value=v['ref_cliente'])
            
            # Columna 12: INTERCAMBIO
            ws.cell(row=fila, column=12, value=v['intercambio'])
            
            # Columna 14: LUGAR DE CARGA + nota si hay adicional
            ws.cell(row=fila, column=14, value=v['lugar_carga'])
            if v['carga_adicional']:
                nota = Comment(f"2 CARGAS: {v['lugar_carga']} + {v['carga_adicional']}", "Bot")
                ws.cell(row=fila, column=14).comment = nota
            
            # Columna 17: LUGAR DE DESCARGA + nota si hay adicional
            ws.cell(row=fila, column=17, value=v['lugar_descarga'])
            if v['descarga_adicional']:
                nota = Comment(f"2 DESCARGAS: {v['lugar_descarga']} + {v['descarga_adicional']}", "Bot")
                ws.cell(row=fila, column=17).comment = nota
            
            # Columna 20: MERCANCÍA
            ws.cell(row=fila, column=20, value=v['mercancia'])
            
            # Columna 23: PRECIO
            ws.cell(row=fila, column=23, value=v['precio'])
            
            # Columna 24: KM
            ws.cell(row=fila, column=24, value=v['km'])
            
            # Columna 25: EUR/KM (fórmula)
            ws.cell(row=fila, column=25, value=f"=W{fila}/X{fila}")
            
            # Columna 28: OBSERVACIONES
            obs = "ZONA: ZONA NORTE"
            if v['carga_adicional']:
                obs += f" | CARGA2: {v['carga_adicional']}"
            if v['descarga_adicional']:
                obs += f" | DESCARGA2: {v['descarga_adicional']}"
            ws.cell(row=fila, column=28, value=obs)
            
            fila += 1
            escritos += 1
            
        except Exception as e:
            print(f"   ⚠️ Error fila {fila}: {e}")
            fila += 1
    
    wb.save(excel_path)
    wb.close()
    
    print(f"   ✅ {escritos} viajes escritos (filas {fila_inicio} a {fila - 1})")
    return escritos

# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("🚀 SIMULACIÓN ZONA NORTE - CON GOOGLE DRIVE")
    print("=" * 60)
    
    if not DRIVE_FILE_ID:
        print("\n❌ No se encontró DRIVE_EXCEL_EMPRESA_ID en .env")
        exit(1)
    
    print(f"\n📁 Drive File ID: {DRIVE_FILE_ID[:20]}...")
    
    # Inicializar Drive
    print("\n🔗 Conectando con Google Drive...")
    drive_service = inicializar_drive()
    if not drive_service:
        print("❌ No se pudo conectar con Drive")
        exit(1)
    print("   ✅ Conectado")
    
    # Descargar Excel
    descargar_excel(drive_service, DRIVE_FILE_ID, EXCEL_LOCAL)
    
    # Generar datos
    print("\n🔄 Generando datos de simulación...")
    camioneros = generar_camioneros(40)
    viajes = generar_viajes(40, camioneros)
    
    print(f"   - 40 camioneros generados")
    print(f"   - 40 viajes generados")
    
    asignados = sum(1 for v in viajes if v['transportista'])
    print(f"   - {asignados} viajes asignados")
    print(f"   - {40 - asignados} viajes pendientes")
    
    # Escribir en Excel
    escritos = escribir_excel(EXCEL_LOCAL, camioneros, viajes)
    
    if escritos > 0:
        # Subir a Drive
        subir_excel(drive_service, DRIVE_FILE_ID, EXCEL_LOCAL)
        
        print("\n" + "=" * 60)
        print("✅ SIMULACIÓN COMPLETADA")
        print("=" * 60)
        print(f"\n📊 Resumen:")
        print(f"   - {escritos} viajes añadidos")
        print(f"   - 40 camioneros con teléfonos")
        print(f"\n🔄 Ahora en Telegram pulsa '🔄 Sincronizar'")
    else:
        print("\n❌ No se escribieron datos")
    
    # Limpiar archivo temporal
    if Path(EXCEL_LOCAL).exists():
        os.remove(EXCEL_LOCAL)
