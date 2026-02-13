"""
MODIFICAR VIAJES EN RUTA (v2.1 - FIX CANCELAR)
========================================================
Sistema para que los admins modifiquen viajes de conductores en ruta.
Ahora soporta hasta 10 cargas y 10 descargas por viaje.

CAMBIOS v2.1:
- FIX: A√±adido CallbackQueryHandler para bot√≥n "‚ùå Cancelar" en InlineKeyboards
- FIX: A√±adido handler cancelar en estados de cargas/descargas
- Antes: El bot se quedaba "enganchado" al pulsar cancelar en InlineKeyboard

Flujo:
1. Admin selecciona "Modificar viaje en ruta"
2. Selecciona zona
3. Ve lista de conductores EN RUTA con: nombre, tel√©fono, ruta, cliente
4. Selecciona conductor
5. Ve detalles del viaje con TODAS las cargas/descargas
6. Puede modificar campos individuales o gestionar cargas/descargas
7. Confirma cambios
8. Se actualiza Excel + Drive + notifica al conductor
"""

import sqlite3
import logging
import openpyxl
import re
from openpyxl.comments import Comment
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from pathlib import Path

from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes,
    ConversationHandler,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters
)

logger = logging.getLogger(__name__)

# ============================================================
# CONSTANTES
# ============================================================
MAX_CARGAS = 10
MAX_DESCARGAS = 10

# ============================================================
# ESTADOS DEL CONVERSATION HANDLER
# ============================================================
MOD_RUTA_ZONA = 100
MOD_RUTA_CONDUCTOR = 101
MOD_RUTA_DETALLE = 102
MOD_RUTA_CAMPO = 103
MOD_RUTA_VALOR = 104
MOD_RUTA_CONFIRMAR = 105

# Estados para gesti√≥n de cargas/descargas en ruta
MOD_RUTA_CARGAS_MENU = 110
MOD_RUTA_CARGAS_EDITAR = 111
MOD_RUTA_CARGAS_ELIMINAR = 112
MOD_RUTA_DESCARGAS_MENU = 113
MOD_RUTA_DESCARGAS_EDITAR = 114
MOD_RUTA_DESCARGAS_ELIMINAR = 115

# Zonas disponibles (seg√∫n Excel PRUEBO.xlsx)
ZONAS = ["ZONA NORTE", "ZONA CORTOS NORTE", "ZONA RESTO NACIONAL", "ZONA MURCIA"]


# ============================================================
# HELPERS MULTI-CARGA/DESCARGA
# ============================================================

def _extraer_cargas_de_obs(observaciones: str) -> list:
    """Extrae cargas adicionales (CARGA2..CARGA10) de observaciones"""
    extra = []
    if not observaciones:
        return extra
    for i in range(2, MAX_CARGAS + 1):
        match = re.search(rf'CARGA{i}:\s*([^|]+)', observaciones)
        if match:
            extra.append(match.group(1).strip())
    return extra


def _extraer_descargas_de_obs(observaciones: str) -> list:
    """Extrae descargas adicionales (DESCARGA2..DESCARGA10) de observaciones"""
    extra = []
    if not observaciones:
        return extra
    for i in range(2, MAX_DESCARGAS + 1):
        match = re.search(rf'DESCARGA{i}:\s*([^|]+)', observaciones)
        if match:
            extra.append(match.group(1).strip())
    return extra


def _inicializar_cargas_conductor(conductor: dict):
    """Asegura que conductor tenga listas de cargas/descargas"""
    if 'cargas' not in conductor:
        cargas = []
        if conductor.get('lugar_carga'):
            cargas.append(conductor['lugar_carga'])
        obs = conductor.get('observaciones', '') or ''
        cargas.extend(_extraer_cargas_de_obs(obs))
        conductor['cargas'] = cargas
    
    if 'descargas' not in conductor:
        descargas = []
        lugar = conductor.get('lugar_entrega') or conductor.get('lugar_descarga', '')
        if lugar:
            descargas.append(lugar)
        obs = conductor.get('observaciones', '') or ''
        descargas.extend(_extraer_descargas_de_obs(obs))
        conductor['descargas'] = descargas


def _sync_compat_conductor(conductor: dict):
    """Sincroniza listas con campos legacy"""
    cargas = conductor.get('cargas', [])
    conductor['lugar_carga'] = cargas[0] if cargas else ''
    
    descargas = conductor.get('descargas', [])
    conductor['lugar_entrega'] = descargas[0] if descargas else ''


def _generar_observaciones_ruta(conductor: dict) -> str:
    """Genera string de observaciones con todas las cargas/descargas"""
    partes = []
    
    # Mantener zona si existe
    obs_original = conductor.get('observaciones', '') or ''
    zona_match = re.search(r'ZONA:\s*([^|]+)', obs_original)
    if zona_match:
        partes.append(f"ZONA: {zona_match.group(1).strip()}")
    
    cargas = conductor.get('cargas', [])
    for i, c in enumerate(cargas[1:], 2):
        partes.append(f"CARGA{i}: {c}")
    
    descargas = conductor.get('descargas', [])
    for i, d in enumerate(descargas[1:], 2):
        partes.append(f"DESCARGA{i}: {d}")
    
    # Mantener partes de observaciones originales que no sean CARGA/DESCARGA/ZONA
    if obs_original:
        for parte in obs_original.split('|'):
            parte = parte.strip()
            if parte and not re.match(r'(ZONA|CARGA\d+|DESCARGA\d+):', parte):
                partes.append(parte)
    
    return " | ".join(partes)


def _generar_comentario_cargas(cargas: list) -> str:
    if len(cargas) <= 1:
        return ""
    return f"{len(cargas)} CARGAS: " + " + ".join(cargas)


def _generar_comentario_descargas(descargas: list) -> str:
    if len(descargas) <= 1:
        return ""
    return f"{len(descargas)} DESCARGAS: " + " + ".join(descargas)


# ============================================================
# CLASE PRINCIPAL
# ============================================================

class ModificadorViajesRuta:
    """
    Gestiona la modificaci√≥n de viajes de conductores en ruta.
    Incluye notificaciones y sincronizaci√≥n con Drive.
    Soporta hasta 10 cargas y 10 descargas por viaje.
    """
    
    def __init__(self, excel_path: str, db_path: str, es_admin_func, 
                 subir_drive_func=None, bot=None, movildata_api=None):
        self.excel_path = excel_path
        self.db_path = db_path
        self.es_admin = es_admin_func
        self.subir_drive = subir_drive_func
        self.bot = bot
        self.movildata = movildata_api
        
        # Campos modificables del viaje (sin cargas/descargas, que tienen su submen√∫)
        self.campos_viaje = {
            '1': ('cargas', 'üìç Cargas'),
            '2': ('descargas', 'üìç Descargas'),
            '3': ('mercancia', 'üì¶ Mercanc√≠a'),
            '4': ('observaciones', 'üìù Observaciones'),
            '5': ('hora_carga', '‚è∞ Hora de Carga'),
            '6': ('hora_descarga', '‚è∞ Hora de Descarga'),
            '7': ('cliente', 'üè¢ Cliente'),
            '8': ('precio', 'üí∞ Precio'),
            '9': ('km', 'üìè Kil√≥metros'),
        }
        
        logger.info("[MOD_RUTA] Modificador de viajes en ruta v2.1 inicializado")
    
    def get_conversation_handler(self):
        """Devuelve el ConversationHandler para modificar viajes en ruta"""
        return ConversationHandler(
            entry_points=[
                MessageHandler(filters.Regex("^üîÑ Modificar viaje en ruta$"), self.inicio),
                CommandHandler("modificar_ruta", self.inicio),
            ],
            states={
                MOD_RUTA_ZONA: [
                    CallbackQueryHandler(self.seleccionar_zona, pattern="^zona_"),
                    CallbackQueryHandler(self.cancelar, pattern="^cancelar$"),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                ],
                MOD_RUTA_CONDUCTOR: [
                    CallbackQueryHandler(self.seleccionar_conductor, pattern="^conductor_"),
                    CallbackQueryHandler(self.volver_zonas, pattern="^volver_zonas$"),
                    CallbackQueryHandler(self.cancelar, pattern="^cancelar$"),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                ],
                MOD_RUTA_DETALLE: [
                    CallbackQueryHandler(self.seleccionar_campo, pattern="^campo_"),
                    CallbackQueryHandler(self.llamar_conductor, pattern="^llamar_"),
                    CallbackQueryHandler(self.volver_conductores, pattern="^volver_conductores$"),
                    CallbackQueryHandler(self.confirmar_cambios, pattern="^confirmar_si$"),
                    CallbackQueryHandler(self.cancelar, pattern="^cancelar$"),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                ],
                MOD_RUTA_VALOR: [
                    MessageHandler(filters.Regex("^‚¨ÖÔ∏è Volver$"), self.volver_detalle),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.guardar_valor),
                ],
                MOD_RUTA_CONFIRMAR: [
                    CallbackQueryHandler(self.confirmar_cambios, pattern="^confirmar_si$"),
                    CallbackQueryHandler(self.volver_detalle, pattern="^confirmar_no$"),
                    CallbackQueryHandler(self.cancelar, pattern="^cancelar$"),
                ],
                # === GESTI√ìN CARGAS EN RUTA ===
                MOD_RUTA_CARGAS_MENU: [
                    MessageHandler(filters.Regex("^‚ûï A√±adir carga$"), self._ruta_cargas_a√±adir),
                    MessageHandler(filters.Regex("^üóëÔ∏è Eliminar carga$"), self._ruta_cargas_pedir_eliminar),
                    MessageHandler(filters.Regex("^‚¨ÖÔ∏è Volver$"), self._ruta_cargas_volver),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._ruta_cargas_seleccionar),
                ],
                MOD_RUTA_CARGAS_EDITAR: [
                    MessageHandler(filters.Regex("^‚¨ÖÔ∏è Volver$"), self._ruta_mostrar_cargas_menu),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._ruta_cargas_guardar_edicion),
                ],
                MOD_RUTA_CARGAS_ELIMINAR: [
                    MessageHandler(filters.Regex("^‚¨ÖÔ∏è Volver$"), self._ruta_mostrar_cargas_menu),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._ruta_cargas_confirmar_eliminar),
                ],
                # === GESTI√ìN DESCARGAS EN RUTA ===
                MOD_RUTA_DESCARGAS_MENU: [
                    MessageHandler(filters.Regex("^‚ûï A√±adir descarga$"), self._ruta_descargas_a√±adir),
                    MessageHandler(filters.Regex("^üóëÔ∏è Eliminar descarga$"), self._ruta_descargas_pedir_eliminar),
                    MessageHandler(filters.Regex("^‚¨ÖÔ∏è Volver$"), self._ruta_descargas_volver),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._ruta_descargas_seleccionar),
                ],
                MOD_RUTA_DESCARGAS_EDITAR: [
                    MessageHandler(filters.Regex("^‚¨ÖÔ∏è Volver$"), self._ruta_mostrar_descargas_menu),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._ruta_descargas_guardar_edicion),
                ],
                MOD_RUTA_DESCARGAS_ELIMINAR: [
                    MessageHandler(filters.Regex("^‚¨ÖÔ∏è Volver$"), self._ruta_mostrar_descargas_menu),
                    MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._ruta_descargas_confirmar_eliminar),
                ],
            },
            fallbacks=[
                CommandHandler("cancelar", self.cancelar),
                MessageHandler(filters.Regex("^‚ùå Cancelar$"), self.cancelar),
                CallbackQueryHandler(self.cancelar, pattern="^cancelar$"),
            ],
        )

    # ============================================================
    # FUNCIONES DE BASE DE DATOS
    # ============================================================
    
    def _query(self, query: str, params: tuple = (), fetch_one: bool = False):
        """Ejecuta una consulta SQL"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute(query, params)
                if fetch_one:
                    row = cursor.fetchone()
                    return dict(row) if row else None
                return [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            logger.error(f"[MOD_RUTA] Error SQL: {e}")
            return None
    
    def _obtener_conductores_en_ruta(self, zona: str = None) -> List[Dict]:
        """Obtiene conductores que est√°n EN RUTA (con viaje activo)."""
        query = """
            SELECT DISTINCT 
                c.nombre,
                c.telefono,
                c.tractora,
                c.remolque,
                c.zona,
                c.telegram_id,
                v.id as viaje_id,
                v.cliente,
                v.lugar_carga,
                v.lugar_entrega,
                v.mercancia,
                v.precio,
                v.km,
                v.observaciones,
                v.fila_excel,
                v.estado
            FROM conductores_empresa c
            INNER JOIN viajes_empresa v ON v.conductor_asignado LIKE '%' || c.nombre || '%'
            WHERE v.estado IN ('pendiente', 'en_ruta', 'asignado')
        """
        
        params = ()
        if zona:
            # Normalizar zona (quitar espacios extras)
            zona_normalizada = zona.strip()
            query += " AND (TRIM(c.zona) LIKE ? OR TRIM(v.zona) LIKE ?)"
            params = (f"%{zona_normalizada}%", f"%{zona_normalizada}%")
        
        query += " ORDER BY c.nombre"
        
        conductores = self._query(query, params)
        
        # A√±adir estado GPS si est√° disponible
        if self.movildata and conductores:
            for c in conductores:
                if c.get('tractora'):
                    try:
                        estado = self.movildata.get_estado_vehiculo(c['tractora'])
                        if estado:
                            c['estado_gps'] = estado.get('estado', 'desconocido')
                            c['velocidad'] = estado.get('velocidad', 0)
                    except:
                        pass
        
        # Inicializar cargas/descargas para cada conductor
        if conductores:
            for c in conductores:
                _inicializar_cargas_conductor(c)
        
        return conductores or []
    
    def _obtener_telegram_id_conductor(self, nombre: str) -> Optional[int]:
        """Obtiene el telegram_id de un conductor"""
        result = self._query(
            "SELECT telegram_id FROM conductores_empresa WHERE nombre LIKE ?",
            (f"%{nombre}%",),
            fetch_one=True
        )
        return result.get('telegram_id') if result else None
    
    # ============================================================
    # FUNCIONES DE EXCEL
    # ============================================================
    
    def _detectar_formato_columna(self, ws, columna: int) -> dict:
        """Detecta el formato de una columna bas√°ndose en las primeras filas."""
        formato = {
            'number_format': 'General',
            'bold': False,
            'font_name': 'Calibri',
            'font_size': 11,
            'tipo_valor': 'str',
        }
        
        for fila in range(3, 8):
            celda = ws.cell(row=fila, column=columna)
            if celda.value is not None:
                formato['number_format'] = celda.number_format
                if celda.font:
                    formato['bold'] = celda.font.bold or False
                    formato['font_name'] = celda.font.name
                    formato['font_size'] = celda.font.size
                
                if isinstance(celda.value, int):
                    formato['tipo_valor'] = 'int'
                elif isinstance(celda.value, float):
                    formato['tipo_valor'] = 'float'
                else:
                    formato['tipo_valor'] = 'str'
                break
        
        return formato
    
    def _aplicar_formato(self, valor, campo, formato):
        """Convierte valor al tipo correcto seg√∫n formato detectado."""
        if campo in ['precio', 'km']:
            try:
                valor_limpio = str(valor).replace('‚Ç¨', '').replace(' ', '').replace(',', '.').replace('km', '').replace('KM', '')
                if formato['tipo_valor'] == 'int':
                    return int(float(valor_limpio))
                elif formato['tipo_valor'] == 'float':
                    return float(valor_limpio)
                else:
                    return float(valor_limpio)
            except:
                return valor
        elif campo in ['lugar_carga', 'lugar_entrega', 'cliente', 'mercancia']:
            return str(valor).strip().upper()
        else:
            return valor

    def _actualizar_excel(self, fila: int, campo: str, valor) -> bool:
        """Actualiza un campo en el Excel copiando formato."""
        from openpyxl.styles import Font
        
        COLUMNAS_EXCEL = {
            'lugar_carga': 14,
            'lugar_entrega': 17,
            'mercancia': 20,
            'precio': 23,
            'km': 24,
            'observaciones': 28,
            'cliente': 9,
            'hora_carga': 3,
            'hora_descarga': 4,
        }
        
        columna = COLUMNAS_EXCEL.get(campo)
        if not columna:
            logger.error(f"[MOD_RUTA] Campo no encontrado: {campo}")
            return False
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            fila_real = fila + 1  # Corregir desfase BD
            
            formato = self._detectar_formato_columna(ws, columna)
            valor_convertido = self._aplicar_formato(valor, campo, formato)
            
            celda = ws.cell(row=fila_real, column=columna)
            valor_anterior = celda.value
            
            celda.value = valor_convertido
            celda.number_format = formato['number_format']
            celda.font = Font(
                bold=formato['bold'],
                name=formato['font_name'] or 'Calibri',
                size=formato['font_size'] or 11
            )
            
            comentario = f"Modificado: {datetime.now().strftime('%d/%m/%Y %H:%M')}\nAnterior: {valor_anterior}"
            celda.comment = Comment(comentario, "Bot")
            
            wb.save(self.excel_path)
            wb.close()
            
            logger.info(f"[MOD_RUTA] Excel actualizado: fila {fila_real}, {campo} = {valor_convertido}")
            return True
            
        except Exception as e:
            logger.error(f"[MOD_RUTA] Error actualizando Excel: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def _actualizar_excel_cargas_descargas(self, fila: int, conductor: dict) -> bool:
        """Actualiza cargas, descargas y observaciones completas en Excel."""
        from openpyxl.styles import Font
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            fila_real = fila + 1
            
            cargas = conductor.get('cargas', [])
            descargas = conductor.get('descargas', [])
            
            # Columna 14: Carga principal + comentario
            if cargas:
                ws.cell(row=fila_real, column=14, value=cargas[0])
                ws.cell(row=fila_real, column=14).comment = None
                if len(cargas) > 1:
                    texto = _generar_comentario_cargas(cargas)
                    ws.cell(row=fila_real, column=14).comment = Comment(texto, "Bot")
            
            # Columna 17: Descarga principal + comentario
            if descargas:
                ws.cell(row=fila_real, column=17, value=descargas[0])
                ws.cell(row=fila_real, column=17).comment = None
                if len(descargas) > 1:
                    texto = _generar_comentario_descargas(descargas)
                    ws.cell(row=fila_real, column=17).comment = Comment(texto, "Bot")
            
            # Columna 28: Observaciones con todas las cargas/descargas
            observaciones = _generar_observaciones_ruta(conductor)
            ws.cell(row=fila_real, column=28, value=observaciones)
            
            wb.save(self.excel_path)
            wb.close()
            
            logger.info(f"[MOD_RUTA] Excel cargas/descargas actualizadas: fila {fila_real}, {len(cargas)}C/{len(descargas)}D")
            return True
            
        except Exception as e:
            logger.error(f"[MOD_RUTA] Error actualizando cargas/descargas Excel: {e}")
            return False
    
    def _actualizar_bd(self, viaje_id: int, campo: str, valor) -> bool:
        """Actualiza un campo en la BD"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    f"UPDATE viajes_empresa SET {campo} = ? WHERE id = ?",
                    (valor, viaje_id)
                )
                conn.commit()
                return True
        except Exception as e:
            logger.error(f"[MOD_RUTA] Error actualizando BD: {e}")
            return False
    
    def _sync_to_drive(self) -> bool:
        """Sincroniza el Excel con Google Drive"""
        if self.subir_drive:
            try:
                result = self.subir_drive()
                if result:
                    logger.info("[MOD_RUTA] ‚úÖ Excel sincronizado con Drive")
                return result
            except Exception as e:
                logger.error(f"[MOD_RUTA] Error sincronizando: {e}")
                return False
        else:
            logger.warning("[MOD_RUTA] ‚ö†Ô∏è subir_drive es None, no se sincroniza")
        return True
    
    # ============================================================
    # NOTIFICACIONES
    # ============================================================
    
    async def _notificar_conductor(self, telegram_id: int, mensaje: str) -> bool:
        """Env√≠a notificaci√≥n al conductor"""
        if not self.bot or not telegram_id:
            return False
        try:
            await self.bot.send_message(
                chat_id=telegram_id,
                text=mensaje,
                parse_mode='Markdown'
            )
            logger.info(f"[MOD_RUTA] Notificaci√≥n enviada a {telegram_id}")
            return True
        except Exception as e:
            logger.error(f"[MOD_RUTA] Error enviando notificaci√≥n: {e}")
            return False
    
    def _generar_mensaje_modificacion(self, conductor: str, cambios: Dict) -> str:
        """Genera el mensaje de notificaci√≥n para el conductor"""
        mensaje = f"‚ö†Ô∏è *CAMBIO EN TU VIAJE*\n\n"
        mensaje += f"üë§ Conductor: {conductor}\n"
        mensaje += f"üìÖ Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
        mensaje += "üìù *Cambios realizados:*\n"
        
        for campo, valores in cambios.items():
            if campo in ('cargas', 'descargas'):
                emoji = "üìç"
                nombre = "Cargas" if campo == 'cargas' else "Descargas"
                mensaje += f"\n{emoji} {nombre}:\n"
                mensaje += f"   ‚úÖ Ahora: {', '.join(valores.get('nuevo', []))}\n"
            else:
                nombre_campo = self.campos_viaje.get(campo, (campo, campo))
                if isinstance(nombre_campo, tuple):
                    nombre_campo = nombre_campo[1]
                mensaje += f"\n{nombre_campo}:\n"
                mensaje += f"   ‚ùå Antes: {valores['anterior']}\n"
                mensaje += f"   ‚úÖ Ahora: {valores['nuevo']}\n"
        
        mensaje += "\n_Contacta con oficina si tienes dudas._"
        return mensaje
    
    # ============================================================
    # HANDLERS DEL CONVERSATION
    # ============================================================
    
    async def inicio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Inicio: muestra selecci√≥n de zonas"""
        user = update.effective_user
        
        if not self.es_admin(user.id):
            await update.message.reply_text("‚ùå Solo administradores pueden acceder.")
            return ConversationHandler.END
        
        context.user_data.clear()
        
        keyboard = []
        for zona in ZONAS:
            keyboard.append([InlineKeyboardButton(zona, callback_data=f"zona_{zona}")])
        keyboard.append([InlineKeyboardButton("‚ùå Cancelar", callback_data="cancelar")])
        
        await update.message.reply_text(
            "üîÑ *MODIFICAR VIAJE EN RUTA*\n\n"
            "Selecciona la zona:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MOD_RUTA_ZONA
    
    async def seleccionar_zona(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handler cuando admin selecciona una zona"""
        query = update.callback_query
        await query.answer()
        
        zona = query.data.replace("zona_", "")
        context.user_data['zona_seleccionada'] = zona
        
        conductores = self._obtener_conductores_en_ruta(zona)
        
        if not conductores:
            await query.edit_message_text(
                f"‚ùå No hay conductores en ruta en {zona}.\n\n"
                "Selecciona otra zona:",
                reply_markup=self._get_keyboard_zonas()
            )
            return MOD_RUTA_ZONA
        
        context.user_data['conductores'] = conductores
        
        texto = f"üöõ *CONDUCTORES EN RUTA - {zona}*\n\n"
        keyboard = []
        
        for i, c in enumerate(conductores):
            nombre = c.get('nombre', 'N/A')
            telefono = c.get('telefono', 'Sin tel√©fono')
            cliente = c.get('cliente', 'N/A')
            carga = c.get('lugar_carga', '?')
            descarga = c.get('lugar_entrega', '?')
            estado_gps = c.get('estado_gps', '')
            n_cargas = len(c.get('cargas', []))
            n_descargas = len(c.get('descargas', []))
            
            estado_emoji = "üü¢" if estado_gps == 'en_ruta' else "üîµ"
            
            texto += f"{estado_emoji} *{nombre}*\n"
            texto += f"   üìû {telefono}\n"
            texto += f"   üöõ {carga} ‚Üí {descarga}"
            if n_cargas > 1 or n_descargas > 1:
                texto += f" ({n_cargas}C/{n_descargas}D)"
            texto += f"\n   üè¢ {cliente}\n\n"
            
            nombre_corto = nombre.split()[0] if nombre else f"Conductor {i+1}"
            keyboard.append([
                InlineKeyboardButton(
                    f"{estado_emoji} {nombre_corto} - {carga}‚Üí{descarga}",
                    callback_data=f"conductor_{i}"
                )
            ])
        
        keyboard.append([InlineKeyboardButton("‚¨ÖÔ∏è Volver", callback_data="volver_zonas")])
        keyboard.append([InlineKeyboardButton("‚ùå Cancelar", callback_data="cancelar")])
        
        await query.edit_message_text(
            texto,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MOD_RUTA_CONDUCTOR
    
    async def seleccionar_conductor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handler cuando admin selecciona un conductor"""
        query = update.callback_query
        await query.answer()
        
        indice = int(query.data.replace("conductor_", ""))
        conductores = context.user_data.get('conductores', [])
        
        if indice >= len(conductores):
            await query.edit_message_text("‚ùå Error: conductor no encontrado.")
            return ConversationHandler.END
        
        conductor = conductores[indice]
        _inicializar_cargas_conductor(conductor)
        context.user_data['conductor_seleccionado'] = conductor
        context.user_data['cambios_pendientes'] = {}
        
        texto = self._formatear_detalle_viaje(conductor, context.user_data.get('cambios_pendientes', {}))
        keyboard = self._get_keyboard_campos(conductor)
        
        await query.edit_message_text(
            texto,
            parse_mode="Markdown",
            reply_markup=keyboard
        )
        return MOD_RUTA_DETALLE
    
    def _formatear_detalle_viaje(self, conductor: Dict, cambios: Dict = None) -> str:
        """Formatea los detalles del viaje con todas las cargas/descargas"""
        if cambios is None:
            cambios = {}
        
        texto = "üìã *DETALLE DEL VIAJE*\n"
        texto += "‚ïê" * 25 + "\n\n"
        
        texto += f"üë§ *Conductor:* {conductor.get('nombre', 'N/A')}\n"
        texto += f"üìû *Tel√©fono:* {conductor.get('telefono', 'Sin tel√©fono')}\n"
        texto += f"üöõ *Tractora:* {conductor.get('tractora', 'N/A')}\n\n"
        
        texto += "‚îÄ" * 25 + "\n"
        texto += "*VIAJE ACTUAL*\n"
        texto += "‚îÄ" * 25 + "\n\n"
        
        texto += f"üè¢ *Cliente:* {conductor.get('cliente', 'N/A')}\n"
        texto += f"üì¶ *Mercanc√≠a:* {conductor.get('mercancia', 'N/A')}\n\n"
        
        # Cargas (todas)
        cargas = conductor.get('cargas', [])
        texto += f"üì• *Cargas ({len(cargas)}):*\n"
        for i, c in enumerate(cargas):
            etiqueta = "Principal" if i == 0 else f"#{i+1}"
            texto += f"   {etiqueta}: {c}\n"
        if not cargas:
            texto += "   _Sin cargas_\n"
        texto += "\n"
        
        # Descargas (todas)
        descargas = conductor.get('descargas', [])
        texto += f"üì§ *Descargas ({len(descargas)}):*\n"
        for i, d in enumerate(descargas):
            etiqueta = "Principal" if i == 0 else f"#{i+1}"
            texto += f"   {etiqueta}: {d}\n"
        if not descargas:
            texto += "   _Sin descargas_\n"
        texto += "\n"
        
        texto += f"üìè *Km:* {conductor.get('km', 'N/A')}\n"
        texto += f"üí∞ *Precio:* {conductor.get('precio', 'N/A')}‚Ç¨\n"
        
        if conductor.get('observaciones'):
            obs = conductor.get('observaciones', '')
            # Limpiar c√≥digos internos
            obs_limpia = obs
            for n in range(2, MAX_CARGAS + 1):
                obs_limpia = re.sub(rf'\s*\|?\s*CARGA{n}:[^|]+', '', obs_limpia)
            for n in range(2, MAX_DESCARGAS + 1):
                obs_limpia = re.sub(rf'\s*\|?\s*DESCARGA{n}:[^|]+', '', obs_limpia)
            obs_limpia = re.sub(r'\s*\|?\s*ZONA:[^|]+', '', obs_limpia).strip()
            if obs_limpia:
                texto += f"\nüìù *Obs:* {obs_limpia[:100]}\n"
        
        # Cambios pendientes
        if cambios:
            texto += "\n" + "‚ïê" * 25 + "\n"
            texto += "‚ö†Ô∏è *CAMBIOS PENDIENTES:*\n"
            for campo_key, valores in cambios.items():
                if campo_key in ('cargas', 'descargas'):
                    emoji = "üìç"
                    nombre = "Cargas" if campo_key == 'cargas' else "Descargas"
                    nuevo = valores.get('nuevo', [])
                    texto += f"‚Ä¢ {emoji} {nombre}: {', '.join(nuevo)}\n"
                else:
                    texto += f"‚Ä¢ {valores['nombre']}: `{valores['anterior']}` ‚Üí `{valores['nuevo']}`\n"
        
        texto += "\n_Selecciona un campo para modificar o confirma:_"
        return texto
    
    def _get_keyboard_campos(self, conductor: Dict) -> InlineKeyboardMarkup:
        """Genera el teclado con campos modificables"""
        cargas = conductor.get('cargas', [])
        descargas = conductor.get('descargas', [])
        
        keyboard = [
            [
                InlineKeyboardButton(f"üì• Cargas ({len(cargas)})", callback_data="campo_1"),
                InlineKeyboardButton(f"üì§ Descargas ({len(descargas)})", callback_data="campo_2"),
            ],
            [
                InlineKeyboardButton("üì¶ Mercanc√≠a", callback_data="campo_3"),
                InlineKeyboardButton("üìù Observaciones", callback_data="campo_4"),
            ],
            [
                InlineKeyboardButton("üè¢ Cliente", callback_data="campo_7"),
                InlineKeyboardButton("üí∞ Precio", callback_data="campo_8"),
            ],
            [
                InlineKeyboardButton("üìè Km", callback_data="campo_9"),
            ],
            [
                InlineKeyboardButton(
                    f"üìû Llamar ({conductor.get('telefono', 'N/A')})",
                    callback_data=f"llamar_{conductor.get('telefono', '')}"
                ),
            ],
            [
                InlineKeyboardButton("‚úÖ Confirmar cambios", callback_data="confirmar_si"),
                InlineKeyboardButton("‚¨ÖÔ∏è Volver", callback_data="volver_conductores"),
            ],
        ]
        
        return InlineKeyboardMarkup(keyboard)
    
    def _get_keyboard_zonas(self) -> InlineKeyboardMarkup:
        keyboard = []
        for zona in ZONAS:
            keyboard.append([InlineKeyboardButton(zona, callback_data=f"zona_{zona}")])
        keyboard.append([InlineKeyboardButton("‚ùå Cancelar", callback_data="cancelar")])
        return InlineKeyboardMarkup(keyboard)
    
    async def seleccionar_campo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handler cuando admin selecciona un campo para modificar"""
        query = update.callback_query
        await query.answer()
        
        num_campo = query.data.replace("campo_", "")
        campo_info = self.campos_viaje.get(num_campo)
        
        if not campo_info:
            await query.answer("Campo no v√°lido", show_alert=True)
            return MOD_RUTA_DETALLE
        
        campo_key, campo_nombre = campo_info
        
        # Si es cargas o descargas, abrir submen√∫ de gesti√≥n
        if campo_key == 'cargas':
            return await self._ruta_mostrar_cargas_menu(update, context)
        elif campo_key == 'descargas':
            return await self._ruta_mostrar_descargas_menu(update, context)
        
        context.user_data['campo_editando'] = num_campo
        context.user_data['campo_key'] = campo_key
        context.user_data['campo_nombre'] = campo_nombre
        
        conductor = context.user_data.get('conductor_seleccionado', {})
        valor_actual = conductor.get(campo_key, 'N/A')
        
        keyboard = [["‚¨ÖÔ∏è Volver", "‚ùå Cancelar"]]
        
        await query.message.reply_text(
            f"‚úèÔ∏è *MODIFICAR {campo_nombre.upper()}*\n\n"
            f"Valor actual: `{valor_actual}`\n\n"
            "Escribe el nuevo valor:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MOD_RUTA_VALOR
    
    async def guardar_valor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda el nuevo valor del campo"""
        nuevo_valor = update.message.text.strip()
        
        campo_key = context.user_data.get('campo_key')
        campo_nombre = context.user_data.get('campo_nombre')
        conductor = context.user_data.get('conductor_seleccionado', {})
        valor_anterior = conductor.get(campo_key, '')
        
        if 'cambios_pendientes' not in context.user_data:
            context.user_data['cambios_pendientes'] = {}
        
        context.user_data['cambios_pendientes'][campo_key] = {
            'anterior': valor_anterior,
            'nuevo': nuevo_valor,
            'nombre': campo_nombre
        }
        
        conductor[campo_key] = nuevo_valor
        
        await update.message.reply_text(
            f"‚úÖ {campo_nombre} actualizado a: `{nuevo_valor}`\n\n"
            "Puedes modificar m√°s campos o confirmar los cambios.",
            parse_mode="Markdown"
        )
        
        # Volver al detalle
        cambios = context.user_data.get('cambios_pendientes', {})
        texto = self._formatear_detalle_viaje(conductor, cambios)
        keyboard = self._get_keyboard_campos(conductor)
        
        await update.message.reply_text(
            texto,
            parse_mode="Markdown",
            reply_markup=keyboard
        )
        return MOD_RUTA_DETALLE
    
    async def llamar_conductor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra tel√©fono del conductor"""
        query = update.callback_query
        telefono = query.data.replace("llamar_", "")
        
        if not telefono or telefono == "None":
            await query.answer("‚ùå No hay tel√©fono registrado", show_alert=True)
            return MOD_RUTA_DETALLE
        
        conductor = context.user_data.get('conductor_seleccionado', {})
        nombre = conductor.get('nombre', 'Conductor')
        
        await query.answer(f"üìû {telefono}", show_alert=True)
        
        await query.message.reply_text(
            f"üìû *LLAMAR A {nombre}*\n\n"
            f"Tel√©fono: `{telefono}`\n\n"
            f"[Llamar](tel:{telefono})",
            parse_mode="Markdown"
        )
        return MOD_RUTA_DETALLE
    
    # ============================================================
    # GESTI√ìN DE CARGAS EN RUTA
    # ============================================================
    
    async def _ruta_mostrar_cargas_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra men√∫ de gesti√≥n de cargas"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        _inicializar_cargas_conductor(conductor)
        cargas = conductor.get('cargas', [])
        n = len(cargas)
        
        mensaje = f"üì• *GESTI√ìN DE CARGAS* ({n}/{MAX_CARGAS})\n‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ\n"
        if cargas:
            for i, c in enumerate(cargas):
                etiqueta = "Principal" if i == 0 else f"#{i+1}"
                mensaje += f"*{i+1}.* üì• {etiqueta}: *{c}*\n"
        else:
            mensaje += "_Sin cargas definidas_\n"
        
        mensaje += "‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ\n\n"
        mensaje += "Escribe un *n√∫mero* para editar esa carga"
        
        botones = []
        if n < MAX_CARGAS:
            botones.append("‚ûï A√±adir carga")
        if n > 1:
            botones.append("üóëÔ∏è Eliminar carga")
        
        keyboard = []
        if botones:
            keyboard.append(botones)
        keyboard.append(["‚¨ÖÔ∏è Volver"])
        
        # Responder seg√∫n si viene de CallbackQuery o Message
        if update.callback_query:
            await update.callback_query.message.reply_text(
                mensaje, parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
        else:
            await update.message.reply_text(
                mensaje, parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
        return MOD_RUTA_CARGAS_MENU
    
    async def _ruta_cargas_seleccionar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Selecciona carga a editar por n√∫mero"""
        texto = update.message.text.strip()
        conductor = context.user_data.get('conductor_seleccionado', {})
        cargas = conductor.get('cargas', [])
        
        try:
            num = int(texto)
            if num < 1 or num > len(cargas):
                await update.message.reply_text(f"‚ö†Ô∏è N√∫mero del 1 al {len(cargas)}:")
                return MOD_RUTA_CARGAS_MENU
            
            context.user_data['_editando_carga_idx'] = num - 1
            keyboard = [["‚¨ÖÔ∏è Volver", "‚ùå Cancelar"]]
            await update.message.reply_text(
                f"‚úèÔ∏è *EDITAR CARGA #{num}*\n\n"
                f"Valor actual: *{cargas[num-1]}*\n\n"
                "Escribe el nuevo lugar de carga:",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return MOD_RUTA_CARGAS_EDITAR
        except ValueError:
            await update.message.reply_text("‚ö†Ô∏è Introduce un n√∫mero v√°lido:")
            return MOD_RUTA_CARGAS_MENU
    
    async def _ruta_cargas_guardar_edicion(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda edici√≥n de carga"""
        nuevo_valor = update.message.text.strip().upper()
        idx = context.user_data.get('_editando_carga_idx', 0)
        conductor = context.user_data.get('conductor_seleccionado', {})
        
        if idx == -1:
            # Nueva carga
            conductor.setdefault('cargas', []).append(nuevo_valor)
            await update.message.reply_text(f"‚úÖ Carga *{nuevo_valor}* a√±adida.", parse_mode="Markdown")
        else:
            conductor['cargas'][idx] = nuevo_valor
            await update.message.reply_text(f"‚úÖ Carga #{idx+1} actualizada a *{nuevo_valor}*", parse_mode="Markdown")
        
        _sync_compat_conductor(conductor)
        
        # Marcar cambio pendiente
        if 'cambios_pendientes' not in context.user_data:
            context.user_data['cambios_pendientes'] = {}
        context.user_data['cambios_pendientes']['cargas'] = {
            'nombre': 'üìç Cargas',
            'nuevo': conductor.get('cargas', [])
        }
        
        return await self._ruta_mostrar_cargas_menu(update, context)
    
    async def _ruta_cargas_a√±adir(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """A√±adir nueva carga"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        cargas = conductor.get('cargas', [])
        if len(cargas) >= MAX_CARGAS:
            await update.message.reply_text(f"‚ö†Ô∏è M√°ximo de {MAX_CARGAS} cargas alcanzado.")
            return MOD_RUTA_CARGAS_MENU
        
        n = len(cargas)
        keyboard = [["‚¨ÖÔ∏è Volver", "‚ùå Cancelar"]]
        await update.message.reply_text(
            f"üì• *NUEVA CARGA #{n+1}*\n\nEscribe el lugar de carga:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        context.user_data['_editando_carga_idx'] = -1
        return MOD_RUTA_CARGAS_EDITAR
    
    async def _ruta_cargas_pedir_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide qu√© carga eliminar"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        cargas = conductor.get('cargas', [])
        if len(cargas) <= 1:
            await update.message.reply_text("‚ö†Ô∏è Debe haber al menos 1 carga.")
            return MOD_RUTA_CARGAS_MENU
        
        mensaje = f"üóëÔ∏è *¬øQu√© carga eliminar?* (2-{len(cargas)})\n\n"
        mensaje += "_Nota: la carga #1 (principal) no se puede eliminar_\n\n"
        for i, c in enumerate(cargas):
            mensaje += f"*{i+1}.* {c}\n"
        
        keyboard = [["‚¨ÖÔ∏è Volver"]]
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MOD_RUTA_CARGAS_ELIMINAR
    
    async def _ruta_cargas_confirmar_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Confirma eliminaci√≥n de carga"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        cargas = conductor.get('cargas', [])
        try:
            num = int(update.message.text.strip())
            if num < 2 or num > len(cargas):
                await update.message.reply_text(f"‚ö†Ô∏è N√∫mero del 2 al {len(cargas)}:")
                return MOD_RUTA_CARGAS_ELIMINAR
            
            eliminada = cargas.pop(num - 1)
            _sync_compat_conductor(conductor)
            
            if 'cambios_pendientes' not in context.user_data:
                context.user_data['cambios_pendientes'] = {}
            context.user_data['cambios_pendientes']['cargas'] = {
                'nombre': 'üìç Cargas',
                'nuevo': conductor.get('cargas', [])
            }
            
            await update.message.reply_text(f"‚úÖ Carga *{eliminada}* eliminada.", parse_mode="Markdown")
            return await self._ruta_mostrar_cargas_menu(update, context)
        except ValueError:
            await update.message.reply_text("‚ö†Ô∏è Introduce un n√∫mero:")
            return MOD_RUTA_CARGAS_ELIMINAR
    
    async def _ruta_cargas_volver(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Volver al detalle desde cargas"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        cambios = context.user_data.get('cambios_pendientes', {})
        
        texto = self._formatear_detalle_viaje(conductor, cambios)
        keyboard = self._get_keyboard_campos(conductor)
        
        await update.message.reply_text(
            texto, parse_mode="Markdown",
            reply_markup=keyboard
        )
        return MOD_RUTA_DETALLE
    
    # ============================================================
    # GESTI√ìN DE DESCARGAS EN RUTA
    # ============================================================
    
    async def _ruta_mostrar_descargas_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra men√∫ de gesti√≥n de descargas"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        _inicializar_cargas_conductor(conductor)
        descargas = conductor.get('descargas', [])
        n = len(descargas)
        
        mensaje = f"üì§ *GESTI√ìN DE DESCARGAS* ({n}/{MAX_DESCARGAS})\n‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ\n"
        if descargas:
            for i, d in enumerate(descargas):
                etiqueta = "Principal" if i == 0 else f"#{i+1}"
                mensaje += f"*{i+1}.* üì§ {etiqueta}: *{d}*\n"
        else:
            mensaje += "_Sin descargas definidas_\n"
        
        mensaje += "‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ‚îÅ\n\n"
        mensaje += "Escribe un *n√∫mero* para editar esa descarga"
        
        botones = []
        if n < MAX_DESCARGAS:
            botones.append("‚ûï A√±adir descarga")
        if n > 1:
            botones.append("üóëÔ∏è Eliminar descarga")
        
        keyboard = []
        if botones:
            keyboard.append(botones)
        keyboard.append(["‚¨ÖÔ∏è Volver"])
        
        if update.callback_query:
            await update.callback_query.message.reply_text(
                mensaje, parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
        else:
            await update.message.reply_text(
                mensaje, parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
        return MOD_RUTA_DESCARGAS_MENU
    
    async def _ruta_descargas_seleccionar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Selecciona descarga a editar por n√∫mero"""
        texto = update.message.text.strip()
        conductor = context.user_data.get('conductor_seleccionado', {})
        descargas = conductor.get('descargas', [])
        
        try:
            num = int(texto)
            if num < 1 or num > len(descargas):
                await update.message.reply_text(f"‚ö†Ô∏è N√∫mero del 1 al {len(descargas)}:")
                return MOD_RUTA_DESCARGAS_MENU
            
            context.user_data['_editando_descarga_idx'] = num - 1
            keyboard = [["‚¨ÖÔ∏è Volver", "‚ùå Cancelar"]]
            await update.message.reply_text(
                f"‚úèÔ∏è *EDITAR DESCARGA #{num}*\n\n"
                f"Valor actual: *{descargas[num-1]}*\n\n"
                "Escribe el nuevo lugar de descarga:",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return MOD_RUTA_DESCARGAS_EDITAR
        except ValueError:
            await update.message.reply_text("‚ö†Ô∏è Introduce un n√∫mero v√°lido:")
            return MOD_RUTA_DESCARGAS_MENU
    
    async def _ruta_descargas_guardar_edicion(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda edici√≥n de descarga"""
        nuevo_valor = update.message.text.strip().upper()
        idx = context.user_data.get('_editando_descarga_idx', 0)
        conductor = context.user_data.get('conductor_seleccionado', {})
        
        if idx == -1:
            conductor.setdefault('descargas', []).append(nuevo_valor)
            await update.message.reply_text(f"‚úÖ Descarga *{nuevo_valor}* a√±adida.", parse_mode="Markdown")
        else:
            conductor['descargas'][idx] = nuevo_valor
            await update.message.reply_text(f"‚úÖ Descarga #{idx+1} actualizada a *{nuevo_valor}*", parse_mode="Markdown")
        
        _sync_compat_conductor(conductor)
        
        if 'cambios_pendientes' not in context.user_data:
            context.user_data['cambios_pendientes'] = {}
        context.user_data['cambios_pendientes']['descargas'] = {
            'nombre': 'üìç Descargas',
            'nuevo': conductor.get('descargas', [])
        }
        
        return await self._ruta_mostrar_descargas_menu(update, context)
    
    async def _ruta_descargas_a√±adir(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """A√±adir nueva descarga"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        descargas = conductor.get('descargas', [])
        if len(descargas) >= MAX_DESCARGAS:
            await update.message.reply_text(f"‚ö†Ô∏è M√°ximo de {MAX_DESCARGAS} descargas alcanzado.")
            return MOD_RUTA_DESCARGAS_MENU
        
        n = len(descargas)
        keyboard = [["‚¨ÖÔ∏è Volver", "‚ùå Cancelar"]]
        await update.message.reply_text(
            f"üì§ *NUEVA DESCARGA #{n+1}*\n\nEscribe el lugar de descarga:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        context.user_data['_editando_descarga_idx'] = -1
        return MOD_RUTA_DESCARGAS_EDITAR
    
    async def _ruta_descargas_pedir_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide qu√© descarga eliminar"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        descargas = conductor.get('descargas', [])
        if len(descargas) <= 1:
            await update.message.reply_text("‚ö†Ô∏è Debe haber al menos 1 descarga.")
            return MOD_RUTA_DESCARGAS_MENU
        
        mensaje = f"üóëÔ∏è *¬øQu√© descarga eliminar?* (2-{len(descargas)})\n\n"
        mensaje += "_Nota: la descarga #1 (principal) no se puede eliminar_\n\n"
        for i, d in enumerate(descargas):
            mensaje += f"*{i+1}.* {d}\n"
        
        keyboard = [["‚¨ÖÔ∏è Volver"]]
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MOD_RUTA_DESCARGAS_ELIMINAR
    
    async def _ruta_descargas_confirmar_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Confirma eliminaci√≥n de descarga"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        descargas = conductor.get('descargas', [])
        try:
            num = int(update.message.text.strip())
            if num < 2 or num > len(descargas):
                await update.message.reply_text(f"‚ö†Ô∏è N√∫mero del 2 al {len(descargas)}:")
                return MOD_RUTA_DESCARGAS_ELIMINAR
            
            eliminada = descargas.pop(num - 1)
            _sync_compat_conductor(conductor)
            
            if 'cambios_pendientes' not in context.user_data:
                context.user_data['cambios_pendientes'] = {}
            context.user_data['cambios_pendientes']['descargas'] = {
                'nombre': 'üìç Descargas',
                'nuevo': conductor.get('descargas', [])
            }
            
            await update.message.reply_text(f"‚úÖ Descarga *{eliminada}* eliminada.", parse_mode="Markdown")
            return await self._ruta_mostrar_descargas_menu(update, context)
        except ValueError:
            await update.message.reply_text("‚ö†Ô∏è Introduce un n√∫mero:")
            return MOD_RUTA_DESCARGAS_ELIMINAR
    
    async def _ruta_descargas_volver(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Volver al detalle desde descargas"""
        conductor = context.user_data.get('conductor_seleccionado', {})
        cambios = context.user_data.get('cambios_pendientes', {})
        
        texto = self._formatear_detalle_viaje(conductor, cambios)
        keyboard = self._get_keyboard_campos(conductor)
        
        await update.message.reply_text(
            texto, parse_mode="Markdown",
            reply_markup=keyboard
        )
        return MOD_RUTA_DETALLE
    
    # ============================================================
    # CONFIRMAR Y APLICAR CAMBIOS
    # ============================================================
    
    async def confirmar_cambios(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Confirma y aplica todos los cambios"""
        logger.info("üîî [MOD_RUTA] ENTRANDO EN confirmar_cambios")
        query = update.callback_query
        await query.answer()
        
        cambios = context.user_data.get('cambios_pendientes', {})
        conductor = context.user_data.get('conductor_seleccionado', {})
        
        if not cambios:
            await query.answer("No hay cambios pendientes", show_alert=True)
            return MOD_RUTA_DETALLE
        
        fila_excel = conductor.get('fila_excel')
        viaje_id = conductor.get('viaje_id')
        nombre_conductor = conductor.get('nombre', '')
        telegram_id = conductor.get('telegram_id') or self._obtener_telegram_id_conductor(nombre_conductor)
        
        errores = []
        hay_cambio_cargas = False
        
        for campo_key, valores in cambios.items():
            if campo_key in ('cargas', 'descargas'):
                hay_cambio_cargas = True
                continue  # Se procesan aparte
            
            # Actualizar Excel
            if fila_excel:
                if not self._actualizar_excel(fila_excel, campo_key, valores['nuevo']):
                    errores.append(f"Excel: {campo_key}")
            
            # Actualizar BD
            if viaje_id:
                if not self._actualizar_bd(viaje_id, campo_key, valores['nuevo']):
                    errores.append(f"BD: {campo_key}")
        
        # Si hubo cambios en cargas/descargas, actualizar Excel completo
        if hay_cambio_cargas and fila_excel:
            if not self._actualizar_excel_cargas_descargas(fila_excel, conductor):
                errores.append("Excel: cargas/descargas")
            
            # Actualizar campos legacy en BD
            if viaje_id:
                _sync_compat_conductor(conductor)
                self._actualizar_bd(viaje_id, 'lugar_carga', conductor.get('lugar_carga', ''))
                self._actualizar_bd(viaje_id, 'lugar_entrega', conductor.get('lugar_entrega', ''))
                obs = _generar_observaciones_ruta(conductor)
                self._actualizar_bd(viaje_id, 'observaciones', obs)
        
        # Sync Drive
        drive_ok = self._sync_to_drive()
        
        # Notificar al conductor
        notificado = False
        if telegram_id:
            mensaje_notif = self._generar_mensaje_modificacion(nombre_conductor, cambios)
            notificado = await self._notificar_conductor(telegram_id, mensaje_notif)
        
        # Mensaje al admin
        if errores:
            msg = f"‚ö†Ô∏è Cambios aplicados con errores:\n" + "\n".join(errores)
        else:
            n_cargas = len(conductor.get('cargas', []))
            n_descargas = len(conductor.get('descargas', []))
            
            msg = f"‚úÖ *¬°VIAJE MODIFICADO!*\n\n"
            msg += f"üë§ {nombre_conductor}\n"
            msg += f"üì• Cargas: {n_cargas} | üì§ Descargas: {n_descargas}\n"
            msg += f"üìù {len(cambios)} campo(s) modificado(s)\n\n"
            msg += "‚òÅÔ∏è _Drive sincronizado_\n" if drive_ok else "‚ö†Ô∏è _Error sincronizando Drive_\n"
            msg += "üì≤ _Conductor notificado_" if notificado else "‚ö†Ô∏è _No se pudo notificar_"
        
        from teclados import teclado_admin
        
        await query.message.reply_text(
            msg,
            parse_mode="Markdown",
            reply_markup=teclado_admin
        )
        
        context.user_data.clear()
        return ConversationHandler.END
    
    # ============================================================
    # NAVEGACI√ìN
    # ============================================================
    
    async def volver_zonas(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve a la selecci√≥n de zonas"""
        query = update.callback_query
        await query.answer()
        
        await query.edit_message_text(
            "üîÑ *MODIFICAR VIAJE EN RUTA*\n\nSelecciona la zona:",
            parse_mode="Markdown",
            reply_markup=self._get_keyboard_zonas()
        )
        return MOD_RUTA_ZONA
    
    async def volver_conductores(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve a la lista de conductores"""
        query = update.callback_query
        await query.answer()
        
        zona = context.user_data.get('zona_seleccionada', '')
        conductores = self._obtener_conductores_en_ruta(zona)
        context.user_data['conductores'] = conductores
        
        if not conductores:
            await query.edit_message_text(
                f"‚ùå No hay conductores en ruta en {zona}.",
                reply_markup=self._get_keyboard_zonas()
            )
            return MOD_RUTA_ZONA
        
        texto = f"üöõ *CONDUCTORES EN RUTA - {zona}*\n\n"
        keyboard = []
        
        for i, c in enumerate(conductores):
            nombre = c.get('nombre', 'N/A')
            telefono = c.get('telefono', 'Sin tel√©fono')
            carga = c.get('lugar_carga', '?')
            descarga = c.get('lugar_entrega', '?')
            n_cargas = len(c.get('cargas', []))
            n_descargas = len(c.get('descargas', []))
            
            texto += f"‚Ä¢ *{nombre}* - üìû {telefono}\n"
            texto += f"   üöõ {carga} ‚Üí {descarga}"
            if n_cargas > 1 or n_descargas > 1:
                texto += f" ({n_cargas}C/{n_descargas}D)"
            texto += "\n\n"
            
            nombre_corto = nombre.split()[0] if nombre else f"Conductor {i+1}"
            keyboard.append([
                InlineKeyboardButton(
                    f"{nombre_corto} - {carga}‚Üí{descarga}",
                    callback_data=f"conductor_{i}"
                )
            ])
        
        keyboard.append([InlineKeyboardButton("‚¨ÖÔ∏è Volver", callback_data="volver_zonas")])
        
        await query.edit_message_text(
            texto,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MOD_RUTA_CONDUCTOR
    
    async def volver_detalle(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve al detalle del viaje"""
        if update.callback_query:
            query = update.callback_query
            await query.answer()
            message = query.message
        else:
            message = update.message
        
        conductor = context.user_data.get('conductor_seleccionado', {})
        cambios = context.user_data.get('cambios_pendientes', {})
        
        texto = self._formatear_detalle_viaje(conductor, cambios)
        keyboard = self._get_keyboard_campos(conductor)
        
        if update.callback_query:
            await update.callback_query.edit_message_text(
                texto,
                parse_mode="Markdown",
                reply_markup=keyboard
            )
        else:
            await message.reply_text(
                texto,
                parse_mode="Markdown",
                reply_markup=keyboard
            )
        return MOD_RUTA_DETALLE
    
    async def cancelar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela la operaci√≥n"""
        context.user_data.clear()
        
        from teclados import teclado_admin
        
        if update.callback_query:
            await update.callback_query.answer()
            await update.callback_query.message.reply_text(
                "‚ùå Operaci√≥n cancelada.\n\n¬øQu√© m√°s necesitas?",
                reply_markup=teclado_admin
            )
        else:
            await update.message.reply_text(
                "‚ùå Operaci√≥n cancelada.\n\n¬øQu√© m√°s necesitas?",
                reply_markup=teclado_admin
            )
        
        return ConversationHandler.END


# ============================================================
# FUNCI√ìN PARA INTEGRAR EN BOT_TRANSPORTE.PY
# ============================================================

def crear_modificador_viajes_ruta(excel_path: str, db_path: str, es_admin_func,
                                    subir_drive_func=None, bot=None, movildata_api=None):
    """
    Crea una instancia del modificador de viajes en ruta.
    
    Uso en bot_transporte.py:
    
        from modificador_viajes_ruta import crear_modificador_viajes_ruta, ModificadorViajesRuta
        
        # En main():
        modificador_ruta = crear_modificador_viajes_ruta(
            config.EXCEL_EMPRESA,
            config.DB_PATH,
            es_admin,
            subir_excel_a_drive,
            app.bot,
            movildata
        )
        app.add_handler(modificador_ruta.get_conversation_handler())
    """
    return ModificadorViajesRuta(
        excel_path=excel_path,
        db_path=db_path,
        es_admin_func=es_admin_func,
        subir_drive_func=subir_drive_func,
        bot=bot,
        movildata_api=movildata_api
    )
