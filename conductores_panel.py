"""
PANEL DE CONDUCTORES UNIFICADO v1.1
====================================
Mejora M1: Admin -> Flota -> Conductores

Cambios v1.1:
- Corregido bug: botón "Ver ficha" no funcionaba después de editar

Funcionalidades:
- Lista todos los conductores con estado visual
- Buscar por nombre (escribir filtra)
- Pinchar en uno -> ver ficha completa
- Editar cualquier campo
- Cambios se guardan en Excel + Drive

Estados:
- 🟢 Disponible (sin viaje activo)
- 🚛 En ruta
- 📥 En carga
- 📤 En descarga
- 🔴 BAJA / VACACIONES
"""

import sqlite3
import logging
from typing import Dict, List, Optional
from pathlib import Path

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes,
    ConversationHandler,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters
)
import openpyxl
from openpyxl.comments import Comment

logger = logging.getLogger(__name__)

# ============================================================
# ESTADOS DE LA CONVERSACIÓN
# ============================================================
COND_LISTA = 500
COND_BUSCAR = 501
COND_FICHA = 502
COND_EDITAR_MENU = 503
COND_EDITAR_VALOR = 504

# Número de conductores por página
CONDUCTORES_POR_PAGINA = 8


class ConductoresPanel:
    """Panel unificado de gestión de conductores"""
    
    def __init__(self, excel_path: str, db_path: str, es_admin_func, subir_drive_func=None):
        self.excel_path = excel_path
        self.db_path = db_path
        self.es_admin = es_admin_func
        self.subir_drive = subir_drive_func
        
        self.campos_editables = {
            'nombre': '👤 Nombre',
            'telefono': '📱 Teléfono',
            'tractora': '🚛 Tractora',
            'remolque': '📦 Remolque',
            'ubicacion': '📍 Ubicación',
            'zona': '🗺️ Zona',
            'absentismo': '🔴 Estado (Activo/Baja/Vacaciones)',
        }
        
        logger.info("[CONDUCTORES_PANEL] Panel de conductores v1.1 inicializado")
    
    def get_conversation_handler(self):
        """Devuelve el ConversationHandler"""
        return ConversationHandler(
            entry_points=[
                MessageHandler(filters.Regex("^👥 Conductores$"), self.inicio),
                CommandHandler("conductores_panel", self.inicio),
            ],
            states={
                COND_LISTA: [
                    CallbackQueryHandler(self.ver_ficha, pattern="^cond_ver_"),
                    CallbackQueryHandler(self.pagina, pattern="^cond_pag_"),
                    CallbackQueryHandler(self.modo_buscar, pattern="^cond_buscar$"),
                    CallbackQueryHandler(self.volver_lista, pattern="^cond_volver_lista$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cond_cancelar$"),
                ],
                COND_BUSCAR: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.buscar_conductor),
                    CallbackQueryHandler(self.volver_lista, pattern="^cond_volver_lista$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cond_cancelar$"),
                ],
                COND_FICHA: [
                    CallbackQueryHandler(self.ver_ficha, pattern="^cond_ver_"),
                    CallbackQueryHandler(self.menu_editar, pattern="^cond_editar$"),
                    CallbackQueryHandler(self.volver_ficha, pattern="^cond_volver_ficha$"),
                    CallbackQueryHandler(self.volver_lista, pattern="^cond_volver_lista$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cond_cancelar$"),
                ],
                COND_EDITAR_MENU: [
                    CallbackQueryHandler(self.ver_ficha, pattern="^cond_ver_"),
                    CallbackQueryHandler(self.elegir_campo, pattern="^cond_campo_"),
                    CallbackQueryHandler(self.volver_ficha, pattern="^cond_volver_ficha$"),
                    CallbackQueryHandler(self.volver_lista, pattern="^cond_volver_lista$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cond_cancelar$"),
                ],
                COND_EDITAR_VALOR: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.guardar_campo),
                    CallbackQueryHandler(self.ver_ficha, pattern="^cond_ver_"),
                    CallbackQueryHandler(self.volver_editar_menu, pattern="^cond_volver_editar$"),
                    CallbackQueryHandler(self.volver_lista, pattern="^cond_volver_lista$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^cond_cancelar$"),
                ],
            },
            fallbacks=[
                CommandHandler("cancelar", self.cancelar),
                MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                MessageHandler(filters.Regex("^⬅️ Volver al menú$"), self.cancelar),
            ],
        )
    
    # ============================================================
    # OBTENER DATOS
    # ============================================================
    
    def _obtener_conductores(self) -> List[Dict]:
        """Obtiene todos los conductores con su estado actual"""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Obtener conductores
            cursor.execute("""
                SELECT id, nombre, tractora, remolque, ubicacion, zona, 
                       absentismo, telefono, fila_excel, telegram_id
                FROM conductores_empresa
                ORDER BY nombre
            """)
            conductores_raw = [dict(row) for row in cursor.fetchall()]
            
            # Obtener viajes activos para determinar estado
            cursor.execute("""
                SELECT conductor_asignado, estado 
                FROM viajes_empresa 
                WHERE estado IN ('pendiente', 'en_ruta', 'asignado', 'en_carga', 'en_descarga')
            """)
            viajes_activos = {row['conductor_asignado'].upper(): row['estado'] 
                           for row in cursor.fetchall() if row['conductor_asignado']}
            
            conn.close()
            
            # Asignar estado visual a cada conductor
            for c in conductores_raw:
                nombre_upper = (c['nombre'] or '').upper()
                absentismo = (c['absentismo'] or '').upper()
                
                if 'BAJA' in absentismo:
                    c['estado_emoji'] = '🔴'
                    c['estado_texto'] = 'BAJA'
                elif 'VACACIONES' in absentismo:
                    c['estado_emoji'] = '🔴'
                    c['estado_texto'] = 'VACACIONES'
                elif nombre_upper in viajes_activos:
                    estado_viaje = viajes_activos[nombre_upper]
                    if estado_viaje == 'en_carga':
                        c['estado_emoji'] = '📥'
                        c['estado_texto'] = 'En carga'
                    elif estado_viaje == 'en_descarga':
                        c['estado_emoji'] = '📤'
                        c['estado_texto'] = 'En descarga'
                    else:
                        c['estado_emoji'] = '🚛'
                        c['estado_texto'] = 'En ruta'
                else:
                    c['estado_emoji'] = '🟢'
                    c['estado_texto'] = 'Disponible'
            
            return conductores_raw
            
        except Exception as e:
            logger.error(f"[CONDUCTORES_PANEL] Error obteniendo conductores: {e}")
            return []
    
    def _obtener_conductor_por_id(self, conductor_id: int) -> Optional[Dict]:
        """Obtiene un conductor por su ID"""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM conductores_empresa WHERE id = ?
            """, (conductor_id,))
            
            row = cursor.fetchone()
            conn.close()
            
            if row:
                return dict(row)
            return None
            
        except Exception as e:
            logger.error(f"[CONDUCTORES_PANEL] Error: {e}")
            return None
    
    # ============================================================
    # HANDLERS
    # ============================================================
    
    async def inicio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra la lista de conductores"""
        user = update.effective_user
        
        if not self.es_admin(user.id):
            from teclados import teclado_admin
            await update.message.reply_text(
                "❌ Solo para responsables.",
                reply_markup=teclado_admin
            )
            return ConversationHandler.END
        
        conductores = self._obtener_conductores()
        context.user_data['conductores'] = conductores
        context.user_data['conductores_filtrados'] = conductores
        context.user_data['pagina'] = 0
        
        return await self._mostrar_lista(update, context, es_mensaje=True)
    
    async def _mostrar_lista(self, update: Update, context: ContextTypes.DEFAULT_TYPE, 
                             es_mensaje: bool = False, filtro: str = None):
        """Muestra lista paginada de conductores"""
        conductores = context.user_data.get('conductores', [])
        pagina = context.user_data.get('pagina', 0)
        
        # Aplicar filtro si existe
        if filtro:
            filtro_upper = filtro.upper()
            conductores = [c for c in conductores if filtro_upper in (c['nombre'] or '').upper()]
            context.user_data['conductores_filtrados'] = conductores
        else:
            conductores = context.user_data.get('conductores_filtrados', conductores)
        
        total = len(conductores)
        inicio = pagina * CONDUCTORES_POR_PAGINA
        fin = inicio + CONDUCTORES_POR_PAGINA
        conductores_pagina = conductores[inicio:fin]
        
        # Construir mensaje
        texto = f"👥 CONDUCTORES ({total})\n"
        if filtro:
            texto += f"🔍 Filtro: {filtro}\n"
        texto += "━━━━━━━━━━━━━━━━━━━━\n\n"
        
        if not conductores_pagina:
            texto += "No hay conductores"
            if filtro:
                texto += f" que coincidan con '{filtro}'"
            texto += ".\n"
        
        # Construir botones inline
        keyboard = []
        
        for c in conductores_pagina:
            emoji = c.get('estado_emoji', '⚪')
            nombre = c.get('nombre', '?')[:20]
            ubicacion = c.get('ubicacion', '?')[:15]
            
            # Botón para ver ficha
            keyboard.append([
                InlineKeyboardButton(
                    f"{emoji} {nombre} | {ubicacion}",
                    callback_data=f"cond_ver_{c['id']}"
                )
            ])
        
        # Navegación de páginas
        nav_buttons = []
        if pagina > 0:
            nav_buttons.append(InlineKeyboardButton("⬅️ Anterior", callback_data=f"cond_pag_{pagina-1}"))
        if fin < total:
            nav_buttons.append(InlineKeyboardButton("Siguiente ➡️", callback_data=f"cond_pag_{pagina+1}"))
        if nav_buttons:
            keyboard.append(nav_buttons)
        
        # Botón de búsqueda y cancelar
        keyboard.append([InlineKeyboardButton("🔍 Buscar por nombre", callback_data="cond_buscar")])
        keyboard.append([InlineKeyboardButton("❌ Cerrar", callback_data="cond_cancelar")])
        
        if es_mensaje:
            await update.message.reply_text(
                texto,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        else:
            query = update.callback_query
            await query.edit_message_text(
                texto,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        
        return COND_LISTA
    
    async def pagina(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cambia de página"""
        query = update.callback_query
        await query.answer()
        
        pagina = int(query.data.replace("cond_pag_", ""))
        context.user_data['pagina'] = pagina
        
        return await self._mostrar_lista(update, context)
    
    async def modo_buscar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Activa modo búsqueda"""
        query = update.callback_query
        await query.answer()
        
        keyboard = [[InlineKeyboardButton("⬅️ Volver a lista", callback_data="cond_volver_lista")]]
        
        await query.edit_message_text(
            "🔍 BUSCAR CONDUCTOR\n\n"
            "Escribe el nombre (o parte) del conductor:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return COND_BUSCAR
    
    async def buscar_conductor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Busca conductores por nombre"""
        filtro = update.message.text.strip()
        
        # Recargar conductores
        conductores = self._obtener_conductores()
        context.user_data['conductores'] = conductores
        context.user_data['pagina'] = 0
        
        return await self._mostrar_lista(update, context, es_mensaje=True, filtro=filtro)
    
    async def volver_lista(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve a la lista de conductores"""
        query = update.callback_query
        await query.answer()
        
        # Recargar y limpiar filtro
        conductores = self._obtener_conductores()
        context.user_data['conductores'] = conductores
        context.user_data['conductores_filtrados'] = conductores
        context.user_data['pagina'] = 0
        
        return await self._mostrar_lista(update, context)
    
    async def ver_ficha(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra la ficha del conductor"""
        query = update.callback_query
        await query.answer()
        
        conductor_id = int(query.data.replace("cond_ver_", ""))
        conductor = self._obtener_conductor_por_id(conductor_id)
        
        if not conductor:
            await query.edit_message_text("❌ Conductor no encontrado.")
            return COND_LISTA
        
        context.user_data['conductor_actual'] = conductor
        
        return await self._mostrar_ficha_query(query, conductor)
    
    async def _mostrar_ficha_query(self, query, conductor: Dict):
        """Muestra la ficha del conductor (desde callback query)"""
        absentismo = (conductor.get('absentismo') or '').upper()
        if 'BAJA' in absentismo:
            estado = "🔴 BAJA"
        elif 'VACACIONES' in absentismo:
            estado = "🔴 VACACIONES"
        else:
            estado = "🟢 Activo"
        
        texto = (
            f"📋 FICHA CONDUCTOR\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👤 Nombre: {conductor.get('nombre', '-')}\n"
            f"📱 Teléfono: {conductor.get('telefono', '-') or '-'}\n"
            f"🚛 Tractora: {conductor.get('tractora', '-')}\n"
            f"📦 Remolque: {conductor.get('remolque', '-') or '-'}\n"
            f"📍 Ubicación: {conductor.get('ubicacion', '-')}\n"
            f"🗺️ Zona: {conductor.get('zona', '-')}\n"
            f"📊 Estado: {estado}\n"
        )
        
        if conductor.get('telegram_id'):
            texto += f"📲 Vinculado: ✅\n"
        else:
            texto += f"📲 Vinculado: ❌\n"
        
        keyboard = [
            [InlineKeyboardButton("✏️ Editar", callback_data="cond_editar")],
            [InlineKeyboardButton("⬅️ Volver a lista", callback_data="cond_volver_lista")],
            [InlineKeyboardButton("❌ Cerrar", callback_data="cond_cancelar")]
        ]
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return COND_FICHA
    
    async def menu_editar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra menú de edición"""
        query = update.callback_query
        await query.answer()
        
        conductor = context.user_data.get('conductor_actual', {})
        
        texto = (
            f"✏️ EDITAR: {conductor.get('nombre', '?')}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"¿Qué campo quieres editar?"
        )
        
        keyboard = []
        for campo, etiqueta in self.campos_editables.items():
            keyboard.append([
                InlineKeyboardButton(etiqueta, callback_data=f"cond_campo_{campo}")
            ])
        
        keyboard.append([InlineKeyboardButton("⬅️ Volver a ficha", callback_data="cond_volver_ficha")])
        keyboard.append([InlineKeyboardButton("❌ Cancelar", callback_data="cond_cancelar")])
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return COND_EDITAR_MENU
    
    async def elegir_campo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide el nuevo valor para el campo"""
        query = update.callback_query
        await query.answer()
        
        campo = query.data.replace("cond_campo_", "")
        context.user_data['campo_editar'] = campo
        
        conductor = context.user_data.get('conductor_actual', {})
        etiqueta = self.campos_editables.get(campo, campo)
        valor_actual = conductor.get(campo, '-') or '-'
        
        texto = (
            f"✏️ EDITAR {etiqueta.upper()}\n\n"
            f"Valor actual: {valor_actual}\n\n"
        )
        
        if campo == 'absentismo':
            texto += "Escribe: ACTIVO, BAJA o VACACIONES"
        else:
            texto += f"Escribe el nuevo valor:"
        
        keyboard = [
            [InlineKeyboardButton("⬅️ Volver", callback_data="cond_volver_editar")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="cond_cancelar")]
        ]
        
        await query.edit_message_text(
            texto,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return COND_EDITAR_VALOR
    
    async def guardar_campo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda el nuevo valor del campo"""
        nuevo_valor = update.message.text.strip()
        campo = context.user_data.get('campo_editar')
        conductor = context.user_data.get('conductor_actual', {})
        fila_excel = conductor.get('fila_excel')
        
        if not campo or not fila_excel:
            await update.message.reply_text("❌ Error: datos incompletos.")
            return ConversationHandler.END
        
        # Normalizar absentismo
        if campo == 'absentismo':
            nuevo_valor_upper = nuevo_valor.upper()
            if nuevo_valor_upper in ['ACTIVO', 'ACTIVE', '']:
                nuevo_valor = ''
            elif 'BAJA' in nuevo_valor_upper:
                nuevo_valor = 'BAJA'
            elif 'VACACIONES' in nuevo_valor_upper:
                nuevo_valor = 'VACACIONES'
        
        try:
            # Actualizar Excel
            exito_excel = self._actualizar_excel(fila_excel, campo, nuevo_valor)
            
            # Actualizar BD
            self._actualizar_bd(conductor['id'], campo, nuevo_valor)
            
            # Subir a Drive
            drive_ok = False
            if exito_excel and self.subir_drive:
                try:
                    self.subir_drive()
                    drive_ok = True
                except Exception as e:
                    logger.error(f"[CONDUCTORES_PANEL] Error subiendo a Drive: {e}")
            
            # Actualizar conductor en contexto
            conductor[campo] = nuevo_valor
            context.user_data['conductor_actual'] = conductor
            
            texto = (
                f"✅ CAMPO ACTUALIZADO\n\n"
                f"{self.campos_editables.get(campo, campo)}: {nuevo_valor}\n\n"
            )
            if drive_ok:
                texto += "☁️ Sincronizado con Drive"
            else:
                texto += "⚠️ Guardado local (Drive no sincronizado)"
            
            conductor_id = conductor.get('id')
            keyboard = [
                [InlineKeyboardButton("✏️ Editar otro campo", callback_data="cond_editar")],
                [InlineKeyboardButton("👁️ Ver ficha", callback_data=f"cond_ver_{conductor_id}")],
                [InlineKeyboardButton("⬅️ Volver a lista", callback_data="cond_volver_lista")]
            ]
            
            await update.message.reply_text(
                texto,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            
            # Retornamos COND_FICHA porque ahora tiene el handler cond_ver_
            return COND_FICHA
            
        except Exception as e:
            logger.error(f"[CONDUCTORES_PANEL] Error guardando: {e}")
            await update.message.reply_text(f"❌ Error: {e}")
            return COND_FICHA
    
    def _actualizar_excel(self, fila_excel: int, campo: str, valor: str) -> bool:
        """Actualiza el campo en el Excel"""
        try:
            if not Path(self.excel_path).exists():
                logger.error(f"[CONDUCTORES_PANEL] Excel no encontrado: {self.excel_path}")
                return False
            
            # Mapeo campo -> columna Excel
            columnas = {
                'ubicacion': 2,      # B
                'nombre': 5,         # E
                'absentismo': 6,     # F
                'tractora': 7,       # G
                'remolque': 8,       # H
            }
            
            columna = columnas.get(campo)
            
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            # fila_excel es 0-indexed, openpyxl es 1-indexed
            fila_openpyxl = fila_excel + 1
            
            if fila_openpyxl > ws.max_row:
                logger.error(f"[CONDUCTORES_PANEL] Fila {fila_openpyxl} fuera de rango")
                wb.close()
                return False
            
            if columna:
                celda = ws.cell(row=fila_openpyxl, column=columna)
                celda.value = valor
            
            # Teléfono va en comentario de columna E (nombre)
            if campo == 'telefono':
                celda_nombre = ws.cell(row=fila_openpyxl, column=5)
                if valor:
                    celda_nombre.comment = Comment(f"Tel. empresa: {valor}", "Bot")
                else:
                    celda_nombre.comment = None
            
            # Zona va en observaciones
            if campo == 'zona':
                # La zona se guarda en la columna de observaciones o se detecta por sección
                pass  # Por ahora no modificamos zona en Excel
            
            wb.save(self.excel_path)
            wb.close()
            
            logger.info(f"[CONDUCTORES_PANEL] Excel actualizado: Fila {fila_openpyxl}, {campo} = {valor}")
            return True
            
        except Exception as e:
            logger.error(f"[CONDUCTORES_PANEL] Error actualizando Excel: {e}")
            return False
    
    def _actualizar_bd(self, conductor_id: int, campo: str, valor: str):
        """Actualiza el campo en la BD"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Campos válidos
            campos_validos = ['nombre', 'telefono', 'tractora', 'remolque', 
                            'ubicacion', 'zona', 'absentismo']
            
            if campo not in campos_validos:
                conn.close()
                return
            
            cursor.execute(f"""
                UPDATE conductores_empresa 
                SET {campo} = ?
                WHERE id = ?
            """, (valor, conductor_id))
            
            conn.commit()
            conn.close()
            
            logger.info(f"[CONDUCTORES_PANEL] BD actualizada: conductor {conductor_id}, {campo} = {valor}")
            
        except Exception as e:
            logger.error(f"[CONDUCTORES_PANEL] Error actualizando BD: {e}")
    
    async def volver_ficha(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve a la ficha del conductor"""
        query = update.callback_query
        await query.answer()
        
        conductor = context.user_data.get('conductor_actual', {})
        
        if not conductor:
            return await self.volver_lista(update, context)
        
        # Recargar conductor
        conductor = self._obtener_conductor_por_id(conductor['id'])
        if conductor:
            context.user_data['conductor_actual'] = conductor
        
        return await self._mostrar_ficha_query(query, conductor)
    
    async def _mostrar_ficha(self, update: Update, context: ContextTypes.DEFAULT_TYPE, conductor: Dict):
        """Muestra la ficha del conductor (reutilizable)"""
        query = update.callback_query
        return await self._mostrar_ficha_query(query, conductor)
    
    async def volver_editar_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve al menú de edición"""
        return await self.menu_editar(update, context)
    
    async def cancelar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela (mensaje)"""
        context.user_data.clear()
        
        from teclados import teclado_flota
        
        await update.message.reply_text(
            "❌ Panel de conductores cerrado.",
            reply_markup=teclado_flota
        )
        
        return ConversationHandler.END
    
    async def cancelar_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela (callback)"""
        query = update.callback_query
        await query.answer()
        
        context.user_data.clear()
        
        from teclados import teclado_flota
        
        await query.edit_message_text("❌ Panel de conductores cerrado.")
        await query.message.reply_text(
            "¿Qué más necesitas?",
            reply_markup=teclado_flota
        )
        
        return ConversationHandler.END


# ============================================================
# FUNCIÓN PARA INTEGRAR EN BOT
# ============================================================

def crear_conductores_panel(excel_path: str, db_path: str, es_admin_func, subir_drive_func=None):
    """
    Crea el panel de conductores.
    
    Uso en bot_transporte.py:
    
        from conductores_panel import crear_conductores_panel
        
        panel = crear_conductores_panel(
            config.EXCEL_EMPRESA,
            config.DB_PATH,
            es_admin,
            subir_excel_a_drive if config.DRIVE_ENABLED else None
        )
        app.add_handler(panel.get_conversation_handler())
    """
    return ConductoresPanel(
        excel_path=excel_path,
        db_path=db_path,
        es_admin_func=es_admin_func,
        subir_drive_func=subir_drive_func
    )
