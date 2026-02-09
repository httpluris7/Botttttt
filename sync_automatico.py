#!/usr/bin/env python3
"""
SINCRONIZACIÓN AUTOMÁTICA
==========================
Descarga Excel de Drive y sincroniza con la BD.
Ejecutar con cron cada 5 minutos.

Uso:
    python3 sync_automatico.py

Cron (cada 5 minutos):
    */5 * * * * cd /root/bot-transporte/Botttttt && /root/bot-transporte/venv/bin/python3 sync_automatico.py >> /var/log/sync_transporte.log 2>&1
"""

import os
import sys
import logging
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Cargar variables de entorno
load_dotenv()

# Configuración
EXCEL_EMPRESA = os.getenv("EXCEL_EMPRESA", "PRUEBO.xlsx")
DB_PATH = os.getenv("DB_PATH", "logistica.db")
DRIVE_ENABLED = os.getenv("DRIVE_ENABLED", "false").lower() == "true"
DRIVE_CREDENTIALS = os.getenv("DRIVE_CREDENTIALS", "credentials.json")
DRIVE_EXCEL_EMPRESA_ID = os.getenv("DRIVE_EXCEL_EMPRESA_ID", "")

# Google Drive
SCOPES = ['https://www.googleapis.com/auth/drive']
drive_service = None


def inicializar_drive():
    """Inicializa conexión con Google Drive"""
    global drive_service
    
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        import pickle
        
        creds = None
        
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                if not os.path.exists(DRIVE_CREDENTIALS):
                    logger.error(f"No se encontró {DRIVE_CREDENTIALS}")
                    return False
                
                flow = InstalledAppFlow.from_client_secrets_file(DRIVE_CREDENTIALS, SCOPES)
                creds = flow.run_local_server(port=0)
            
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)
        
        drive_service = build('drive', 'v3', credentials=creds)
        return True
        
    except Exception as e:
        logger.error(f"Error inicializando Drive: {e}")
        return False


def descargar_excel_desde_drive() -> bool:
    """Descarga PRUEBO.xlsx desde Drive"""
    global drive_service
    
    if not drive_service:
        if not inicializar_drive():
            return False
    
    if not DRIVE_EXCEL_EMPRESA_ID:
        logger.warning("No hay ID de Excel en Drive configurado")
        return False
    
    try:
        from googleapiclient.http import MediaIoBaseDownload
        import io
        
        request = drive_service.files().get_media(fileId=DRIVE_EXCEL_EMPRESA_ID)
        
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        with open(EXCEL_EMPRESA, 'wb') as f:
            f.write(fh.getvalue())
        
        logger.info(f"✅ Excel descargado: {EXCEL_EMPRESA}")
        return True
        
    except Exception as e:
        logger.error(f"Error descargando Excel: {e}")
        return False


def sincronizar_bd() -> dict:
    """Sincroniza Excel con la BD"""
    try:
        from separador_excel_empresa import SeparadorExcelEmpresa
        from extractor_telefonos import sincronizar_telefonos
        from generador_direcciones import sincronizar_direcciones
        
        separador = SeparadorExcelEmpresa(DB_PATH)
        resultado = separador.sincronizar_desde_archivo(EXCEL_EMPRESA, forzar=False)
        
        if resultado.get('cambios'):
            # Sincronizar teléfonos y direcciones si hubo cambios
            sincronizar_telefonos(EXCEL_EMPRESA, DB_PATH)
            sincronizar_direcciones(DB_PATH)
            logger.info(f"✅ BD sincronizada: {resultado}")
        else:
            logger.debug("Sin cambios en el Excel")
        
        return resultado
        
    except Exception as e:
        logger.error(f"Error sincronizando BD: {e}")
        return {"exito": False, "error": str(e)}


def sincronizar_transportistas() -> dict:
    """
    Sincroniza la columna TRANSPORTISTA (V) del Excel con la BD.
    
    Bidireccional:
    1. BD → Excel: Si viaje tiene conductor_asignado pero columna V vacía → Escribir
    2. Excel → BD: Si columna V tiene nombre pero BD vacía → Actualizar BD
    """
    try:
        import sqlite3
        from openpyxl import load_workbook
        
        if not Path(EXCEL_EMPRESA).exists():
            return {"exito": False, "error": "Excel no encontrado"}
        
        wb = load_workbook(EXCEL_EMPRESA)
        ws = wb.active
        
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Columna V = 22 en openpyxl
        COL_TRANSPORTISTA = 22
        
        actualizados_excel = 0
        actualizados_bd = 0
        
        # Obtener todos los viajes con fila_excel
        cursor.execute("""
            SELECT id, conductor_asignado, fila_excel 
            FROM viajes_empresa 
            WHERE fila_excel IS NOT NULL
        """)
        viajes = cursor.fetchall()
        
        for viaje in viajes:
            viaje_id = viaje['id']
            conductor_bd = (viaje['conductor_asignado'] or '').strip()
            fila_excel = viaje['fila_excel']
            
            # fila_excel es 0-indexed, openpyxl es 1-indexed
            fila_openpyxl = fila_excel + 1
            
            if fila_openpyxl > ws.max_row:
                continue
            
            celda = ws.cell(row=fila_openpyxl, column=COL_TRANSPORTISTA)
            transportista_excel = (str(celda.value) if celda.value else '').strip()
            
            # Caso 1: BD tiene conductor, Excel vacío → Escribir en Excel
            if conductor_bd and not transportista_excel:
                celda.value = conductor_bd
                actualizados_excel += 1
                logger.info(f"[SYNC] Excel fila {fila_openpyxl}: '{conductor_bd}'")
            
            # Caso 2: Excel tiene nombre, BD vacío → Actualizar BD
            elif transportista_excel and not conductor_bd:
                cursor.execute("""
                    UPDATE viajes_empresa 
                    SET conductor_asignado = ? 
                    WHERE id = ?
                """, (transportista_excel, viaje_id))
                actualizados_bd += 1
                logger.info(f"[SYNC] BD viaje {viaje_id}: '{transportista_excel}'")
        
        # Guardar cambios
        if actualizados_excel > 0:
            wb.save(EXCEL_EMPRESA)
        wb.close()
        
        if actualizados_bd > 0:
            conn.commit()
        conn.close()
        
        return {
            "exito": True,
            "actualizados_excel": actualizados_excel,
            "actualizados_bd": actualizados_bd
        }
        
    except Exception as e:
        logger.error(f"Error sincronizando transportistas: {e}")
        return {"exito": False, "error": str(e)}


def subir_excel_a_drive() -> bool:
    """Sube el Excel actualizado a Drive"""
    global drive_service
    
    if not drive_service:
        if not inicializar_drive():
            return False
    
    if not DRIVE_EXCEL_EMPRESA_ID:
        return False
    
    try:
        from googleapiclient.http import MediaFileUpload
        
        media = MediaFileUpload(EXCEL_EMPRESA, resumable=True)
        drive_service.files().update(
            fileId=DRIVE_EXCEL_EMPRESA_ID,
            media_body=media
        ).execute()
        
        logger.info("✅ Excel subido a Drive")
        return True
        
    except Exception as e:
        logger.error(f"Error subiendo Excel: {e}")
        return False


def main():
    """Función principal"""
    logger.info("="*50)
    logger.info(f"🔄 Sync automático - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Paso 1: Descargar Excel de Drive (si está habilitado)
    if DRIVE_ENABLED and DRIVE_EXCEL_EMPRESA_ID:
        if not descargar_excel_desde_drive():
            logger.warning("No se pudo descargar Excel de Drive, usando local")
    
    # Paso 2: Verificar que existe el Excel
    if not Path(EXCEL_EMPRESA).exists():
        logger.error(f"Excel no encontrado: {EXCEL_EMPRESA}")
        sys.exit(1)
    
    # Paso 3: Sincronizar con BD
    resultado = sincronizar_bd()
    
    if resultado.get('exito'):
        if resultado.get('cambios'):
            logger.info(f"✅ Sync completado: {resultado.get('conductores', 0)} conductores, "
                       f"{resultado.get('viajes', 0)} viajes")
        else:
            logger.debug("✅ Sin cambios")
    else:
        logger.error(f"❌ Error en sync: {resultado.get('error', 'desconocido')}")
        sys.exit(1)
    
    # Paso 4: Sincronizar columna TRANSPORTISTA (bidireccional)
    resultado_trans = sincronizar_transportistas()
    
    if resultado_trans.get('exito'):
        excel_upd = resultado_trans.get('actualizados_excel', 0)
        bd_upd = resultado_trans.get('actualizados_bd', 0)
        
        if excel_upd > 0 or bd_upd > 0:
            logger.info(f"✅ Transportistas: {excel_upd} Excel, {bd_upd} BD")
            
            # Subir Excel a Drive si hubo cambios
            if excel_upd > 0 and DRIVE_ENABLED:
                subir_excel_a_drive()


if __name__ == "__main__":
    main()
