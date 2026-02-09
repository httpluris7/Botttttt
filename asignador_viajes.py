"""
ASIGNADOR AUTOMÁTICO DE VIAJES v3.0
====================================
Asigna viajes a conductores con ENCADENAMIENTO INTELIGENTE.

PROBLEMA RESUELTO:
- Evita asignar viajes incompatibles
- Si el conductor descarga en MÉRIDA, el siguiente viaje debe cargar CERCA de MÉRIDA
- No en AZAGRA (650km de distancia)

PRIORIZACIÓN DE VIAJES:
1. Urgencia (fecha límite de carga)
2. Precio (viajes mejor pagados primero)  
3. Tipo de mercancía (congelado > refrigerado > seco)

SELECCIÓN DE CONDUCTOR:
1. Estado (disponible, no en descanso)
2. Horas disponibles (tacógrafo)
3. Tipo de remolque (frigorífico si necesita frío)
4. ENCADENAMIENTO: Nueva carga cerca de última descarga
5. Distancia al punto de carga (más cercano)
"""

import sqlite3
import logging
import math
import re
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


# ============================================================
# COORDENADAS DE LUGARES CONOCIDOS
# ============================================================

COORDENADAS_LUGARES = {
    # NAVARRA
    "AZAGRA": (42.3167, -1.8833),
    "MELIDA": (42.3833, -1.5500),
    "MÉLIDA": (42.3833, -1.5500),
    "TUDELA": (42.0617, -1.6067),
    "PAMPLONA": (42.8125, -1.6458),
    "SAN ADRIAN": (42.3417, -1.9333),
    "SAN ADRIÁN": (42.3417, -1.9333),
    "PERALTA": (42.3333, -1.8000),
    "FALCES": (42.3833, -1.8000),
    "TAFALLA": (42.5167, -1.6667),
    "OLITE": (42.4833, -1.6500),
    "ESTELLA": (42.6667, -2.0333),
    "MENDAVIA": (42.4333, -2.2000),
    "LODOSA": (42.4333, -2.0833),
    "SARTAGUDA": (42.3833, -2.0500),
    "CORELLA": (42.1167, -1.7833),
    "CINTRUENIGO": (42.0833, -1.8000),
    "CAPARROSO": (42.3333, -1.6333),
    "CARCASTILLO": (42.3667, -1.4667),
    
    # LA RIOJA
    "CALAHORRA": (42.3050, -1.9653),
    "LOGROÑO": (42.4650, -2.4456),
    "ALFARO": (42.1833, -1.7500),
    "ARNEDO": (42.2167, -2.1000),
    "AUTOL": (42.2167, -2.0000),
    "QUEL": (42.2333, -2.0500),
    "ALDEANUEVA": (42.2333, -1.9000),
    "ALDEANUEVA DE EBRO": (42.2333, -1.9000),
    "PRADEJÓN": (42.3000, -2.0333),
    "PRADEJON": (42.3000, -2.0333),
    "RINCON DE SOTO": (42.2333, -1.8500),
    "HARO": (42.5833, -2.8500),
    
    # ARAGÓN
    "ZARAGOZA": (41.6488, -0.8891),
    "HUESCA": (42.1401, -0.4089),
    "TERUEL": (40.3456, -1.1065),
    "CALATAYUD": (41.3500, -1.6333),
    "EJEA": (42.1333, -1.1333),
    "TARAZONA": (41.9000, -1.7167),
    
    # CATALUÑA
    "BARCELONA": (41.3851, 2.1734),
    "VIC": (41.9304, 2.2546),
    "LLEIDA": (41.6176, 0.6200),
    "TARRAGONA": (41.1189, 1.2445),
    "GIRONA": (41.9794, 2.8214),
    "REUS": (41.1561, 1.1069),
    "FIGUERES": (42.2667, 2.9617),
    "MANRESA": (41.7286, 1.8265),
    "SABADELL": (41.5463, 2.1086),
    "TERRASSA": (41.5630, 2.0089),
    "IGUALADA": (41.5833, 1.6167),
    "MARTORELL": (41.4739, 1.9303),
    "MOLLET": (41.5400, 2.2136),
    "GRANOLLERS": (41.6083, 2.2875),
    
    # MADRID Y CENTRO
    "MADRID": (40.4168, -3.7038),
    "MERCAMADRID": (40.3833, -3.6500),
    "TORREJON DE ARDOZ": (40.4603, -3.4689),
    "GETAFE": (40.3047, -3.7311),
    "ALCALA DE HENARES": (40.4819, -3.3635),
    "MOSTOLES": (40.3228, -3.8650),
    "LEGANES": (40.3281, -3.7642),
    "FUENLABRADA": (40.2839, -3.8000),
    "ALCORCON": (40.3489, -3.8317),
    "TOLEDO": (39.8628, -4.0273),
    "GUADALAJARA": (40.6337, -3.1667),
    "ARANJUEZ": (40.0333, -3.6000),
    "ARGANDA": (40.3000, -3.4333),
    
    # PAÍS VASCO
    "BILBAO": (43.2630, -2.9350),
    "VITORIA": (42.8467, -2.6728),
    "VITORIA-GASTEIZ": (42.8467, -2.6728),
    "SAN SEBASTIAN": (43.3183, -1.9812),
    "DONOSTIA": (43.3183, -1.9812),
    "IRUN": (43.3378, -1.7889),
    "EIBAR": (43.1847, -2.4722),
    "DURANGO": (43.1700, -2.6333),
    "BASAURI": (43.2333, -2.8833),
    "BARAKALDO": (43.2956, -2.9906),
    
    # CANTABRIA Y ASTURIAS
    "SANTANDER": (43.4623, -3.8100),
    "TORRELAVEGA": (43.3500, -4.0500),
    "OVIEDO": (43.3614, -5.8494),
    "GIJON": (43.5453, -5.6615),
    "GIJÓN": (43.5453, -5.6615),
    "AVILES": (43.5578, -5.9250),
    "LANGREO": (43.3000, -5.6833),
    "MIERES": (43.2500, -5.7667),
    
    # GALICIA
    "VIGO": (42.2314, -8.7124),
    "A CORUÑA": (43.3713, -8.3960),
    "LA CORUÑA": (43.3713, -8.3960),
    "CORUÑA": (43.3713, -8.3960),
    "SANTIAGO": (42.8782, -8.5448),
    "OURENSE": (42.3400, -7.8648),
    "LUGO": (43.0097, -7.5567),
    "PONTEVEDRA": (42.4310, -8.6447),
    "FERROL": (43.4833, -8.2333),
    
    # VALENCIA Y MURCIA
    "VALENCIA": (39.4699, -0.3763),
    "MERCAVALENCIA": (39.4500, -0.3833),
    "ALICANTE": (38.3452, -0.4815),
    "CASTELLON": (39.9864, -0.0513),
    "SAGUNTO": (39.6833, -0.2667),
    "GANDIA": (38.9667, -0.1833),
    "ALZIRA": (39.1500, -0.4333),
    "MURCIA": (37.9922, -1.1307),
    "MERCAMURCIA": (37.9667, -1.1500),
    "ALCANTARILLA": (37.9694, -1.2136),
    "CARTAGENA": (37.6057, -0.9916),
    "LORCA": (37.6775, -1.7014),
    "ELCHE": (38.2669, -0.6983),
    
    # ANDALUCÍA
    "SEVILLA": (37.3891, -5.9845),
    "MERCASEVILLA": (37.3500, -5.9667),
    "MALAGA": (36.7213, -4.4214),
    "CORDOBA": (37.8882, -4.7794),
    "GRANADA": (37.1773, -3.5986),
    "ALMERIA": (36.8340, -2.4637),
    "JAEN": (37.7796, -3.7849),
    "HUELVA": (37.2571, -6.9497),
    "CADIZ": (36.5271, -6.2886),
    "JEREZ": (36.6817, -6.1378),
    "ALGECIRAS": (36.1408, -5.4536),
    "MOTRIL": (36.7500, -3.5167),
    "ANTEQUERA": (37.0167, -4.5500),
    
    # EXTREMADURA
    "MERIDA": (38.9161, -6.3436),
    "MÉRIDA": (38.9161, -6.3436),
    "BADAJOZ": (38.8794, -6.9706),
    "CACERES": (39.4753, -6.3724),
    "PLASENCIA": (40.0303, -6.0906),
    "DON BENITO": (38.9553, -5.8614),
    "VILLANUEVA": (38.9833, -5.8000),
    "ALMENDRALEJO": (38.6833, -6.4000),
    "ZAFRA": (38.4167, -6.4167),
    
    # CASTILLA Y LEÓN
    "VALLADOLID": (41.6523, -4.7245),
    "BURGOS": (42.3439, -3.6969),
    "SALAMANCA": (40.9701, -5.6635),
    "LEON": (42.5987, -5.5671),
    "PALENCIA": (42.0096, -4.5288),
    "ZAMORA": (41.5034, -5.7467),
    "AVILA": (40.6566, -4.6819),
    "SEGOVIA": (40.9429, -4.1088),
    "SORIA": (41.7636, -2.4649),
    "ARANDA DE DUERO": (41.6703, -3.6892),
    "MIRANDA DE EBRO": (42.6867, -2.9472),
    "BENAVENTE": (42.0028, -5.6783),
    "PONFERRADA": (42.5500, -6.5833),
    "ASTORGA": (42.4583, -6.0500),
    
    # CASTILLA LA MANCHA
    "ALBACETE": (38.9943, -1.8585),
    "CIUDAD REAL": (38.9848, -3.9274),
    "CUENCA": (40.0704, -2.1374),
    "TALAVERA": (39.9635, -4.8307),
    "PUERTOLLANO": (38.6870, -4.1072),
    "TOMELLOSO": (39.1582, -3.0241),
    "ALCAZAR DE SAN JUAN": (39.3897, -3.2089),
    "MANZANARES": (38.9981, -3.3697),
    "VALDEPENAS": (38.7622, -3.3847),
}

# Distancia máxima para considerar viajes "encadenables" (km)
MAX_DISTANCIA_ENCADENAMIENTO = 150


# ============================================================
# MODELOS DE DATOS
# ============================================================

@dataclass
class ConductorDisponible:
    """Datos de un conductor disponible para asignar"""
    nombre: str
    matricula: str
    lat: float
    lon: float
    horas_restantes_hoy: float
    horas_restantes_semana: float
    tipo_remolque: str
    estado: str
    # Datos del último viaje asignado
    ultima_descarga: str = ""
    lat_ultima_descarga: float = 0.0
    lon_ultima_descarga: float = 0.0
    tiene_viajes_asignados: bool = False
    num_viajes_asignados: int = 0
    # Cálculos
    distancia_a_carga: float = 0.0
    puede_hacer_viaje: bool = True
    motivo_no_puede: str = ""


@dataclass
class ViajeParaAsignar:
    """Datos de un viaje pendiente de asignar"""
    id: int
    cliente: str
    lugar_carga: str
    lugar_entrega: str
    mercancia: str
    km: int
    precio: float
    lat_carga: float
    lon_carga: float
    lat_descarga: float
    lon_descarga: float
    necesita_frio: bool
    horas_estimadas: float
    # Campos de priorización
    fecha_carga: Optional[datetime] = None
    urgente: bool = False
    prioridad: int = 0
    observaciones: str = ""


# ============================================================
# CLASE PRINCIPAL: ASIGNADOR DE VIAJES v3.0
# ============================================================

class AsignadorViajes:
    """Asigna viajes con encadenamiento inteligente"""
    
    VELOCIDAD_MEDIA = 70
    MARGEN_HORAS = 1.0
    
    PALABRAS_FRIO = ['REFRIG', 'CONGEL', 'FRIO', 'FRÍO', '-18', '-20', '-25', '+2', '+4', '+5']
    PALABRAS_URGENTE = ['URGENTE', 'HOY', 'INMEDIATO', 'PRIORIDAD', 'ASAP', 'EXPRESS', '⚠️', '🚨']
    
    def __init__(self, db_path: str, movildata_api, excel_path: str = None, 
                 on_excel_updated: callable = None):
        """
        Inicializa el asignador de viajes.
        
        Args:
            db_path: Ruta a la base de datos SQLite
            movildata_api: Instancia de MovildataAPI para GPS
            excel_path: Ruta al archivo Excel para actualizar la columna TRANSPORTISTA
            on_excel_updated: Callback que se llama cuando se actualiza el Excel
                              (para subir a Drive, por ejemplo)
        """
        self.db_path = db_path
        self.movildata = movildata_api
        self.excel_path = excel_path
        self.on_excel_updated = on_excel_updated
        logger.info("[ASIGNADOR] Inicializado v3.0 con ENCADENAMIENTO INTELIGENTE")
    
    # ═══════════════════════════════════════════════════════════
    # MÉTODOS DE UTILIDAD
    # ═══════════════════════════════════════════════════════════
    
    def _calcular_distancia(self, lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """Calcula distancia en km usando Haversine"""
        if not all([lat1, lon1, lat2, lon2]):
            return 9999
        R = 6371
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    def _obtener_coordenadas(self, lugar: str) -> Tuple[float, float]:
        """Obtiene coordenadas de un lugar"""
        if not lugar:
            return 0.0, 0.0
        
        lugar_upper = lugar.upper().strip()
        
        # Coincidencia exacta
        if lugar_upper in COORDENADAS_LUGARES:
            return COORDENADAS_LUGARES[lugar_upper]
        
        # Coincidencia parcial
        for nombre, coords in COORDENADAS_LUGARES.items():
            if nombre in lugar_upper or lugar_upper in nombre:
                return coords
        
        # Buscar en la dirección completa
        for nombre, coords in COORDENADAS_LUGARES.items():
            if nombre in lugar_upper:
                return coords
        
        logger.warning(f"[ASIGNADOR] ⚠️ Lugar sin coordenadas: {lugar}")
        return 0.0, 0.0
    
    def _necesita_frigorifico(self, mercancia: str) -> bool:
        if not mercancia:
            return False
        mercancia_upper = mercancia.upper()
        return any(palabra in mercancia_upper for palabra in self.PALABRAS_FRIO)
    
    def _es_urgente(self, observaciones: str, cliente: str = "") -> bool:
        texto = f"{observaciones} {cliente}".upper()
        return any(palabra in texto for palabra in self.PALABRAS_URGENTE)
    
    def _estimar_horas_viaje(self, km: int) -> float:
        if not km or km <= 0:
            return 2.0
        return (km / self.VELOCIDAD_MEDIA) + 1.0
    
    def _calcular_prioridad(self, viaje: ViajeParaAsignar) -> int:
        """Calcula prioridad del viaje (0-100)"""
        prioridad = 50
        
        if viaje.urgente:
            prioridad += 40
        
        if viaje.precio:
            if viaje.precio >= 800:
                prioridad += 25
            elif viaje.precio >= 600:
                prioridad += 20
            elif viaje.precio >= 400:
                prioridad += 10
        
        if viaje.mercancia:
            mercancia_upper = viaje.mercancia.upper()
            if 'CONGEL' in mercancia_upper or '-18' in mercancia_upper:
                prioridad += 15
            elif 'REFRIG' in mercancia_upper or '+2' in mercancia_upper:
                prioridad += 10
        
        if viaje.fecha_carga:
            hoy = datetime.now().date()
            dias_restantes = (viaje.fecha_carga.date() - hoy).days
            if dias_restantes <= 0:
                prioridad += 25
            elif dias_restantes == 1:
                prioridad += 15
            elif dias_restantes <= 3:
                prioridad += 5
        
        return min(100, prioridad)
    
    # ═══════════════════════════════════════════════════════════
    # OBTENCIÓN DE DATOS
    # ═══════════════════════════════════════════════════════════
    
    def _obtener_viajes_conductor(self, nombre_conductor: str) -> List[dict]:
        """Obtiene los viajes ya asignados a un conductor"""
        viajes = []
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM viajes_empresa 
                WHERE conductor_asignado = ?
                ORDER BY id DESC
            """, (nombre_conductor,))
            
            for row in cursor.fetchall():
                viajes.append(dict(row))
            
            conn.close()
        except Exception as e:
            logger.error(f"[ASIGNADOR] Error obteniendo viajes de {nombre_conductor}: {e}")
        
        return viajes
    
    def obtener_viajes_pendientes(self) -> List[ViajeParaAsignar]:
        """Obtiene viajes sin conductor, ordenados por prioridad"""
        viajes = []
        
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM viajes_empresa 
                WHERE (conductor_asignado IS NULL OR conductor_asignado = '')
                AND cliente IS NOT NULL AND cliente != ''
            """)
            
            for row in cursor.fetchall():
                viaje_dict = dict(row)
                lugar_carga = viaje_dict.get('lugar_carga', '')
                lugar_entrega = viaje_dict.get('lugar_entrega', '')
                mercancia = viaje_dict.get('mercancia', '')
                km = viaje_dict.get('km', 0) or 0
                precio = viaje_dict.get('precio', 0) or 0
                observaciones = viaje_dict.get('observaciones', '') or ''
                cliente = viaje_dict.get('cliente', '')
                
                lat_carga, lon_carga = self._obtener_coordenadas(lugar_carga)
                lat_descarga, lon_descarga = self._obtener_coordenadas(lugar_entrega)
                
                urgente = self._es_urgente(observaciones, cliente)
                
                # Parsear fecha si existe
                fecha_carga = None
                fecha_match = re.search(r'(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?', observaciones)
                if fecha_match:
                    try:
                        dia = int(fecha_match.group(1))
                        mes = int(fecha_match.group(2))
                        año = int(fecha_match.group(3)) if fecha_match.group(3) else datetime.now().year
                        if año < 100:
                            año += 2000
                        fecha_carga = datetime(año, mes, dia)
                    except:
                        pass
                
                viaje = ViajeParaAsignar(
                    id=viaje_dict.get('id', 0),
                    cliente=cliente,
                    lugar_carga=lugar_carga,
                    lugar_entrega=lugar_entrega,
                    mercancia=mercancia,
                    km=km,
                    precio=precio,
                    lat_carga=lat_carga,
                    lon_carga=lon_carga,
                    lat_descarga=lat_descarga,
                    lon_descarga=lon_descarga,
                    necesita_frio=self._necesita_frigorifico(mercancia),
                    horas_estimadas=self._estimar_horas_viaje(km),
                    fecha_carga=fecha_carga,
                    urgente=urgente,
                    observaciones=observaciones
                )
                viaje.prioridad = self._calcular_prioridad(viaje)
                viajes.append(viaje)
            
            conn.close()
            
            # ORDENAR POR PRIORIDAD
            viajes.sort(key=lambda v: (-v.prioridad, -v.precio, v.km))
            
            logger.info(f"[ASIGNADOR] {len(viajes)} viajes pendientes")
            
        except Exception as e:
            logger.error(f"[ASIGNADOR] Error obteniendo viajes: {e}")
        
        return viajes
    
    def obtener_conductores_disponibles(self) -> List[ConductorDisponible]:
        """Obtiene conductores con datos de viajes previos"""
        conductores = []
        
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM conductores_empresa")
            
            for row in cursor.fetchall():
                conductor_dict = dict(row)
                nombre = conductor_dict.get('nombre', '')
                matricula = conductor_dict.get('tractora', '')
                
                if not matricula:
                    continue
                
                # GPS actual
                pos = self.movildata.get_last_location_plate(matricula)
                if not pos:
                    continue
                
                lat = pos.get('latitud', 0)
                lon = pos.get('longitud', 0)
                
                # Estado
                estado_vehiculo = self.movildata.get_vehicle_status(matricula)
                estado = estado_vehiculo.get('estado', 'DESCONOCIDO') if estado_vehiculo else 'DESCONOCIDO'
                
                # Remolque
                vehiculos = self.movildata.get_vehiculos()
                vehiculo = next((v for v in vehiculos if v.get('matricula') == matricula), None)
                tipo_remolque = vehiculo.get('tipo_remolque', 'FRIGORIFICO') if vehiculo else 'FRIGORIFICO'
                
                # Disponibilidad
                disponibilidad = self.movildata.get_disponibilidad_por_nombre(nombre)
                if disponibilidad:
                    horas_restantes_hoy = disponibilidad.get('horas_restantes_hoy', 9.0)
                    horas_restantes_semana = disponibilidad.get('horas_restantes_semana', 56.0)
                else:
                    horas_restantes_hoy = 9.0
                    horas_restantes_semana = 56.0
                
                # ═══════════════════════════════════════════════════
                # OBTENER VIAJES YA ASIGNADOS (PARA ENCADENAMIENTO)
                # ═══════════════════════════════════════════════════
                viajes_asignados = self._obtener_viajes_conductor(nombre)
                tiene_viajes = len(viajes_asignados) > 0
                
                ultima_descarga = ""
                lat_ultima_descarga = 0.0
                lon_ultima_descarga = 0.0
                
                if tiene_viajes:
                    # Obtener el último viaje (más reciente)
                    ultimo_viaje = viajes_asignados[0]
                    ultima_descarga = ultimo_viaje.get('lugar_entrega', '')
                    lat_ultima_descarga, lon_ultima_descarga = self._obtener_coordenadas(ultima_descarga)
                    
                    logger.info(f"[ASIGNADOR] 📍 {nombre} tiene {len(viajes_asignados)} viaje(s), última descarga: {ultima_descarga}")
                
                conductor = ConductorDisponible(
                    nombre=nombre,
                    matricula=matricula,
                    lat=lat,
                    lon=lon,
                    horas_restantes_hoy=horas_restantes_hoy,
                    horas_restantes_semana=horas_restantes_semana,
                    tipo_remolque=tipo_remolque,
                    estado=estado,
                    ultima_descarga=ultima_descarga,
                    lat_ultima_descarga=lat_ultima_descarga,
                    lon_ultima_descarga=lon_ultima_descarga,
                    tiene_viajes_asignados=tiene_viajes,
                    num_viajes_asignados=len(viajes_asignados)
                )
                conductores.append(conductor)
            
            conn.close()
            logger.info(f"[ASIGNADOR] {len(conductores)} conductores disponibles")
            
        except Exception as e:
            logger.error(f"[ASIGNADOR] Error obteniendo conductores: {e}")
        
        return conductores
    
    # ═══════════════════════════════════════════════════════════
    # FILTRADO CON ENCADENAMIENTO
    # ═══════════════════════════════════════════════════════════
    
    def filtrar_conductores_para_viaje(self, viaje: ViajeParaAsignar, conductores: List[ConductorDisponible]) -> List[ConductorDisponible]:
        """
        Filtra conductores que pueden hacer el viaje.
        
        NUEVO: Verifica ENCADENAMIENTO
        - Si el conductor tiene viajes previos, la nueva carga debe estar
          cerca de su última descarga (máx 150km)
        """
        candidatos = []
        
        for c in conductores:
            c.puede_hacer_viaje = True
            c.motivo_no_puede = ""
            
            # 1. Estado
            if c.estado in ['DESCANSO', 'AVERIA', 'OTROS_TRABAJOS']:
                c.puede_hacer_viaje = False
                c.motivo_no_puede = f"Estado: {c.estado}"
                continue
            
            # 2. Horas
            horas_necesarias = viaje.horas_estimadas + self.MARGEN_HORAS
            if c.horas_restantes_hoy < horas_necesarias:
                c.puede_hacer_viaje = False
                c.motivo_no_puede = f"Horas insuficientes: {c.horas_restantes_hoy:.1f}h (necesita {horas_necesarias:.1f}h)"
                continue
            
            if c.horas_restantes_semana < horas_necesarias:
                c.puede_hacer_viaje = False
                c.motivo_no_puede = f"Horas semana: {c.horas_restantes_semana:.1f}h"
                continue
            
            # 3. Remolque
            if viaje.necesita_frio and c.tipo_remolque != 'FRIGORIFICO':
                c.puede_hacer_viaje = False
                c.motivo_no_puede = "Necesita frigorífico"
                continue
            
            # ═══════════════════════════════════════════════════
            # 4. VERIFICAR ENCADENAMIENTO
            # ═══════════════════════════════════════════════════
            if c.tiene_viajes_asignados:
                # El conductor ya tiene viajes → verificar que el nuevo sea compatible
                if c.lat_ultima_descarga and c.lon_ultima_descarga:
                    distancia_encadenamiento = self._calcular_distancia(
                        c.lat_ultima_descarga, c.lon_ultima_descarga,
                        viaje.lat_carga, viaje.lon_carga
                    )
                    
                    if distancia_encadenamiento > MAX_DISTANCIA_ENCADENAMIENTO:
                        c.puede_hacer_viaje = False
                        c.motivo_no_puede = f"Encadenamiento: {c.ultima_descarga} → {viaje.lugar_carga} = {distancia_encadenamiento:.0f}km (máx {MAX_DISTANCIA_ENCADENAMIENTO}km)"
                        logger.info(f"[ASIGNADOR] ❌ {c.nombre}: {c.motivo_no_puede}")
                        continue
                    else:
                        # Viaje encadenable - usar distancia desde última descarga
                        c.distancia_a_carga = distancia_encadenamiento
                        logger.info(f"[ASIGNADOR] ✅ {c.nombre}: Encadenamiento OK ({c.ultima_descarga} → {viaje.lugar_carga} = {distancia_encadenamiento:.0f}km)")
                else:
                    # Sin coordenadas de última descarga, usar GPS actual
                    if viaje.lat_carga and viaje.lon_carga:
                        c.distancia_a_carga = self._calcular_distancia(c.lat, c.lon, viaje.lat_carga, viaje.lon_carga)
                    else:
                        c.distancia_a_carga = 9999
            else:
                # Sin viajes previos - usar posición GPS actual
                if viaje.lat_carga and viaje.lon_carga and c.lat and c.lon:
                    c.distancia_a_carga = self._calcular_distancia(c.lat, c.lon, viaje.lat_carga, viaje.lon_carga)
                else:
                    c.distancia_a_carga = 9999
            
            candidatos.append(c)
        
        # Ordenar por distancia (más cercano primero)
        candidatos.sort(key=lambda x: x.distancia_a_carga)
        return candidatos
    
    def asignar_viaje(self, viaje: ViajeParaAsignar, conductor: ConductorDisponible) -> bool:
        """Asigna viaje a conductor en BD y actualiza Excel"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Obtener fila_excel antes de actualizar
            cursor.execute("SELECT fila_excel FROM viajes_empresa WHERE id = ?", (viaje.id,))
            row = cursor.fetchone()
            fila_excel = row[0] if row else None
            
            # Actualizar BD
            cursor.execute("""
                UPDATE viajes_empresa 
                SET conductor_asignado = ?, tractora_asignada = ?
                WHERE id = ?
            """, (conductor.nombre, conductor.matricula, viaje.id))
            conn.commit()
            conn.close()
            
            encadenado = "🔗" if conductor.tiene_viajes_asignados else "🆕"
            logger.info(f"[ASIGNADOR] ✅ {encadenado} {viaje.cliente} ({viaje.lugar_carga}→{viaje.lugar_entrega}) → {conductor.nombre}")
            
            # Actualizar Excel si está configurado
            if self.excel_path and fila_excel:
                self._actualizar_excel_transportista(fila_excel, conductor.nombre)
            
            return True
        except Exception as e:
            logger.error(f"[ASIGNADOR] Error: {e}")
            return False
    
    def _actualizar_excel_transportista(self, fila_excel: int, nombre_conductor: str) -> bool:
        """
        Actualiza la columna TRANSPORTISTA (columna V) en el Excel.
        
        Args:
            fila_excel: Número de fila (0-indexed como viene de la BD)
            nombre_conductor: Nombre del conductor asignado
        """
        try:
            from openpyxl import load_workbook
            from pathlib import Path
            
            if not Path(self.excel_path).exists():
                logger.error(f"[ASIGNADOR] Excel no encontrado: {self.excel_path}")
                return False
            
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # fila_excel es 0-indexed, openpyxl es 1-indexed
            fila_openpyxl = fila_excel + 1
            columna_transportista = 22  # Columna V
            
            if fila_openpyxl > ws.max_row:
                logger.error(f"[ASIGNADOR] Fila {fila_openpyxl} fuera de rango")
                return False
            
            celda = ws.cell(row=fila_openpyxl, column=columna_transportista)
            valor_anterior = celda.value
            celda.value = nombre_conductor
            
            wb.save(self.excel_path)
            wb.close()
            
            logger.info(f"[ASIGNADOR] 📝 Excel actualizado: Fila {fila_openpyxl}, "
                       f"TRANSPORTISTA = '{nombre_conductor}' (antes: '{valor_anterior}')")
            
            # Llamar callback si existe (para subir a Drive)
            if self.on_excel_updated:
                try:
                    self.on_excel_updated()
                except Exception as e:
                    logger.error(f"[ASIGNADOR] Error en callback: {e}")
            
            return True
            
        except Exception as e:
            logger.error(f"[ASIGNADOR] Error actualizando Excel: {e}")
            return False
    
    def asignar_viajes_pendientes(self) -> Dict:
        """Asigna todos los viajes pendientes con encadenamiento"""
        resultado = {
            "viajes_pendientes": 0,
            "viajes_asignados": 0,
            "viajes_sin_conductor": 0,
            "viajes_encadenados": 0,
            "asignaciones": [],
            "rechazados": []
        }
        
        viajes = self.obtener_viajes_pendientes()
        resultado["viajes_pendientes"] = len(viajes)
        
        if not viajes:
            return resultado
        
        conductores = self.obtener_conductores_disponibles()
        if not conductores:
            resultado["viajes_sin_conductor"] = len(viajes)
            return resultado
        
        # Actualizar dinámicamente los viajes asignados durante el proceso
        viajes_asignados_ahora = {}  # conductor -> último lugar de descarga
        
        for viaje in viajes:
            # Actualizar conductores con viajes asignados en esta ronda
            for c in conductores:
                if c.nombre in viajes_asignados_ahora:
                    c.tiene_viajes_asignados = True
                    c.ultima_descarga = viajes_asignados_ahora[c.nombre]['lugar']
                    c.lat_ultima_descarga = viajes_asignados_ahora[c.nombre]['lat']
                    c.lon_ultima_descarga = viajes_asignados_ahora[c.nombre]['lon']
            
            candidatos = self.filtrar_conductores_para_viaje(viaje, conductores)
            
            if candidatos:
                mejor = candidatos[0]
                if self.asignar_viaje(viaje, mejor):
                    resultado["viajes_asignados"] += 1
                    
                    # Registrar si fue encadenado
                    if mejor.tiene_viajes_asignados:
                        resultado["viajes_encadenados"] += 1
                    
                    # Actualizar última descarga para siguientes asignaciones
                    viajes_asignados_ahora[mejor.nombre] = {
                        'lugar': viaje.lugar_entrega,
                        'lat': viaje.lat_descarga,
                        'lon': viaje.lon_descarga
                    }
                    
                    resultado["asignaciones"].append({
                        "viaje_id": viaje.id,
                        "cliente": viaje.cliente,
                        "ruta": f"{viaje.lugar_carga} → {viaje.lugar_entrega}",
                        "conductor": mejor.nombre,
                        "matricula": mejor.matricula,
                        "distancia_a_carga": round(mejor.distancia_a_carga, 1),
                        "prioridad": viaje.prioridad,
                        "urgente": viaje.urgente,
                        "precio": viaje.precio,
                        "horas_disponibles": mejor.horas_restantes_hoy,
                        "encadenado": mejor.tiene_viajes_asignados,
                        "desde": mejor.ultima_descarga if mejor.tiene_viajes_asignados else "GPS"
                    })
            else:
                resultado["viajes_sin_conductor"] += 1
                resultado["rechazados"].append({
                    "viaje_id": viaje.id,
                    "cliente": viaje.cliente,
                    "ruta": f"{viaje.lugar_carga} → {viaje.lugar_entrega}"
                })
        
        logger.info(f"[ASIGNADOR] Resultado: {resultado['viajes_asignados']}/{resultado['viajes_pendientes']} (🔗{resultado['viajes_encadenados']} encadenados)")
        return resultado


# Funciones globales
_instancia_asignador = None

def inicializar_asignador(db_path: str, movildata_api, excel_path: str = None,
                          on_excel_updated: callable = None) -> AsignadorViajes:
    """
    Inicializa el asignador de viajes.
    
    Args:
        db_path: Ruta a la base de datos
        movildata_api: API de Movildata para GPS
        excel_path: Ruta al Excel para actualizar columna TRANSPORTISTA
        on_excel_updated: Función a llamar cuando se actualiza el Excel (ej: subir a Drive)
    """
    global _instancia_asignador
    _instancia_asignador = AsignadorViajes(db_path, movildata_api, excel_path, on_excel_updated)
    return _instancia_asignador

def obtener_asignador() -> Optional[AsignadorViajes]:
    return _instancia_asignador
