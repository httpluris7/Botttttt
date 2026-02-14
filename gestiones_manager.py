"""
GESTIONES - CAMIONEROS Y VIAJES (v2.4 - BOTONES INLINE + ASIGNAR CONDUCTOR)
===============================================================
Sistema completo para añadir y modificar camioneros y viajes.

CAMBIOS v2.4:
- Lista de viajes con botones inline
- Campos editables con botones
- Nuevo campo "Asignar Conductor" con lista de conductores
- Notificación automática al conductor cuando se le asigna viaje

CAMBIOS v2.3:
- FIX: _get_viajes_sin_asignar usaba columna 5 (conductor) en vez de 22 (transportista viaje)

CAMBIOS v2.2:
- DESACTIVADO handler "✏️ Modificar viaje" → usar modificador_viajes_pendientes.py

CAMBIOS v2.1:
- Mejorado formato de lista de viajes (similar a modificar en ruta)

CAMBIOS v2.0:
- Hasta 10 cargas y 10 descargas por viaje
- Botón "Volver atrás" durante creación
- Botón "Editar" antes de confirmar
- Modificar viajes/camioneros existentes con gestión de cargas/descargas
- Sincronización con Google Drive
"""

import logging
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes, 
    ConversationHandler, 
    CommandHandler, 
    MessageHandler, 
    CallbackQueryHandler,
    filters
)
import openpyxl
from openpyxl.comments import Comment
from pathlib import Path
from teclados import teclado_admin
import re
from validaciones import (
    validar_telefono,
    validar_matricula_tractora,
    validar_matricula_remolque,
    validar_nombre,
    validar_lugar_carga,
    validar_lugar_descarga,
    validar_cliente,
    validar_mercancia,
    validar_precio,
    validar_km,
    validar_zona,
    validar_observaciones,
    normalizar_ciudad,
    formatear_precio
)

# ============================================================
# CONSTANTES
# ============================================================
MAX_CARGAS = 10
MAX_DESCARGAS = 10

def interpretar_texto(texto: str) -> dict:
    """Interpreta texto libre del usuario"""
    texto_lower = texto.lower().strip()
    
    if texto_lower in ['si', 'sí', 'sip', 'yes', 'ok', 'vale', 'claro']:
        return {'tipo': 'si'}
    if texto_lower.startswith('➕'):
        return {'tipo': 'si'}
    
    if texto_lower in ['no', 'nop', 'nope', 'nel', 'negativo', 'nada']:
        return {'tipo': 'no'}
    if texto_lower.startswith('➡️'):
        return {'tipo': 'no'}
    
    if texto_lower in ['cancelar', 'cancela', 'salir', 'dejalo', 'déjalo']:
        return {'tipo': 'cancelar'}
    if '❌' in texto_lower:
        return {'tipo': 'cancelar'}
    
    if texto_lower in ['volver', 'atras', 'atrás', 'back']:
        return {'tipo': 'volver'}
    if '⬅️' in texto_lower:
        return {'tipo': 'volver'}
    
    return {'tipo': 'dato', 'valor': texto}


logger = logging.getLogger(__name__)

# ============================================================
# ESTADOS DE LA CONVERSACIÓN
# ============================================================
MENU_PRINCIPAL = 0
MENU_TIPO = 1
MENU_ACCION = 2

# Camionero
CAM_NOMBRE = 10
CAM_TELEFONO = 11
CAM_TRACTORA = 12
CAM_REMOLQUE = 13
CAM_UBICACION = 14
CAM_CONFIRMAR = 15
CAM_EDITAR_CAMPO = 16

# Viaje - Añadir
VIA_ZONA = 20
VIA_CLIENTE = 21
VIA_NUM_PEDIDO = 22
VIA_REF_CLIENTE = 23
VIA_INTERCAMBIO = 24
VIA_LUGAR_CARGA = 25
VIA_CARGA_ADICIONAL = 26       # ¿Añadir otra carga? (loop)
VIA_CARGA_ADICIONAL_LUGAR = 27  # Escribir nueva carga
VIA_LUGAR_DESCARGA = 28
VIA_DESCARGA_ADICIONAL = 29     # ¿Añadir otra descarga? (loop)
VIA_DESCARGA_ADICIONAL_LUGAR = 30  # Escribir nueva descarga
VIA_MERCANCIA = 31
VIA_KM = 32
VIA_PRECIO = 33
VIA_CONFIRMAR = 34
VIA_EDITAR_CAMPO = 35
VIA_EDITAR_VALOR = 36

# Modificar existente
MOD_ELEGIR_VIAJE = 40
MOD_ELEGIR_CAMIONERO = 41
MOD_CAMPO = 42
MOD_VALOR = 43

# Submenú gestión cargas/descargas (compartido add-edit y mod)
CARGAS_MENU = 50       # Lista cargas + opciones
CARGAS_EDITAR = 51     # Escribir nuevo valor para carga N
CARGAS_ELIMINAR = 52   # Elegir cuál eliminar
DESCARGAS_MENU = 53    # Lista descargas + opciones
DESCARGAS_EDITAR = 54  # Escribir nuevo valor para descarga N
DESCARGAS_ELIMINAR = 55 # Elegir cuál eliminar


# ============================================================
# HELPERS
# ============================================================

def _inicializar_cargas(viaje: dict):
    """Asegura que viaje tenga listas de cargas/descargas"""
    if 'cargas' not in viaje:
        cargas = []
        if viaje.get('lugar_carga'):
            cargas.append(viaje['lugar_carga'])
        if viaje.get('carga_adicional'):
            cargas.append(viaje['carga_adicional'])
        # Buscar CARGA3..CARGA10 en observaciones
        obs = viaje.get('observaciones', '') or ''
        for i in range(3, MAX_CARGAS + 1):
            match = re.search(rf'CARGA{i}:\s*([^|]+)', obs)
            if match:
                cargas.append(match.group(1).strip())
        viaje['cargas'] = cargas if cargas else []
    
    if 'descargas' not in viaje:
        descargas = []
        if viaje.get('lugar_descarga') or viaje.get('lugar_entrega'):
            descargas.append(viaje.get('lugar_descarga') or viaje.get('lugar_entrega', ''))
        if viaje.get('descarga_adicional'):
            descargas.append(viaje['descarga_adicional'])
        obs = viaje.get('observaciones', '') or ''
        for i in range(3, MAX_DESCARGAS + 1):
            match = re.search(rf'DESCARGA{i}:\s*([^|]+)', obs)
            if match:
                descargas.append(match.group(1).strip())
        viaje['descargas'] = descargas if descargas else []


def _sync_compat(viaje: dict):
    """Sincroniza listas con campos legacy para compatibilidad"""
    cargas = viaje.get('cargas', [])
    viaje['lugar_carga'] = cargas[0] if cargas else ''
    viaje['carga_adicional'] = cargas[1] if len(cargas) > 1 else None
    
    descargas = viaje.get('descargas', [])
    viaje['lugar_descarga'] = descargas[0] if descargas else ''
    viaje['descarga_adicional'] = descargas[1] if len(descargas) > 1 else None


def _formatear_lista_cargas(cargas: list, tipo: str = "carga") -> str:
    """Formatea lista de cargas/descargas para mostrar"""
    if not cargas:
        return "_Sin definir_"
    emoji = "📥" if tipo == "carga" else "📤"
    lineas = []
    for i, c in enumerate(cargas):
        etiqueta = "Principal" if i == 0 else f"#{i+1}"
        lineas.append(f"  {emoji} {etiqueta}: *{c}*")
    return "\n".join(lineas)


def _generar_observaciones(viaje: dict) -> str:
    """Genera string de observaciones con todas las cargas/descargas"""
    partes = []
    if viaje.get('zona'):
        partes.append(f"ZONA: {viaje['zona']}")
    
    cargas = viaje.get('cargas', [])
    for i, c in enumerate(cargas[1:], 2):  # Desde la 2ª
        partes.append(f"CARGA{i}: {c}")
    
    descargas = viaje.get('descargas', [])
    for i, d in enumerate(descargas[1:], 2):  # Desde la 2ª
        partes.append(f"DESCARGA{i}: {d}")
    
    return " | ".join(partes)


def _generar_comentario_cargas(cargas: list) -> str:
    """Genera texto para Comment de Excel"""
    if len(cargas) <= 1:
        return ""
    return f"{len(cargas)} CARGAS: " + " + ".join(cargas)


def _generar_comentario_descargas(descargas: list) -> str:
    """Genera texto para Comment de Excel"""
    if len(descargas) <= 1:
        return ""
    return f"{len(descargas)} DESCARGAS: " + " + ".join(descargas)


# ============================================================
# CLASE PRINCIPAL
# ============================================================

class GestionesManager:
    """Gestiona altas y modificaciones de camioneros y viajes"""
    
    def __init__(self, excel_path: str, db_path: str, es_admin_func, subir_drive_func=None):
        self.excel_path = excel_path
        self.db_path = db_path
        self.es_admin = es_admin_func
        self.subir_drive = subir_drive_func
        
        # Mapeo de campos de viaje (para editar al añadir y para modificar)
        self.campos_viaje = {
            '1': ('zona', '🗺️ Zona'),
            '2': ('cliente', '🏢 Cliente'),
            '3': ('num_pedido', '🔢 Nº Pedido'),
            '4': ('ref_cliente', '🏷️ Ref. Cliente'),
            '5': ('intercambio', '🔄 Intercambio'),
            '6': ('cargas', '📍 Cargas'),
            '7': ('descargas', '📍 Descargas'),
            '8': ('mercancia', '📦 Mercancía'),
            '9': ('km', '📏 Kilómetros'),
            '10': ('precio', '💰 Precio'),
            '11': ('transportista', '🚛 Conductor'),
        }
        
        # Mapeo de campos de camionero
        self.campos_camionero = {
            '1': ('nombre', '👤 Nombre'),
            '2': ('telefono', '📱 Teléfono'),
            '3': ('tractora', '🚛 Tractora'),
            '4': ('remolque', '📦 Remolque'),
            '5': ('ubicacion', '📍 Ubicación'),
        }

    # ============================================================
    # CONVERSATION HANDLER
    # ============================================================
        
    def get_conversation_handler(self):
        """Devuelve el ConversationHandler"""
        return ConversationHandler(
            entry_points=[
                MessageHandler(filters.Regex("^🛠️ Gestiones$"), self.inicio),
                CommandHandler("gestiones", self.inicio),
                CommandHandler("anadir_conductor", self.inicio_añadir_conductor),
                CommandHandler("anadir_viaje", self.inicio_añadir_viaje),
                CommandHandler("modificar_conductor", self.inicio_modificar_conductor),
                # CommandHandler("modificar_viaje", self.inicio_modificar_viaje),  # DESACTIVADO - usar modificador_viajes_pendientes.py
                # Nuevos botones desde submenús
                MessageHandler(filters.Regex("^➕ Añadir camionero$"), self.inicio_añadir_conductor),
                MessageHandler(filters.Regex("^✏️ Modificar camionero$"), self.inicio_modificar_conductor),
                MessageHandler(filters.Regex("^➕ Añadir viaje$"), self.inicio_añadir_viaje),
                # MessageHandler(filters.Regex("^✏️ Modificar viaje$"), self.inicio_modificar_viaje),  # DESACTIVADO
                MessageHandler(filters.Regex("^🔄 Sincronizar$"), self.sincronizar_drive),
            ],
            states={
                MENU_PRINCIPAL: [
                    MessageHandler(filters.Regex("^🚛 Camionero$"), self.menu_camionero),
                    MessageHandler(filters.Regex("^📦 Viaje$"), self.menu_viaje),
                    MessageHandler(filters.Regex("^🔄 Sincronizar$"), self.sincronizar_drive),
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                ],
                MENU_ACCION: [
                    MessageHandler(filters.Regex("^➕ Añadir$"), self.accion_añadir),
                    MessageHandler(filters.Regex("^✏️ Modificar$"), self.accion_modificar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.inicio),
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                ],
                
                # === CAMIONERO ===
                CAM_NOMBRE: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.cam_nombre),
                ],
                CAM_TELEFONO: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.cam_volver_nombre),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.cam_telefono),
                ],
                CAM_TRACTORA: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.cam_volver_telefono),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.cam_tractora),
                ],
                CAM_REMOLQUE: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.cam_volver_tractora),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.cam_remolque),
                ],
                CAM_UBICACION: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.cam_volver_remolque),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.cam_ubicacion),
                ],
                CAM_CONFIRMAR: [
                    MessageHandler(filters.Regex("^✅ Confirmar$"), self.cam_guardar),
                    MessageHandler(filters.Regex("^✏️ Editar$"), self.cam_editar),
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                ],
                CAM_EDITAR_CAMPO: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.cam_mostrar_resumen),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.cam_editar_campo),
                ],
                
                # === VIAJE - AÑADIR ===
                VIA_ZONA: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.menu_viaje),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_zona),
                ],
                VIA_CLIENTE: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_volver_zona),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_cliente),
                ],
                VIA_NUM_PEDIDO: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_volver_cliente),
                    MessageHandler(filters.Regex("^⏭️ Saltar$"), self.via_saltar_num_pedido),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_num_pedido),
                ],
                VIA_REF_CLIENTE: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_volver_num_pedido),
                    MessageHandler(filters.Regex("^⏭️ Saltar$"), self.via_saltar_ref_cliente),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_ref_cliente),
                ],
                VIA_INTERCAMBIO: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_volver_ref_cliente),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_intercambio),
                ],
                VIA_LUGAR_CARGA: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_volver_intercambio),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_lugar_carga),
                ],
                VIA_CARGA_ADICIONAL: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^➕ Sí$"), self.via_pedir_carga_adicional),
                    MessageHandler(filters.Regex("^➡️ No$"), self.via_saltar_carga_adicional),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._via_volver_desde_carga_adicional),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._handler_carga_adicional),
                ],
                VIA_CARGA_ADICIONAL_LUGAR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._via_preguntar_mas_cargas),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_carga_adicional_lugar),
                ],
                VIA_LUGAR_DESCARGA: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._via_volver_a_cargas),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_lugar_descarga),
                ],
                VIA_DESCARGA_ADICIONAL: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^➕ Sí$"), self.via_pedir_descarga_adicional),
                    MessageHandler(filters.Regex("^➡️ No$"), self.via_saltar_descarga_adicional),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._via_volver_desde_descarga_adicional),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._handler_descarga_adicional),
                ],
                VIA_DESCARGA_ADICIONAL_LUGAR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._via_preguntar_mas_descargas),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_descarga_adicional_lugar),
                ],
                VIA_MERCANCIA: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._via_volver_a_descargas),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_mercancia),
                ],
                VIA_KM: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_volver_mercancia),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_km),
                ],
                VIA_PRECIO: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_volver_km),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_precio),
                ],
                VIA_CONFIRMAR: [
                    MessageHandler(filters.Regex("^✅ Confirmar$"), self.via_guardar),
                    MessageHandler(filters.Regex("^✏️ Editar$"), self.via_editar),
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                ],
                VIA_EDITAR_CAMPO: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_mostrar_resumen),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_editar_campo),
                ],
                VIA_EDITAR_VALOR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.via_editar),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.via_editar_valor),
                ],
                
                # === SUBMENÚ CARGAS/DESCARGAS (compartido) ===
                CARGAS_MENU: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^➕ Añadir carga$"), self._cargas_añadir),
                    MessageHandler(filters.Regex("^🗑️ Eliminar carga$"), self._cargas_pedir_eliminar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._cargas_volver),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._cargas_seleccionar),
                ],
                CARGAS_EDITAR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._mostrar_cargas_menu),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._cargas_guardar_edicion),
                ],
                CARGAS_ELIMINAR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._mostrar_cargas_menu),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._cargas_confirmar_eliminar),
                ],
                DESCARGAS_MENU: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^➕ Añadir descarga$"), self._descargas_añadir),
                    MessageHandler(filters.Regex("^🗑️ Eliminar descarga$"), self._descargas_pedir_eliminar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._descargas_volver),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._descargas_seleccionar),
                ],
                DESCARGAS_EDITAR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._mostrar_descargas_menu),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._descargas_guardar_edicion),
                ],
                DESCARGAS_ELIMINAR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._mostrar_descargas_menu),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self._descargas_confirmar_eliminar),
                ],
                
                # === MODIFICAR EXISTENTE ===
                MOD_ELEGIR_VIAJE: [
                    CallbackQueryHandler(self.mod_elegir_viaje_callback, pattern="^modviaje_"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^mod_cancelar$"),
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                ],
                MOD_ELEGIR_CAMIONERO: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self.menu_camionero),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.mod_elegir_camionero),
                ],
                MOD_CAMPO: [
                    CallbackQueryHandler(self.mod_elegir_campo_callback, pattern="^modcampo_"),
                    CallbackQueryHandler(self.mod_asignar_conductor_callback, pattern="^asignar_"),
                    CallbackQueryHandler(self._mod_mostrar_resumen_viaje_callback, pattern="^mod_volver_resumen$"),
                    CallbackQueryHandler(self.mod_guardar_callback, pattern="^mod_guardar$"),
                    CallbackQueryHandler(self.mod_volver_lista_callback, pattern="^mod_volver_lista$"),
                    CallbackQueryHandler(self.cancelar_callback, pattern="^mod_cancelar$"),
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                ],
                MOD_VALOR: [
                    MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
                    MessageHandler(filters.Regex("^⬅️ Volver$"), self._mod_volver_resumen),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.mod_nuevo_valor),
                ],
            },
            fallbacks=[
                CommandHandler("cancelar", self.cancelar),
                MessageHandler(filters.Regex("^❌ Cancelar$"), self.cancelar),
            ],
            name="gestiones",
            persistent=False,
        )

    # ============================================================
    # MENÚ Y NAVEGACIÓN
    # ============================================================
    
    async def inicio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        if not self.es_admin(user.id):
            await update.message.reply_text("❌ Solo para responsables.")
            return ConversationHandler.END
        
        context.user_data.clear()
        keyboard = [["🚛 Camionero", "📦 Viaje"], ["🔄 Sincronizar"], ["❌ Cancelar"]]
        await update.message.reply_text(
            "🛠️ *GESTIONES*\n\n¿Qué quieres gestionar?",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MENU_PRINCIPAL

    async def menu_camionero(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['tipo'] = 'camionero'
        keyboard = [["➕ Añadir", "✏️ Modificar"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "🚛 *CAMIONERO*\n\n¿Qué quieres hacer?",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MENU_ACCION
    
    async def menu_viaje(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['tipo'] = 'viaje'
        keyboard = [["➕ Añadir", "✏️ Modificar"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *VIAJE*\n\n¿Qué quieres hacer?",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MENU_ACCION
    
    async def sincronizar_drive(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Sincronización bidireccional completa con Google Drive"""
        await update.message.reply_text("🔄 Sincronizando con Drive...\n\n⏳ Descargando Excel...")
        
        try:
            from separador_excel_empresa import SeparadorExcelEmpresa
            from extractor_telefonos import sincronizar_telefonos
            from generador_direcciones import sincronizar_direcciones
            import sqlite3
            from openpyxl import load_workbook
            
            # Paso 1: Descargar Excel de Drive
            descargado = False
            if self.subir_drive:
                try:
                    # Intentar descargar desde Drive
                    from sync_automatico import descargar_excel_desde_drive, DRIVE_ENABLED, DRIVE_EXCEL_EMPRESA_ID
                    if DRIVE_ENABLED and DRIVE_EXCEL_EMPRESA_ID:
                        descargado = descargar_excel_desde_drive()
                except Exception as e:
                    logger.warning(f"[GESTIONES] No se pudo descargar de Drive: {e}")
            
            if descargado:
                await update.message.reply_text("✅ Excel descargado de Drive\n⏳ Sincronizando BD...")
            else:
                await update.message.reply_text("⏳ Usando Excel local, sincronizando BD...")
            
            # Paso 2: Sincronizar BD con Excel
            separador = SeparadorExcelEmpresa(self.db_path)
            resultado = separador.sincronizar_desde_archivo(self.excel_path, forzar=True)
            
            conductores = resultado.get('conductores', 0)
            viajes = resultado.get('viajes', 0)
            
            # Sincronizar teléfonos y direcciones
            sincronizar_telefonos(self.excel_path, self.db_path)
            sincronizar_direcciones(self.db_path)
            
            # Paso 3: Sincronizar columna TRANSPORTISTA (bidireccional)
            await update.message.reply_text("⏳ Sincronizando transportistas...")
            
            wb = load_workbook(self.excel_path)
            ws = wb.active
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            COL_TRANSPORTISTA = 22  # Columna V
            actualizados_excel = 0
            actualizados_bd = 0
            
            cursor.execute("""
                SELECT id, conductor_asignado, fila_excel 
                FROM viajes_empresa 
                WHERE fila_excel IS NOT NULL
            """)
            viajes_db = cursor.fetchall()
            
            for viaje in viajes_db:
                viaje_id = viaje['id']
                conductor_bd = (viaje['conductor_asignado'] or '').strip()
                fila_excel = viaje['fila_excel']
                fila_openpyxl = fila_excel + 1
                
                if fila_openpyxl > ws.max_row:
                    continue
                
                celda = ws.cell(row=fila_openpyxl, column=COL_TRANSPORTISTA)
                transportista_excel = (str(celda.value) if celda.value else '').strip()
                
                # BD tiene conductor, Excel vacío → Escribir en Excel
                if conductor_bd and not transportista_excel:
                    celda.value = conductor_bd
                    actualizados_excel += 1
                
                # Excel tiene nombre, BD vacío → Actualizar BD
                elif transportista_excel and not conductor_bd:
                    cursor.execute("""
                        UPDATE viajes_empresa 
                        SET conductor_asignado = ? 
                        WHERE id = ?
                    """, (transportista_excel, viaje_id))
                    actualizados_bd += 1
            
            if actualizados_excel > 0:
                wb.save(self.excel_path)
            wb.close()
            
            if actualizados_bd > 0:
                conn.commit()
            conn.close()
            
            # Paso 4: Subir Excel a Drive
            if self.subir_drive:
                await update.message.reply_text("⏳ Subiendo a Drive...")
                self.subir_drive()
            
            # Resumen final
            mensaje = "✅ *SINCRONIZACIÓN COMPLETADA*\n\n"
            mensaje += f"📥 Descargado de Drive: {'Sí' if descargado else 'No'}\n"
            mensaje += f"👥 Conductores: {conductores}\n"
            mensaje += f"📦 Viajes: {viajes}\n"
            mensaje += f"📝 Transportistas Excel→BD: {actualizados_bd}\n"
            mensaje += f"📝 Transportistas BD→Excel: {actualizados_excel}\n"
            mensaje += f"📤 Subido a Drive: {'Sí' if self.subir_drive else 'No'}"
            
            await update.message.reply_text(
                mensaje,
                parse_mode="Markdown",
                reply_markup=teclado_admin
            )
            
        except Exception as e:
            logger.error(f"[GESTIONES] Error sincronizando: {e}")
            await update.message.reply_text(
                f"❌ Error: {e}",
                reply_markup=teclado_admin
            )
        
        return ConversationHandler.END
    
    async def accion_añadir(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        tipo = context.user_data.get('tipo', 'viaje')
        if tipo == 'camionero':
            return await self._iniciar_añadir_camionero(update, context)
        else:
            return await self._iniciar_añadir_viaje(update, context)
    
    async def accion_modificar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        tipo = context.user_data.get('tipo', 'viaje')
        if tipo == 'camionero':
            return await self.mod_listar_camioneros(update, context)
        else:
            return await self.mod_listar_viajes(update, context)

    async def inicio_añadir_conductor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        if not self.es_admin(user.id):
            await update.message.reply_text("❌ Solo para responsables.")
            return ConversationHandler.END
        context.user_data.clear()
        context.user_data['tipo'] = 'camionero'
        return await self._iniciar_añadir_camionero(update, context)
    
    async def inicio_añadir_viaje(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        if not self.es_admin(user.id):
            await update.message.reply_text("❌ Solo para responsables.")
            return ConversationHandler.END
        context.user_data.clear()
        context.user_data['tipo'] = 'viaje'
        return await self._iniciar_añadir_viaje(update, context)
    
    async def inicio_modificar_conductor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        if not self.es_admin(user.id):
            await update.message.reply_text("❌ Solo para responsables.")
            return ConversationHandler.END
        context.user_data.clear()
        context.user_data['tipo'] = 'camionero'
        return await self.mod_listar_camioneros(update, context)
    
    async def inicio_modificar_viaje(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        if not self.es_admin(user.id):
            await update.message.reply_text("❌ Solo para responsables.")
            return ConversationHandler.END
        context.user_data.clear()
        context.user_data['tipo'] = 'viaje'
        return await self.mod_listar_viajes(update, context)

    async def cancelar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data.clear()
        await update.message.reply_text(
            "❌ Operación cancelada.",
            reply_markup=teclado_admin
        )
        return ConversationHandler.END

    # ============================================================
    # AÑADIR CAMIONERO (sin cambios significativos)
    # ============================================================

    async def _iniciar_añadir_camionero(self, update, context):
        context.user_data['camionero'] = {}
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "🚛 *NUEVO CAMIONERO*\n\nPaso 1/5\n\n👤 *Nombre completo:*",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CAM_NOMBRE
    
    async def cam_nombre(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_nombre(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return CAM_NOMBRE
        context.user_data['camionero']['nombre'] = resultado['valor']
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "🚛 *NUEVO CAMIONERO*\n\nPaso 2/5\n\n📱 *Teléfono:*",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CAM_TELEFONO
    
    async def cam_telefono(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_telefono(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return CAM_TELEFONO
        context.user_data['camionero']['telefono'] = resultado['valor']
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "🚛 *NUEVO CAMIONERO*\n\nPaso 3/5\n\n🚛 *Matrícula tractora:*\n\n_Ejemplo: 1234ABC_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CAM_TRACTORA
    
    async def cam_tractora(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_matricula_tractora(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return CAM_TRACTORA
        context.user_data['camionero']['tractora'] = resultado['valor']
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "🚛 *NUEVO CAMIONERO*\n\nPaso 4/5\n\n📦 *Matrícula remolque:*\n\n_Ejemplo: R1234BBB_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CAM_REMOLQUE
    
    async def cam_remolque(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_matricula_remolque(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return CAM_REMOLQUE
        context.user_data['camionero']['remolque'] = resultado['valor']
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "🚛 *NUEVO CAMIONERO*\n\nPaso 5/5\n\n📍 *Ubicación base:*\n\n_Ejemplo: CALAHORRA_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CAM_UBICACION
    
    async def cam_ubicacion(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['camionero']['ubicacion'] = update.message.text.strip().upper()
        return await self.cam_mostrar_resumen(update, context)
    
    async def cam_mostrar_resumen(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        cam = context.user_data['camionero']
        keyboard = [["✏️ Editar", "✅ Confirmar"], ["❌ Cancelar"]]
        resumen = "🚛 *NUEVO CAMIONERO*\n\n📋 *RESUMEN:*\n━━━━━━━━━━━━━━━━━━━━\n"
        resumen += f"1. 👤 Nombre: *{cam.get('nombre', '-')}*\n"
        resumen += f"2. 📱 Teléfono: *{cam.get('telefono', '-')}*\n"
        resumen += f"3. 🚛 Tractora: *{cam.get('tractora', '-')}*\n"
        resumen += f"4. 📦 Remolque: *{cam.get('remolque', '-')}*\n"
        resumen += f"5. 📍 Ubicación: *{cam.get('ubicacion', '-')}*\n"
        resumen += "━━━━━━━━━━━━━━━━━━━━\n\n¿Es correcto?"
        await update.message.reply_text(resumen, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return CAM_CONFIRMAR
    
    async def cam_editar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "✏️ *EDITAR CAMIONERO*\n\n¿Qué campo? (1-5)\n\n"
            "1. 👤 Nombre\n2. 📱 Teléfono\n3. 🚛 Tractora\n4. 📦 Remolque\n5. 📍 Ubicación",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CAM_EDITAR_CAMPO
    
    async def cam_editar_campo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        campo_num = update.message.text.strip()
        if campo_num not in self.campos_camionero:
            await update.message.reply_text("⚠️ Introduce un número del 1 al 5:")
            return CAM_EDITAR_CAMPO
        campo_key, campo_nombre = self.campos_camionero[campo_num]
        context.user_data['editando_campo'] = campo_key
        valor_actual = context.user_data['camionero'].get(campo_key, '-')
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"✏️ *EDITAR {campo_nombre.split(' ', 1)[1].upper()}*\n\n"
            f"Valor actual: *{valor_actual}*\n\nEscribe el nuevo valor:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        # Reuse VIA_EDITAR_VALOR but handle differently based on type
        return VIA_EDITAR_VALOR
    
    async def cam_guardar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda camionero en Excel"""
        cam = context.user_data['camionero']
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            fila = 3
            while ws.cell(row=fila, column=5).value:
                fila += 1
            ws.cell(row=fila, column=2, value=cam.get('ubicacion'))
            ws.cell(row=fila, column=5, value=cam.get('nombre'))
            if cam.get('telefono'):
                nota = Comment(f"Tel. empresa: {cam['telefono']}", "Bot")
                ws.cell(row=fila, column=5).comment = nota
            ws.cell(row=fila, column=7, value=cam.get('tractora'))
            ws.cell(row=fila, column=8, value=cam.get('remolque'))
            wb.save(self.excel_path)
            wb.close()
            drive_ok = self._sync_to_drive()
            mensaje = f"✅ *¡CAMIONERO AÑADIDO!*\n\n👤 {cam.get('nombre')}\n📱 {cam.get('telefono')}\n🚛 {cam.get('tractora')}\n\n"
            mensaje += "☁️ _Sincronizado con Drive_" if drive_ok else "⚠️ _Guardado local_"
            await update.message.reply_text(mensaje, parse_mode="Markdown", reply_markup=teclado_admin)
        except Exception as e:
            logger.error(f"[GESTIONES] Error: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=teclado_admin)
        context.user_data.clear()
        return ConversationHandler.END
    
    # Volver atrás en camionero
    async def cam_volver_nombre(self, update, context):
        return await self._iniciar_añadir_camionero(update, context)
    async def cam_volver_telefono(self, update, context):
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"📱 *Teléfono:*\n_Anterior: {context.user_data['camionero'].get('telefono', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return CAM_TELEFONO
    async def cam_volver_tractora(self, update, context):
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"🚛 *Tractora:*\n_Anterior: {context.user_data['camionero'].get('tractora', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return CAM_TRACTORA
    async def cam_volver_remolque(self, update, context):
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"📦 *Remolque:*\n_Anterior: {context.user_data['camionero'].get('remolque', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return CAM_REMOLQUE

    # ============================================================
    # AÑADIR VIAJE - FLUJO PRINCIPAL
    # ============================================================
    
    async def _iniciar_añadir_viaje(self, update, context):
        context.user_data['viaje'] = {'cargas': [], 'descargas': []}
        keyboard = [["ZONA NORTE", "ZONA CORTOS NORTE"], ["ZONA RESTO NACIONAL", "ZONA MURCIA"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 1/10\n\n🗺️ *¿Zona?*",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_ZONA

    async def via_zona(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_zona(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_ZONA
        context.user_data['viaje']['zona'] = resultado['valor']
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 2/10\n\n🏢 *¿Cliente?*\n\n_Ejemplo: EROSKI_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_CLIENTE
    
    async def via_cliente(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_cliente(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_CLIENTE
        context.user_data['viaje']['cliente'] = resultado['valor']
        keyboard = [["⏭️ Saltar"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 3/10\n\n🔢 *¿Nº Pedido?*\n\n_Pulsa Saltar si no hay_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_NUM_PEDIDO
    
    async def via_num_pedido(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['viaje']['num_pedido'] = update.message.text.strip()
        return await self._pedir_ref_cliente(update, context)
    
    async def via_saltar_num_pedido(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['viaje']['num_pedido'] = None
        return await self._pedir_ref_cliente(update, context)
    
    async def _pedir_ref_cliente(self, update, context):
        keyboard = [["⏭️ Saltar"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 4/10\n\n🏷️ *¿Ref. Cliente?*\n\n_Pulsa Saltar si no hay_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_REF_CLIENTE
    
    async def via_ref_cliente(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['viaje']['ref_cliente'] = update.message.text.strip().upper()
        return await self._pedir_intercambio(update, context)
    
    async def via_saltar_ref_cliente(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['viaje']['ref_cliente'] = None
        return await self._pedir_intercambio(update, context)
    
    async def _pedir_intercambio(self, update, context):
        keyboard = [["✅ SÍ", "❌ NO"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 5/10\n\n🔄 *¿Intercambio?*",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_INTERCAMBIO
    
    async def via_intercambio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        texto = update.message.text.upper()
        if "SÍ" in texto or "SI" in texto:
            context.user_data['viaje']['intercambio'] = "SI"
        else:
            context.user_data['viaje']['intercambio'] = "NO"
        return await self._pedir_lugar_carga(update, context)
    
    # ============================================================
    # CARGAS - FLUJO CON LOOP (hasta 10)
    # ============================================================
    
    async def _pedir_lugar_carga(self, update, context):
        """Pide la primera carga (o la siguiente si volvemos)"""
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        num_cargas = len(context.user_data['viaje'].get('cargas', []))
        if num_cargas == 0:
            texto = "📦 *NUEVO VIAJE*\n\nPaso 6/10\n\n📍 *¿Lugar de CARGA?*\n\n_Ejemplo: CALAHORRA_"
        else:
            texto = f"📦 *NUEVO VIAJE*\n\n📍 *¿Lugar de CARGA?*\n\n_Ejemplo: CALAHORRA_"
        await update.message.reply_text(
            texto, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_LUGAR_CARGA
    
    async def via_lugar_carga(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Recibe 1ª carga y pregunta si hay más"""
        resultado = validar_lugar_carga(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_LUGAR_CARGA
        
        viaje = context.user_data['viaje']
        # Si es la primera carga, reiniciar lista
        viaje['cargas'] = [resultado['valor']]
        
        return await self._via_preguntar_mas_cargas(update, context)
    
    async def _via_preguntar_mas_cargas(self, update, context):
        """Pregunta si quiere añadir otra carga"""
        viaje = context.user_data['viaje']
        cargas = viaje.get('cargas', [])
        n = len(cargas)
        
        if n >= MAX_CARGAS:
            await update.message.reply_text(
                f"⚠️ Máximo de {MAX_CARGAS} cargas alcanzado. Continuando...",
                parse_mode="Markdown"
            )
            return await self._pedir_lugar_descarga(update, context)
        
        # Mostrar cargas actuales
        lista = "\n".join([f"  📥 {i+1}. *{c}*" for i, c in enumerate(cargas)])
        
        keyboard = [["➕ Sí", "➡️ No"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"📦 *CARGAS* ({n}/{MAX_CARGAS}):\n{lista}\n\n"
            f"¿Añadir otra carga?",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_CARGA_ADICIONAL
    
    async def _handler_carga_adicional(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handler inteligente para Sí/No en cargas adicionales"""
        texto = update.message.text
        interpretacion = interpretar_texto(texto)
        logger.info(f"[GESTIONES] Carga adicional: '{texto}' -> {interpretacion}")
        
        if interpretacion['tipo'] == 'si':
            return await self.via_pedir_carga_adicional(update, context)
        elif interpretacion['tipo'] == 'no':
            return await self.via_saltar_carga_adicional(update, context)
        elif interpretacion['tipo'] == 'cancelar':
            return await self.cancelar(update, context)
        elif interpretacion['tipo'] == 'volver':
            return await self._via_volver_desde_carga_adicional(update, context)
        else:
            await update.message.reply_text("❓ Por favor responde *Sí* o *No*", parse_mode="Markdown")
            return VIA_CARGA_ADICIONAL
    
    async def via_pedir_carga_adicional(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide nueva carga adicional"""
        n = len(context.user_data['viaje'].get('cargas', []))
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"📦 *NUEVA CARGA*\n\n📍 *Carga #{n + 1}:*\n\n_Ejemplo: TUDELA_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_CARGA_ADICIONAL_LUGAR
    
    async def via_carga_adicional_lugar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda carga adicional y pregunta si hay más"""
        resultado = validar_lugar_carga(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_CARGA_ADICIONAL_LUGAR
        
        context.user_data['viaje']['cargas'].append(resultado['valor'])
        
        # ¿Hay espacio para más?
        return await self._via_preguntar_mas_cargas(update, context)
    
    async def via_saltar_carga_adicional(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """No más cargas, pasar a descarga"""
        return await self._pedir_lugar_descarga(update, context)
    
    async def _via_volver_desde_carga_adicional(self, update, context):
        """Volver: si hay cargas adicionales, eliminar la última; si solo 1, volver a pedir carga"""
        cargas = context.user_data['viaje'].get('cargas', [])
        if len(cargas) > 1:
            # Eliminar última carga añadida
            cargas.pop()
            return await self._via_preguntar_mas_cargas(update, context)
        else:
            # Volver a pedir la primera carga
            return await self._pedir_lugar_carga(update, context)
    
    # ============================================================
    # DESCARGAS - FLUJO CON LOOP (hasta 10)
    # ============================================================
    
    async def _pedir_lugar_descarga(self, update, context):
        """Pide la primera descarga"""
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 7/10\n\n📍 *¿Lugar de DESCARGA?*\n\n_Ejemplo: BARCELONA_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_LUGAR_DESCARGA
    
    async def via_lugar_descarga(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Recibe 1ª descarga y pregunta si hay más"""
        resultado = validar_lugar_descarga(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_LUGAR_DESCARGA
        
        viaje = context.user_data['viaje']
        viaje['descargas'] = [resultado['valor']]
        
        return await self._via_preguntar_mas_descargas(update, context)
    
    async def _via_preguntar_mas_descargas(self, update, context):
        """Pregunta si quiere añadir otra descarga"""
        viaje = context.user_data['viaje']
        descargas = viaje.get('descargas', [])
        n = len(descargas)
        
        if n >= MAX_DESCARGAS:
            await update.message.reply_text(
                f"⚠️ Máximo de {MAX_DESCARGAS} descargas alcanzado. Continuando...",
                parse_mode="Markdown"
            )
            return await self._pedir_mercancia(update, context)
        
        lista = "\n".join([f"  📤 {i+1}. *{d}*" for i, d in enumerate(descargas)])
        
        keyboard = [["➕ Sí", "➡️ No"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"📦 *DESCARGAS* ({n}/{MAX_DESCARGAS}):\n{lista}\n\n"
            f"¿Añadir otra descarga?",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_DESCARGA_ADICIONAL
    
    async def _handler_descarga_adicional(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handler inteligente para Sí/No en descargas adicionales"""
        texto = update.message.text
        interpretacion = interpretar_texto(texto)
        
        if interpretacion['tipo'] == 'si':
            return await self.via_pedir_descarga_adicional(update, context)
        elif interpretacion['tipo'] == 'no':
            return await self.via_saltar_descarga_adicional(update, context)
        elif interpretacion['tipo'] == 'cancelar':
            return await self.cancelar(update, context)
        elif interpretacion['tipo'] == 'volver':
            return await self._via_volver_desde_descarga_adicional(update, context)
        else:
            await update.message.reply_text("❓ Por favor responde *Sí* o *No*", parse_mode="Markdown")
            return VIA_DESCARGA_ADICIONAL
    
    async def via_pedir_descarga_adicional(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide nueva descarga adicional"""
        n = len(context.user_data['viaje'].get('descargas', []))
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"📦 *NUEVA DESCARGA*\n\n📍 *Descarga #{n + 1}:*\n\n_Ejemplo: VALENCIA_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_DESCARGA_ADICIONAL_LUGAR
    
    async def via_descarga_adicional_lugar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda descarga adicional y pregunta si hay más"""
        resultado = validar_lugar_descarga(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_DESCARGA_ADICIONAL_LUGAR
        
        context.user_data['viaje']['descargas'].append(resultado['valor'])
        return await self._via_preguntar_mas_descargas(update, context)
    
    async def via_saltar_descarga_adicional(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """No más descargas, pasar a mercancía"""
        return await self._pedir_mercancia(update, context)
    
    async def _via_volver_desde_descarga_adicional(self, update, context):
        """Volver: si hay descargas adicionales, eliminar la última"""
        descargas = context.user_data['viaje'].get('descargas', [])
        if len(descargas) > 1:
            descargas.pop()
            return await self._via_preguntar_mas_descargas(update, context)
        else:
            return await self._pedir_lugar_descarga(update, context)
    
    async def _via_volver_a_cargas(self, update, context):
        """Volver desde descarga a cargas"""
        return await self._via_preguntar_mas_cargas(update, context)
    
    async def _via_volver_a_descargas(self, update, context):
        """Volver desde mercancía a descargas"""
        return await self._via_preguntar_mas_descargas(update, context)

    # ============================================================
    # RESTO DEL FLUJO DE AÑADIR (mercancía, km, precio)
    # ============================================================
    
    async def _pedir_mercancia(self, update, context):
        keyboard = [
            ["REFRIGERADO +2º", "CONGELADO -18º"],
            ["REFRIGERADO +5º", "SECO"],
            ["⬅️ Volver", "❌ Cancelar"]
        ]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 8/10\n\n📦 *¿Mercancía?*",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_MERCANCIA
    
    async def via_mercancia(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_mercancia(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_MERCANCIA
        context.user_data['viaje']['mercancia'] = resultado['valor']
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 9/10\n\n📏 *¿Kilómetros?*\n\n_Ejemplo: 450_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_KM
    
    async def via_km(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_km(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_KM
        context.user_data['viaje']['km'] = resultado['valor']
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "📦 *NUEVO VIAJE*\n\nPaso 10/10\n\n💰 *¿Precio (€)?*\n\n_Ejemplo: 850_",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_PRECIO
    
    async def via_precio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        resultado = validar_precio(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return VIA_PRECIO
        context.user_data['viaje']['precio'] = resultado['valor']
        return await self.via_mostrar_resumen(update, context)
    
    # Volver atrás
    async def via_volver_zona(self, update, context):
        return await self._iniciar_añadir_viaje(update, context)
    async def via_volver_cliente(self, update, context):
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"🏢 *¿Cliente?*\n_Anterior: {context.user_data['viaje'].get('cliente', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return VIA_CLIENTE
    async def via_volver_num_pedido(self, update, context):
        keyboard = [["⏭️ Saltar"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"🔢 *¿Nº Pedido?*\n_Anterior: {context.user_data['viaje'].get('num_pedido', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return VIA_NUM_PEDIDO
    async def via_volver_ref_cliente(self, update, context):
        keyboard = [["⏭️ Saltar"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"🏷️ *¿Ref. Cliente?*\n_Anterior: {context.user_data['viaje'].get('ref_cliente', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return VIA_REF_CLIENTE
    async def via_volver_intercambio(self, update, context):
        keyboard = [["✅ SÍ", "❌ NO"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"🔄 *¿Intercambio?*\n_Anterior: {context.user_data['viaje'].get('intercambio', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return VIA_INTERCAMBIO
    async def via_volver_mercancia(self, update, context):
        keyboard = [["REFRIGERADO +2º", "CONGELADO -18º"], ["REFRIGERADO +5º", "SECO"], ["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"📦 *¿Mercancía?*\n_Anterior: {context.user_data['viaje'].get('mercancia', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return VIA_MERCANCIA
    async def via_volver_km(self, update, context):
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(f"📏 *¿Kilómetros?*\n_Anterior: {context.user_data['viaje'].get('km', '')}_",
            parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))
        return VIA_KM

    # ============================================================
    # RESUMEN Y CONFIRMAR AL AÑADIR
    # ============================================================
    
    async def via_mostrar_resumen(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra resumen del viaje con todas las cargas/descargas"""
        via = context.user_data['viaje']
        _sync_compat(via)
        
        eur_km = via.get('precio', 0) / via.get('km', 1) if via.get('km', 0) > 0 else 0
        
        keyboard = [["✏️ Editar", "✅ Confirmar"], ["❌ Cancelar"]]
        
        cargas = via.get('cargas', [])
        descargas = via.get('descargas', [])
        
        resumen = "📦 *NUEVO VIAJE*\n\n📋 *RESUMEN:*\n━━━━━━━━━━━━━━━━━━━━\n"
        resumen += f"1. 🗺️ Zona: *{via.get('zona', '')}*\n"
        resumen += f"2. 🏢 Cliente: *{via.get('cliente', '')}*\n"
        resumen += f"3. 🔢 Nº Pedido: *{via.get('num_pedido', '-')}*\n"
        resumen += f"4. 🏷️ Ref. Cliente: *{via.get('ref_cliente', '-')}*\n"
        resumen += f"5. 🔄 Intercambio: *{via.get('intercambio', 'NO')}*\n"
        
        # Cargas
        resumen += f"6. 📍 Cargas ({len(cargas)}):\n"
        for i, c in enumerate(cargas):
            resumen += f"   📥 {i+1}. *{c}*\n"
        
        # Descargas
        resumen += f"7. 📍 Descargas ({len(descargas)}):\n"
        for i, d in enumerate(descargas):
            resumen += f"   📤 {i+1}. *{d}*\n"
        
        resumen += f"8. 📦 Mercancía: *{via.get('mercancia', '')}*\n"
        resumen += f"9. 📏 KM: *{via.get('km', 0):.0f}*\n"
        resumen += f"10. 💰 Precio: *{via.get('precio', 0):.0f}€*\n"
        resumen += f"━━━━━━━━━━━━━━━━━━━━\n"
        resumen += f"📊 Ratio: *{eur_km:.2f} €/km*\n\n"
        resumen += "¿Es correcto?"
        
        await update.message.reply_text(
            resumen, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_CONFIRMAR

    async def via_editar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pregunta qué campo editar"""
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            "✏️ *EDITAR VIAJE*\n\n¿Qué campo quieres editar? (1-10)\n\n"
            "1. 🗺️ Zona\n"
            "2. 🏢 Cliente\n"
            "3. 🔢 Nº Pedido\n"
            "4. 🏷️ Ref. Cliente\n"
            "5. 🔄 Intercambio\n"
            "6. 📍 Cargas (gestionar lista)\n"
            "7. 📍 Descargas (gestionar lista)\n"
            "8. 📦 Mercancía\n"
            "9. 📏 Kilómetros\n"
            "10. 💰 Precio",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_EDITAR_CAMPO
    
    async def via_editar_campo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Recibe campo a editar"""
        campo_num = update.message.text.strip()
        
        if campo_num not in self.campos_viaje:
            await update.message.reply_text("⚠️ Introduce un número del 1 al 10:")
            return VIA_EDITAR_CAMPO
        
        campo_key, campo_nombre = self.campos_viaje[campo_num]
        
        # Si es cargas o descargas, abrir submenú
        if campo_key == 'cargas':
            context.user_data['_cargas_retorno'] = 'via_editar'
            return await self._mostrar_cargas_menu(update, context)
        elif campo_key == 'descargas':
            context.user_data['_descargas_retorno'] = 'via_editar'
            return await self._mostrar_descargas_menu(update, context)
        
        context.user_data['editando_campo'] = campo_key
        valor_actual = context.user_data['viaje'].get(campo_key, '-')
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"✏️ *EDITAR {campo_nombre.split(' ', 1)[1].upper()}*\n\n"
            f"Valor actual: *{valor_actual}*\n\nEscribe el nuevo valor:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return VIA_EDITAR_VALOR
    
    async def via_editar_valor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda nuevo valor y vuelve al resumen"""
        campo = context.user_data.get('editando_campo')
        valor = update.message.text.strip()
        tipo = context.user_data.get('tipo', 'viaje')
        
        # Si venimos de editar camionero
        if tipo == 'camionero':
            if campo == 'telefono':
                valor = valor.replace(" ", "").replace("+34", "")
            else:
                valor = valor.upper()
            context.user_data['camionero'][campo] = valor
            return await self.cam_mostrar_resumen(update, context)
        
        # Convertir según el campo
        if campo in ['km', 'precio', 'num_pedido']:
            try:
                valor = float(valor.replace("€", "").replace("km", "").replace("KM", ""))
            except:
                pass
        elif campo == 'intercambio':
            valor = "SI" if "SÍ" in valor.upper() or "SI" in valor.upper() else "NO"
        else:
            valor = valor.upper()
        
        context.user_data['viaje'][campo] = valor
        
        # Si estamos modificando un viaje existente
        if context.user_data.get('modificando'):
            return await self._mod_mostrar_resumen_viaje(update, context)
        
        return await self.via_mostrar_resumen(update, context)
    
    # ============================================================
    # GUARDAR VIAJE EN EXCEL
    # ============================================================
    
    async def via_guardar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda viaje en Excel con todas las cargas/descargas"""
        via = context.user_data['viaje']
        _sync_compat(via)
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            fila = 3
            while ws.cell(row=fila, column=9).value:
                fila += 1
            
            # Cliente
            ws.cell(row=fila, column=9, value=via.get('cliente'))
            if via.get('num_pedido'):
                ws.cell(row=fila, column=10, value=via['num_pedido'])
            if via.get('ref_cliente'):
                ws.cell(row=fila, column=11, value=via['ref_cliente'])
            ws.cell(row=fila, column=12, value=via.get('intercambio', 'NO'))
            
            # Lugar de carga (1ª carga en columna 14)
            cargas = via.get('cargas', [])
            if cargas:
                ws.cell(row=fila, column=14, value=cargas[0])
                # Comentario con TODAS las cargas si hay más de 1
                if len(cargas) > 1:
                    texto_comment = _generar_comentario_cargas(cargas)
                    ws.cell(row=fila, column=14).comment = Comment(texto_comment, "Bot")
            
            # Lugar de descarga (1ª descarga en columna 17)
            descargas = via.get('descargas', [])
            if descargas:
                ws.cell(row=fila, column=17, value=descargas[0])
                if len(descargas) > 1:
                    texto_comment = _generar_comentario_descargas(descargas)
                    ws.cell(row=fila, column=17).comment = Comment(texto_comment, "Bot")
            
            # Mercancía
            ws.cell(row=fila, column=20, value=via.get('mercancia'))
            
            # Precio con formato
            precio_valor = int(via.get('precio', 0))
            celda_precio = ws.cell(row=fila, column=23, value=precio_valor)
            celda_ref_precio = ws.cell(row=3, column=23)
            celda_precio.number_format = celda_ref_precio.number_format
            if celda_ref_precio.font:
                from openpyxl.styles import Font
                celda_precio.font = Font(
                    bold=celda_ref_precio.font.bold,
                    name=celda_ref_precio.font.name,
                    size=celda_ref_precio.font.size
                )
            
            # KM con formato
            km_valor = int(via.get('km', 0))
            celda_km = ws.cell(row=fila, column=24, value=km_valor)
            celda_ref_km = ws.cell(row=3, column=24)
            celda_km.number_format = celda_ref_km.number_format
            
            # €/KM fórmula
            celda_ratio = ws.cell(row=fila, column=25, value=f"=W{fila}/X{fila}")
            celda_ref_ratio = ws.cell(row=3, column=25)
            celda_ratio.number_format = celda_ref_ratio.number_format
            
            # Observaciones con todas las cargas/descargas adicionales
            observaciones = _generar_observaciones(via)
            ws.cell(row=fila, column=28, value=observaciones)
            
            wb.save(self.excel_path)
            wb.close()
            
            # Formatear mensaje resumen
            carga_str = " → ".join(cargas) if cargas else "-"
            descarga_str = " → ".join(descargas) if descargas else "-"
            
            logger.info(f"[GESTIONES] Viaje añadido: {via.get('cliente')} [{carga_str}]→[{descarga_str}]")
            
            drive_ok = self._sync_to_drive()
            
            mensaje = f"✅ *¡VIAJE AÑADIDO!*\n\n"
            mensaje += f"🏢 {via.get('cliente')}\n"
            mensaje += f"📥 Cargas ({len(cargas)}): {', '.join(cargas)}\n"
            mensaje += f"📤 Descargas ({len(descargas)}): {', '.join(descargas)}\n"
            mensaje += f"💰 {via.get('precio', 0):.0f}€\n\n"
            mensaje += "☁️ _Sincronizado con Drive_" if drive_ok else "⚠️ _Guardado local_"
            
            await update.message.reply_text(mensaje, parse_mode="Markdown", reply_markup=teclado_admin)
            
        except Exception as e:
            logger.error(f"[GESTIONES] Error: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=teclado_admin)
        
        context.user_data.clear()
        return ConversationHandler.END

    # ============================================================
    # SUBMENÚ GESTIÓN DE CARGAS (compartido entre add-edit y mod)
    # ============================================================
    
    async def _mostrar_cargas_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra menú de gestión de cargas"""
        viaje = context.user_data.get('viaje', {})
        _inicializar_cargas(viaje)
        cargas = viaje.get('cargas', [])
        n = len(cargas)
        
        mensaje = f"📍 *GESTIÓN DE CARGAS* ({n}/{MAX_CARGAS})\n━━━━━━━━━━━━━━━━━━━━\n"
        if cargas:
            for i, c in enumerate(cargas):
                etiqueta = "Principal" if i == 0 else f"#{i+1}"
                mensaje += f"*{i+1}.* 📥 {etiqueta}: *{c}*\n"
        else:
            mensaje += "_Sin cargas definidas_\n"
        
        mensaje += "━━━━━━━━━━━━━━━━━━━━\n\n"
        mensaje += "Escribe un *número* para editar esa carga"
        
        botones = []
        if n < MAX_CARGAS:
            botones.append("➕ Añadir carga")
        if n > 1:
            botones.append("🗑️ Eliminar carga")
        
        keyboard = []
        if botones:
            keyboard.append(botones)
        keyboard.append(["⬅️ Volver"])
        
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CARGAS_MENU
    
    async def _cargas_seleccionar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Selecciona carga a editar por número"""
        texto = update.message.text.strip()
        cargas = context.user_data.get('viaje', {}).get('cargas', [])
        
        try:
            num = int(texto)
            if num < 1 or num > len(cargas):
                await update.message.reply_text(f"⚠️ Introduce un número del 1 al {len(cargas)}:")
                return CARGAS_MENU
            
            context.user_data['_editando_carga_idx'] = num - 1
            keyboard = [["⬅️ Volver", "❌ Cancelar"]]
            await update.message.reply_text(
                f"✏️ *EDITAR CARGA #{num}*\n\n"
                f"Valor actual: *{cargas[num-1]}*\n\n"
                "Escribe el nuevo lugar de carga:",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return CARGAS_EDITAR
        except ValueError:
            await update.message.reply_text("⚠️ Introduce un número válido:")
            return CARGAS_MENU
    
    async def _cargas_guardar_edicion(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda edición de carga"""
        resultado = validar_lugar_carga(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return CARGAS_EDITAR
        
        idx = context.user_data.get('_editando_carga_idx', 0)
        context.user_data['viaje']['cargas'][idx] = resultado['valor']
        _sync_compat(context.user_data['viaje'])
        
        await update.message.reply_text(f"✅ Carga #{idx+1} actualizada a *{resultado['valor']}*", parse_mode="Markdown")
        return await self._mostrar_cargas_menu(update, context)
    
    async def _cargas_añadir(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Añadir nueva carga"""
        cargas = context.user_data.get('viaje', {}).get('cargas', [])
        if len(cargas) >= MAX_CARGAS:
            await update.message.reply_text(f"⚠️ Máximo de {MAX_CARGAS} cargas alcanzado.")
            return CARGAS_MENU
        
        n = len(cargas)
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"📍 *NUEVA CARGA #{n+1}*\n\nEscribe el lugar de carga:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        context.user_data['_editando_carga_idx'] = -1  # -1 = nueva
        return CARGAS_EDITAR
    
    async def _cargas_pedir_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide qué carga eliminar"""
        cargas = context.user_data.get('viaje', {}).get('cargas', [])
        if len(cargas) <= 1:
            await update.message.reply_text("⚠️ Debe haber al menos 1 carga.")
            return CARGAS_MENU
        
        mensaje = "🗑️ *¿Qué carga eliminar?* (2-" + str(len(cargas)) + ")\n\n"
        mensaje += "_Nota: la carga #1 (principal) no se puede eliminar_\n\n"
        for i, c in enumerate(cargas):
            mensaje += f"*{i+1}.* {c}\n"
        
        keyboard = [["⬅️ Volver"]]
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return CARGAS_ELIMINAR
    
    async def _cargas_confirmar_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Confirma eliminación de carga"""
        cargas = context.user_data.get('viaje', {}).get('cargas', [])
        try:
            num = int(update.message.text.strip())
            if num < 2 or num > len(cargas):
                await update.message.reply_text(f"⚠️ Número del 2 al {len(cargas)}:")
                return CARGAS_ELIMINAR
            
            eliminada = cargas.pop(num - 1)
            _sync_compat(context.user_data['viaje'])
            await update.message.reply_text(f"✅ Carga *{eliminada}* eliminada.", parse_mode="Markdown")
            return await self._mostrar_cargas_menu(update, context)
        except ValueError:
            await update.message.reply_text("⚠️ Introduce un número:")
            return CARGAS_ELIMINAR
    
    async def _cargas_volver(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Volver al menú anterior desde cargas"""
        _sync_compat(context.user_data.get('viaje', {}))
        retorno = context.user_data.get('_cargas_retorno', 'via_editar')
        if retorno == 'mod':
            return await self._mod_mostrar_resumen_viaje(update, context)
        else:
            return await self.via_editar(update, context)

    # ============================================================
    # SUBMENÚ GESTIÓN DE DESCARGAS (compartido)
    # ============================================================
    
    async def _mostrar_descargas_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra menú de gestión de descargas"""
        viaje = context.user_data.get('viaje', {})
        _inicializar_cargas(viaje)
        descargas = viaje.get('descargas', [])
        n = len(descargas)
        
        mensaje = f"📍 *GESTIÓN DE DESCARGAS* ({n}/{MAX_DESCARGAS})\n━━━━━━━━━━━━━━━━━━━━\n"
        if descargas:
            for i, d in enumerate(descargas):
                etiqueta = "Principal" if i == 0 else f"#{i+1}"
                mensaje += f"*{i+1}.* 📤 {etiqueta}: *{d}*\n"
        else:
            mensaje += "_Sin descargas definidas_\n"
        
        mensaje += "━━━━━━━━━━━━━━━━━━━━\n\n"
        mensaje += "Escribe un *número* para editar esa descarga"
        
        botones = []
        if n < MAX_DESCARGAS:
            botones.append("➕ Añadir descarga")
        if n > 1:
            botones.append("🗑️ Eliminar descarga")
        
        keyboard = []
        if botones:
            keyboard.append(botones)
        keyboard.append(["⬅️ Volver"])
        
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return DESCARGAS_MENU
    
    async def _descargas_seleccionar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Selecciona descarga a editar por número"""
        texto = update.message.text.strip()
        descargas = context.user_data.get('viaje', {}).get('descargas', [])
        
        try:
            num = int(texto)
            if num < 1 or num > len(descargas):
                await update.message.reply_text(f"⚠️ Número del 1 al {len(descargas)}:")
                return DESCARGAS_MENU
            
            context.user_data['_editando_descarga_idx'] = num - 1
            keyboard = [["⬅️ Volver", "❌ Cancelar"]]
            await update.message.reply_text(
                f"✏️ *EDITAR DESCARGA #{num}*\n\n"
                f"Valor actual: *{descargas[num-1]}*\n\n"
                "Escribe el nuevo lugar de descarga:",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return DESCARGAS_EDITAR
        except ValueError:
            await update.message.reply_text("⚠️ Introduce un número válido:")
            return DESCARGAS_MENU
    
    async def _descargas_guardar_edicion(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda edición de descarga"""
        resultado = validar_lugar_descarga(update.message.text)
        if not resultado['valido']:
            await update.message.reply_text(resultado['error'], parse_mode="Markdown")
            return DESCARGAS_EDITAR
        
        idx = context.user_data.get('_editando_descarga_idx', 0)
        viaje = context.user_data['viaje']
        
        if idx == -1:
            # Nueva descarga
            viaje['descargas'].append(resultado['valor'])
            await update.message.reply_text(f"✅ Descarga *{resultado['valor']}* añadida.", parse_mode="Markdown")
        else:
            viaje['descargas'][idx] = resultado['valor']
            await update.message.reply_text(f"✅ Descarga #{idx+1} actualizada a *{resultado['valor']}*", parse_mode="Markdown")
        
        _sync_compat(viaje)
        return await self._mostrar_descargas_menu(update, context)
    
    async def _descargas_añadir(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Añadir nueva descarga"""
        descargas = context.user_data.get('viaje', {}).get('descargas', [])
        if len(descargas) >= MAX_DESCARGAS:
            await update.message.reply_text(f"⚠️ Máximo de {MAX_DESCARGAS} descargas alcanzado.")
            return DESCARGAS_MENU
        
        n = len(descargas)
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"📍 *NUEVA DESCARGA #{n+1}*\n\nEscribe el lugar de descarga:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        context.user_data['_editando_descarga_idx'] = -1
        return DESCARGAS_EDITAR
    
    async def _descargas_pedir_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Pide qué descarga eliminar"""
        descargas = context.user_data.get('viaje', {}).get('descargas', [])
        if len(descargas) <= 1:
            await update.message.reply_text("⚠️ Debe haber al menos 1 descarga.")
            return DESCARGAS_MENU
        
        mensaje = "🗑️ *¿Qué descarga eliminar?* (2-" + str(len(descargas)) + ")\n\n"
        mensaje += "_Nota: la descarga #1 (principal) no se puede eliminar_\n\n"
        for i, d in enumerate(descargas):
            mensaje += f"*{i+1}.* {d}\n"
        
        keyboard = [["⬅️ Volver"]]
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return DESCARGAS_ELIMINAR
    
    async def _descargas_confirmar_eliminar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Confirma eliminación de descarga"""
        descargas = context.user_data.get('viaje', {}).get('descargas', [])
        try:
            num = int(update.message.text.strip())
            if num < 2 or num > len(descargas):
                await update.message.reply_text(f"⚠️ Número del 2 al {len(descargas)}:")
                return DESCARGAS_ELIMINAR
            
            eliminada = descargas.pop(num - 1)
            _sync_compat(context.user_data['viaje'])
            await update.message.reply_text(f"✅ Descarga *{eliminada}* eliminada.", parse_mode="Markdown")
            return await self._mostrar_descargas_menu(update, context)
        except ValueError:
            await update.message.reply_text("⚠️ Introduce un número:")
            return DESCARGAS_ELIMINAR
    
    async def _descargas_volver(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Volver al menú anterior desde descargas"""
        _sync_compat(context.user_data.get('viaje', {}))
        retorno = context.user_data.get('_descargas_retorno', 'via_editar')
        if retorno == 'mod':
            return await self._mod_mostrar_resumen_viaje(update, context)
        else:
            return await self.via_editar(update, context)

    # ============================================================
    # MODIFICAR VIAJE EXISTENTE
    # ============================================================
    
    def _get_viajes_sin_asignar(self, limit=10):
        """Obtiene viajes sin conductor asignado del Excel"""
        viajes = []
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            for fila in range(3, 200):
                cliente = ws.cell(row=fila, column=9).value
                # Columna 22 = TRANSPORTISTA del viaje (no columna 5 que es el conductor de la fila)
                transportista = ws.cell(row=fila, column=22).value
                if not cliente:
                    continue
                # Saltar si tiene transportista asignado
                if transportista and str(transportista).strip():
                    continue
                
                obs = ws.cell(row=fila, column=28).value or ''
                zona_match = re.search(r'ZONA:\s*([^|]+)', obs)
                
                viaje = {
                    'fila': fila,
                    'zona': zona_match.group(1).strip() if zona_match else '-',
                    'cliente': str(cliente),
                    'num_pedido': ws.cell(row=fila, column=10).value,
                    'ref_cliente': ws.cell(row=fila, column=11).value,
                    'intercambio': ws.cell(row=fila, column=12).value or 'NO',
                    'lugar_carga': ws.cell(row=fila, column=14).value or '',
                    'lugar_descarga': ws.cell(row=fila, column=17).value or '',
                    'mercancia': ws.cell(row=fila, column=20).value,
                    'km': ws.cell(row=fila, column=24).value or 0,
                    'precio': ws.cell(row=fila, column=23).value or 0,
                    'observaciones': obs,
                    'transportista': '',  # Sin asignar
                }
                
                # Extraer carga_adicional y descarga_adicional de observaciones
                carga2_match = re.search(r'CARGA2:\s*([^|]+)', obs)
                viaje['carga_adicional'] = carga2_match.group(1).strip() if carga2_match else None
                desc2_match = re.search(r'DESCARGA2:\s*([^|]+)', obs)
                viaje['descarga_adicional'] = desc2_match.group(1).strip() if desc2_match else None
                
                # Inicializar listas
                _inicializar_cargas(viaje)
                
                viajes.append(viaje)
                if len(viajes) >= limit:
                    break
            wb.close()
        except Exception as e:
            logger.error(f"[GESTIONES] Error leyendo viajes: {e}")
        return viajes
    
    def _get_camioneros(self, limit=10):
        """Obtiene camioneros del Excel"""
        camioneros = []
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            for fila in range(3, 200):
                nombre = ws.cell(row=fila, column=5).value
                if not nombre:
                    continue
                
                telefono = ''
                cell = ws.cell(row=fila, column=5)
                if cell.comment:
                    tel_match = re.search(r'Tel[^:]*:\s*(\d+)', cell.comment.text)
                    if tel_match:
                        telefono = tel_match.group(1)
                
                camionero = {
                    'fila': fila,
                    'nombre': str(nombre),
                    'telefono': telefono,
                    'tractora': ws.cell(row=fila, column=7).value or '',
                    'remolque': ws.cell(row=fila, column=8).value or '',
                    'ubicacion': ws.cell(row=fila, column=2).value or '',
                }
                camioneros.append(camionero)
                if len(camioneros) >= limit:
                    break
            wb.close()
        except Exception as e:
            logger.error(f"[GESTIONES] Error: {e}")
        return camioneros
    
    async def mod_listar_viajes(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Lista viajes sin asignar con formato mejorado"""
        viajes = self._get_viajes_sin_asignar(10)
        logger.info(f"[GESTIONES] mod_listar_viajes encontró {len(viajes)} viajes")
        
        if not viajes:
            await update.message.reply_text(
                "📦 No hay viajes sin asignar para modificar.",
                reply_markup=ReplyKeyboardRemove()
            )
            return ConversationHandler.END
        
        context.user_data['viajes_lista'] = viajes
        context.user_data['tipo'] = 'viaje'
        
        mensaje = "✏️ *MODIFICAR VIAJE*\n\n📦 *Viajes pendientes de asignar:*\n"
        keyboard = []
        
        for i, v in enumerate(viajes):
            cliente = v.get('cliente', '?')
            carga = v.get('lugar_carga', '?') or '?'
            descarga = v.get('lugar_descarga', '?') or '?'
            mercancia = v.get('mercancia', '-') or '-'
            km = v.get('km', 0) or 0
            precio = v.get('precio', 0) or 0
            
            precio_str = f"{precio:.0f}€" if precio else "S/P"
            
            mensaje += f"\n• 🏢 *{cliente}*\n"
            mensaje += f"  📥 {carga} → 📤 {descarga}\n"
            
            # Botón con info resumida
            btn_text = f"{cliente[:15]} | {carga[:10]}→{descarga[:10]}"
            keyboard.append([InlineKeyboardButton(btn_text, callback_data=f"modviaje_{i}")])
        
        keyboard.append([InlineKeyboardButton("❌ Cancelar", callback_data="mod_cancelar")])
        
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MOD_ELEGIR_VIAJE
    
    async def mod_elegir_viaje_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Selecciona viaje a modificar (callback)"""
        query = update.callback_query
        await query.answer()
        
        viajes = context.user_data.get('viajes_lista', [])
        idx = int(query.data.replace("modviaje_", ""))
        
        if idx < 0 or idx >= len(viajes):
            await query.answer("Viaje no válido", show_alert=True)
            return MOD_ELEGIR_VIAJE
        
        viaje = viajes[idx]
        context.user_data['viaje'] = viaje.copy()
        context.user_data['viaje_fila'] = viaje['fila']
        context.user_data['modificando'] = True
        context.user_data['tipo'] = 'viaje'
        
        # Asegurar listas
        _inicializar_cargas(context.user_data['viaje'])
        
        logger.info(f"[GESTIONES] Viaje seleccionado fila {viaje['fila']}: {viaje.get('cliente')}")
        return await self._mod_mostrar_resumen_viaje_callback(update, context)
    
    async def mod_elegir_viaje(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Selecciona viaje a modificar"""
        texto = update.message.text.strip()
        viajes = context.user_data.get('viajes_lista', [])
        
        if not viajes:
            return await self.mod_listar_viajes(update, context)
        
        try:
            num = int(texto)
            if num < 1 or num > len(viajes):
                await update.message.reply_text(f"⚠️ Introduce un número del 1 al {len(viajes)}:")
                return MOD_ELEGIR_VIAJE
            
            viaje = viajes[num - 1]
            context.user_data['viaje'] = viaje.copy()
            context.user_data['viaje_fila'] = viaje['fila']
            context.user_data['modificando'] = True
            context.user_data['tipo'] = 'viaje'
            
            # Asegurar listas
            _inicializar_cargas(context.user_data['viaje'])
            
            logger.info(f"[GESTIONES] Viaje seleccionado fila {viaje['fila']}: {viaje.get('cliente')}")
            return await self._mod_mostrar_resumen_viaje(update, context)
            
        except ValueError:
            await update.message.reply_text(f"⚠️ Introduce un número del 1 al {len(viajes)}:")
            return MOD_ELEGIR_VIAJE
    
    def _generar_resumen_viaje_keyboard(self, via: dict):
        """Genera el mensaje y keyboard para el resumen del viaje"""
        _inicializar_cargas(via)
        
        cargas = via.get('cargas', [])
        descargas = via.get('descargas', [])
        
        mensaje = "✏️ *MODIFICAR VIAJE*\n\n📋 *DATOS ACTUALES:*\n━━━━━━━━━━━━━━━━━━━━\n"
        mensaje += f"🗺️ Zona: *{via.get('zona', '-')}*\n"
        mensaje += f"🏢 Cliente: *{via.get('cliente', '-')}*\n"
        mensaje += f"🔢 Nº Pedido: *{via.get('num_pedido', '-') or '-'}*\n"
        mensaje += f"🏷️ Ref. Cliente: *{via.get('ref_cliente', '-') or '-'}*\n"
        mensaje += f"🔄 Intercambio: *{via.get('intercambio', 'NO')}*\n"
        
        # Cargas
        mensaje += f"📥 Cargas ({len(cargas)}):"
        if cargas:
            for i, c in enumerate(cargas):
                mensaje += f" {c}{',' if i < len(cargas)-1 else ''}"
        else:
            mensaje += " _-_"
        mensaje += "\n"
        
        # Descargas
        mensaje += f"📤 Descargas ({len(descargas)}):"
        if descargas:
            for i, d in enumerate(descargas):
                mensaje += f" {d}{',' if i < len(descargas)-1 else ''}"
        else:
            mensaje += " _-_"
        mensaje += "\n"
        
        mensaje += f"📦 Mercancía: *{via.get('mercancia', '-') or '-'}*\n"
        
        km = via.get('km', 0) or 0
        precio = via.get('precio', 0) or 0
        km_str = f"{km:.0f}" if km else "-"
        precio_str = f"{precio:.0f}€" if precio else "-"
        mensaje += f"📏 KM: *{km_str}*\n"
        mensaje += f"💰 Precio: *{precio_str}*\n"
        
        transportista = via.get('transportista', '') or ''
        transp_str = transportista if transportista else "⚠️ SIN ASIGNAR"
        mensaje += f"🚛 Conductor: *{transp_str}*\n"
        
        mensaje += "━━━━━━━━━━━━━━━━━━━━\n\n"
        mensaje += "Pulsa un campo para editarlo:"
        
        # Keyboard con botones para cada campo
        keyboard = [
            [
                InlineKeyboardButton("🗺️ Zona", callback_data="modcampo_1"),
                InlineKeyboardButton("🏢 Cliente", callback_data="modcampo_2"),
            ],
            [
                InlineKeyboardButton("🔢 Nº Pedido", callback_data="modcampo_3"),
                InlineKeyboardButton("🏷️ Ref. Cliente", callback_data="modcampo_4"),
            ],
            [
                InlineKeyboardButton("🔄 Intercambio", callback_data="modcampo_5"),
                InlineKeyboardButton("📦 Mercancía", callback_data="modcampo_8"),
            ],
            [
                InlineKeyboardButton("📥 Cargas", callback_data="modcampo_6"),
                InlineKeyboardButton("📤 Descargas", callback_data="modcampo_7"),
            ],
            [
                InlineKeyboardButton("📏 KM", callback_data="modcampo_9"),
                InlineKeyboardButton("💰 Precio", callback_data="modcampo_10"),
            ],
            [
                InlineKeyboardButton("🚛 Asignar Conductor", callback_data="modcampo_11"),
            ],
            [
                InlineKeyboardButton("✅ Guardar cambios", callback_data="mod_guardar"),
            ],
            [
                InlineKeyboardButton("⬅️ Volver", callback_data="mod_volver_lista"),
                InlineKeyboardButton("❌ Cancelar", callback_data="mod_cancelar"),
            ],
        ]
        
        return mensaje, InlineKeyboardMarkup(keyboard)
    
    async def _mod_mostrar_resumen_viaje(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra resumen del viaje con InlineKeyboard"""
        via = context.user_data.get('viaje', {})
        mensaje, keyboard = self._generar_resumen_viaje_keyboard(via)
        
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=keyboard
        )
        return MOD_CAMPO
    
    async def _mod_mostrar_resumen_viaje_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra resumen del viaje con InlineKeyboard (desde callback)"""
        query = update.callback_query
        via = context.user_data.get('viaje', {})
        mensaje, keyboard = self._generar_resumen_viaje_keyboard(via)
        
        await query.edit_message_text(
            mensaje, parse_mode="Markdown",
            reply_markup=keyboard
        )
        return MOD_CAMPO
    
    async def mod_elegir_campo_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Elige campo a modificar (callback)"""
        query = update.callback_query
        await query.answer()
        
        campo_num = query.data.replace("modcampo_", "")
        tipo = context.user_data.get('tipo', 'viaje')
        
        if tipo == 'viaje':
            if campo_num not in self.campos_viaje:
                await query.answer("Campo no válido", show_alert=True)
                return MOD_CAMPO
            
            campo_key, campo_nombre = self.campos_viaje[campo_num]
            
            # Si es cargas o descargas, abrir submenú (por ahora mensaje)
            if campo_key == 'cargas':
                context.user_data['_cargas_retorno'] = 'mod'
                await query.message.reply_text(
                    "📥 *EDITAR CARGAS*\n\nEscribe las cargas separadas por coma:\nEj: BILBAO, MADRID, VALENCIA",
                    parse_mode="Markdown",
                    reply_markup=ReplyKeyboardMarkup([["⬅️ Volver", "❌ Cancelar"]], resize_keyboard=True)
                )
                context.user_data['editando_campo'] = 'cargas'
                return MOD_VALOR
            elif campo_key == 'descargas':
                context.user_data['_descargas_retorno'] = 'mod'
                await query.message.reply_text(
                    "📤 *EDITAR DESCARGAS*\n\nEscribe las descargas separadas por coma:\nEj: SEVILLA, MALAGA",
                    parse_mode="Markdown",
                    reply_markup=ReplyKeyboardMarkup([["⬅️ Volver", "❌ Cancelar"]], resize_keyboard=True)
                )
                context.user_data['editando_campo'] = 'descargas'
                return MOD_VALOR
            elif campo_key == 'transportista':
                # Mostrar lista de conductores disponibles
                return await self._mostrar_conductores_para_asignar(update, context)
            
            context.user_data['editando_campo'] = campo_key
            valor_actual = context.user_data.get('viaje', {}).get(campo_key, '-')
            
            await query.message.reply_text(
                f"✏️ *EDITAR {campo_nombre.split(' ', 1)[1].upper()}*\n\n"
                f"Valor actual: *{valor_actual or '-'}*\n\nEscribe el nuevo valor:",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup([["⬅️ Volver", "❌ Cancelar"]], resize_keyboard=True)
            )
            return MOD_VALOR
        
        return MOD_CAMPO
    
    async def _mostrar_conductores_para_asignar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra lista de conductores para asignar al viaje"""
        query = update.callback_query
        
        # Obtener conductores de la BD
        conductores = self._get_conductores_disponibles()
        
        if not conductores:
            await query.answer("No hay conductores disponibles", show_alert=True)
            return MOD_CAMPO
        
        context.user_data['conductores_lista'] = conductores
        
        mensaje = "🚛 *ASIGNAR CONDUCTOR*\n\nSelecciona el conductor:\n"
        keyboard = []
        
        for i, c in enumerate(conductores):
            nombre = c.get('nombre', '?')
            ubicacion = c.get('ubicacion', '-')
            tractora = c.get('tractora', '-')
            
            mensaje += f"\n• *{nombre}*\n  📍 {ubicacion} | 🚛 {tractora}\n"
            
            btn_text = f"{nombre[:20]}"
            keyboard.append([InlineKeyboardButton(btn_text, callback_data=f"asignar_{i}")])
        
        keyboard.append([InlineKeyboardButton("⬅️ Volver", callback_data="mod_volver_resumen")])
        keyboard.append([InlineKeyboardButton("❌ Cancelar", callback_data="mod_cancelar")])
        
        await query.edit_message_text(
            mensaje, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
        return MOD_CAMPO
    
    def _get_conductores_disponibles(self):
        """Obtiene conductores disponibles del Excel"""
        conductores = []
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            nombres_vistos = set()
            
            for fila in range(3, 200):
                nombre = ws.cell(row=fila, column=5).value
                if not nombre or nombre in nombres_vistos:
                    continue
                nombres_vistos.add(nombre)
                
                conductor = {
                    'nombre': str(nombre),
                    'tractora': ws.cell(row=fila, column=7).value or '-',
                    'ubicacion': ws.cell(row=fila, column=2).value or '-',
                }
                conductores.append(conductor)
            wb.close()
        except Exception as e:
            logger.error(f"[GESTIONES] Error obteniendo conductores: {e}")
        return conductores
    
    async def mod_asignar_conductor_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Asigna un conductor al viaje"""
        query = update.callback_query
        await query.answer()
        
        conductores = context.user_data.get('conductores_lista', [])
        idx = int(query.data.replace("asignar_", ""))
        
        if idx < 0 or idx >= len(conductores):
            await query.answer("Conductor no válido", show_alert=True)
            return MOD_CAMPO
        
        conductor = conductores[idx]
        viaje = context.user_data.get('viaje', {})
        
        # Guardar transportista en el viaje
        viaje['transportista'] = conductor['nombre']
        viaje['tractora_asignada'] = conductor.get('tractora', '')
        context.user_data['viaje'] = viaje
        context.user_data['conductor_asignado'] = conductor
        
        logger.info(f"[GESTIONES] Conductor {conductor['nombre']} asignado a viaje {viaje.get('cliente')}")
        
        # Volver al resumen
        return await self._mod_mostrar_resumen_viaje_callback(update, context)
    
    async def mod_guardar_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda cambios (callback)"""
        query = update.callback_query
        await query.answer()
        
        # Reutilizar lógica existente
        update.message = query.message
        return await self.mod_guardar(update, context)
    
    async def mod_volver_lista_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Vuelve a la lista de viajes (callback)"""
        query = update.callback_query
        await query.answer()
        
        update.message = query.message
        return await self.mod_listar_viajes(update, context)
    
    async def cancelar_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancela desde callback"""
        query = update.callback_query
        await query.answer()
        context.user_data.clear()
        
        await query.edit_message_text("❌ Operación cancelada.")
        await query.message.reply_text(
            "¿Qué más necesitas?",
            reply_markup=teclado_admin
        )
        return ConversationHandler.END

    async def mod_elegir_campo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Elige campo a modificar"""
        campo_num = update.message.text.strip()
        tipo = context.user_data.get('tipo', 'viaje')
        
        if tipo == 'camionero':
            if campo_num not in self.campos_camionero:
                await update.message.reply_text("⚠️ Introduce un número del 1 al 5:")
                return MOD_CAMPO
            campo_key, campo_nombre = self.campos_camionero[campo_num]
            context.user_data['editando_campo'] = campo_key
            valor_actual = context.user_data.get('camionero', {}).get(campo_key, '-')
        else:
            if campo_num not in self.campos_viaje:
                await update.message.reply_text("⚠️ Introduce un número del 1 al 10:")
                return MOD_CAMPO
            
            campo_key, campo_nombre = self.campos_viaje[campo_num]
            
            # Si es cargas o descargas, abrir submenú
            if campo_key == 'cargas':
                context.user_data['_cargas_retorno'] = 'mod'
                return await self._mostrar_cargas_menu(update, context)
            elif campo_key == 'descargas':
                context.user_data['_descargas_retorno'] = 'mod'
                return await self._mostrar_descargas_menu(update, context)
            
            context.user_data['editando_campo'] = campo_key
            valor_actual = context.user_data.get('viaje', {}).get(campo_key, '-')
        
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            f"✏️ *EDITAR {campo_nombre.split(' ', 1)[1].upper()}*\n\n"
            f"Valor actual: *{valor_actual}*\n\nEscribe el nuevo valor:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MOD_VALOR
    
    async def mod_nuevo_valor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda nuevo valor y muestra resumen"""
        campo = context.user_data.get('editando_campo')
        valor = update.message.text.strip()
        tipo = context.user_data.get('tipo', 'viaje')
        
        if campo in ['km', 'precio']:
            try:
                valor = float(valor.replace("€", "").replace("km", "").replace("KM", ""))
            except:
                pass
        elif campo == 'num_pedido':
            try:
                valor = float(valor)
            except:
                pass
        elif campo == 'intercambio':
            valor = "SI" if "SÍ" in valor.upper() or "SI" in valor.upper() else "NO"
        elif campo == 'telefono':
            valor = valor.replace(" ", "").replace("+34", "")
        else:
            valor = valor.upper()
        
        if tipo == 'camionero':
            if 'camionero' not in context.user_data:
                context.user_data['camionero'] = {}
            context.user_data['camionero'][campo] = valor
            logger.info(f"[GESTIONES] Camionero campo '{campo}' actualizado a: {valor}")
            return await self._mod_mostrar_resumen_camionero(update, context)
        else:
            if 'viaje' not in context.user_data:
                context.user_data['viaje'] = {'cargas': [], 'descargas': []}
            context.user_data['viaje'][campo] = valor
            logger.info(f"[GESTIONES] Viaje campo '{campo}' actualizado a: {valor}")
            return await self._mod_mostrar_resumen_viaje(update, context)
    
    async def _mod_volver_resumen(self, update, context):
        """Volver al resumen desde edición de valor"""
        tipo = context.user_data.get('tipo', 'viaje')
        if tipo == 'camionero':
            return await self._mod_mostrar_resumen_camionero(update, context)
        return await self._mod_mostrar_resumen_viaje(update, context)
    
    async def mod_volver_lista(self, update, context):
        tipo = context.user_data.get('tipo', 'viaje')
        if tipo == 'camionero':
            return await self.mod_listar_camioneros(update, context)
        return await self.mod_listar_viajes(update, context)
    
    async def mod_guardar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda cambios del viaje/camionero modificado en Excel"""
        tipo = context.user_data.get('tipo', 'viaje')
        
        if tipo == 'camionero':
            return await self._mod_guardar_camionero(update, context)
        
        via = context.user_data.get('viaje', {})
        fila = context.user_data.get('viaje_fila')
        conductor_asignado = context.user_data.get('conductor_asignado')
        _sync_compat(via)
        
        if not fila:
            await update.message.reply_text("❌ Error: no se encontró la fila del viaje.")
            return ConversationHandler.END
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            ws.cell(row=fila, column=9, value=via.get('cliente'))
            ws.cell(row=fila, column=10, value=via.get('num_pedido'))
            ws.cell(row=fila, column=11, value=via.get('ref_cliente'))
            ws.cell(row=fila, column=12, value=via.get('intercambio', 'NO'))
            
            # Cargas
            cargas = via.get('cargas', [])
            if cargas:
                ws.cell(row=fila, column=14, value=cargas[0])
                # Limpiar comentario anterior
                ws.cell(row=fila, column=14).comment = None
                if len(cargas) > 1:
                    ws.cell(row=fila, column=14).comment = Comment(_generar_comentario_cargas(cargas), "Bot")
            
            # Descargas
            descargas = via.get('descargas', [])
            if descargas:
                ws.cell(row=fila, column=17, value=descargas[0])
                ws.cell(row=fila, column=17).comment = None
                if len(descargas) > 1:
                    ws.cell(row=fila, column=17).comment = Comment(_generar_comentario_descargas(descargas), "Bot")
            
            ws.cell(row=fila, column=20, value=via.get('mercancia'))
            
            precio = via.get('precio', 0)
            if precio:
                ws.cell(row=fila, column=23, value=int(precio))
            km = via.get('km', 0)
            if km:
                ws.cell(row=fila, column=24, value=int(km))
            
            # Transportista (columna 22)
            transportista = via.get('transportista', '')
            if transportista:
                ws.cell(row=fila, column=22, value=transportista)
            
            # Observaciones
            observaciones = _generar_observaciones(via)
            ws.cell(row=fila, column=28, value=observaciones)
            
            wb.save(self.excel_path)
            wb.close()
            
            drive_ok = self._sync_to_drive()
            
            carga_str = ", ".join(cargas) if cargas else "-"
            descarga_str = ", ".join(descargas) if descargas else "-"
            
            mensaje = f"✅ *¡VIAJE MODIFICADO!*\n\n"
            mensaje += f"🏢 {via.get('cliente')}\n"
            mensaje += f"📥 Cargas: {carga_str}\n"
            mensaje += f"📤 Descargas: {descarga_str}\n"
            if transportista:
                mensaje += f"🚛 Conductor: {transportista}\n"
            mensaje += f"💰 {via.get('precio', 0):.0f}€\n\n"
            mensaje += "☁️ _Sincronizado con Drive_" if drive_ok else "⚠️ _Guardado local_"
            
            await update.message.reply_text(mensaje, parse_mode="Markdown", reply_markup=teclado_admin)
            
            # Notificar al conductor si se asignó uno nuevo
            if conductor_asignado and transportista:
                await self._notificar_conductor_asignacion(via, conductor_asignado)
            
        except Exception as e:
            logger.error(f"[GESTIONES] Error guardando: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=teclado_admin)
        
        context.user_data.clear()
        return ConversationHandler.END
    
    async def _notificar_conductor_asignacion(self, viaje: dict, conductor: dict):
        """Notifica al conductor que se le ha asignado un viaje"""
        try:
            from notificaciones_viajes import obtener_notificador
            notif = obtener_notificador()
            
            if notif:
                # Preparar datos para notificación
                asignacion = {
                    'conductor_asignado': conductor.get('nombre'),
                    'cliente': viaje.get('cliente'),
                    'lugar_carga': viaje.get('cargas', ['?'])[0] if viaje.get('cargas') else viaje.get('lugar_carga', '?'),
                    'lugar_entrega': viaje.get('descargas', ['?'])[0] if viaje.get('descargas') else viaje.get('lugar_descarga', '?'),
                    'mercancia': viaje.get('mercancia', '-'),
                    'km': viaje.get('km', 0),
                    'intercambio': viaje.get('intercambio', 'NO'),
                }
                
                # Usar verificar_y_notificar que detectará el nuevo viaje
                await notif.verificar_y_notificar()
                logger.info(f"[GESTIONES] Notificación enviada a {conductor.get('nombre')}")
        except Exception as e:
            logger.warning(f"[GESTIONES] Error notificando conductor: {e}")

    # ============================================================
    # MODIFICAR CAMIONERO EXISTENTE
    # ============================================================
    
    async def mod_listar_camioneros(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Lista camioneros"""
        camioneros = self._get_camioneros(10)
        
        if not camioneros:
            await update.message.reply_text("🚛 No hay camioneros registrados.", reply_markup=ReplyKeyboardRemove())
            return ConversationHandler.END
        
        context.user_data['camioneros_lista'] = camioneros
        context.user_data['tipo'] = 'camionero'
        
        mensaje = "✏️ *MODIFICAR CAMIONERO*\n\n🚛 Camioneros:\n\n"
        for i, c in enumerate(camioneros, 1):
            mensaje += f"*{i}.* {c.get('nombre', '?')} | {c.get('tractora', '?')}\n"
        
        mensaje += f"\n¿Cuál quieres modificar? (1-{len(camioneros)})"
        keyboard = [["⬅️ Volver", "❌ Cancelar"]]
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MOD_ELEGIR_CAMIONERO
    
    async def mod_elegir_camionero(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Selecciona camionero"""
        texto = update.message.text.strip()
        camioneros = context.user_data.get('camioneros_lista', [])
        
        if not camioneros:
            return await self.mod_listar_camioneros(update, context)
        
        try:
            num = int(texto)
            if num < 1 or num > len(camioneros):
                await update.message.reply_text(f"⚠️ Número del 1 al {len(camioneros)}:")
                return MOD_ELEGIR_CAMIONERO
            
            camionero = camioneros[num - 1]
            context.user_data['camionero'] = camionero.copy()
            context.user_data['camionero_fila'] = camionero['fila']
            context.user_data['modificando'] = True
            context.user_data['tipo'] = 'camionero'
            
            return await self._mod_mostrar_resumen_camionero(update, context)
            
        except ValueError:
            await update.message.reply_text(f"⚠️ Número del 1 al {len(camioneros)}:")
            return MOD_ELEGIR_CAMIONERO
    
    async def _mod_mostrar_resumen_camionero(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        cam = context.user_data.get('camionero', {})
        keyboard = [["✅ Guardar cambios"], ["⬅️ Volver", "❌ Cancelar"]]
        mensaje = "✏️ *MODIFICAR CAMIONERO*\n\n📋 *DATOS ACTUALES:*\n━━━━━━━━━━━━━━━━━━━━\n"
        mensaje += f"1. 👤 Nombre: *{cam.get('nombre', '-')}*\n"
        mensaje += f"2. 📱 Teléfono: *{cam.get('telefono', '-')}*\n"
        mensaje += f"3. 🚛 Tractora: *{cam.get('tractora', '-')}*\n"
        mensaje += f"4. 📦 Remolque: *{cam.get('remolque', '-')}*\n"
        mensaje += f"5. 📍 Ubicación: *{cam.get('ubicacion', '-')}*\n"
        mensaje += "━━━━━━━━━━━━━━━━━━━━\n\n"
        mensaje += "¿Qué campo editar? (1-5)\nO *Guardar cambios* para terminar."
        await update.message.reply_text(
            mensaje, parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return MOD_CAMPO
    
    async def _mod_guardar_camionero(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Guarda cambios del camionero en Excel"""
        cam = context.user_data.get('camionero', {})
        fila = context.user_data.get('camionero_fila')
        
        if not fila:
            await update.message.reply_text("❌ Error: no se encontró la fila.")
            return ConversationHandler.END
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            ws.cell(row=fila, column=2, value=cam.get('ubicacion'))
            ws.cell(row=fila, column=5, value=cam.get('nombre'))
            if cam.get('telefono'):
                ws.cell(row=fila, column=5).comment = Comment(f"Tel. empresa: {cam['telefono']}", "Bot")
            ws.cell(row=fila, column=7, value=cam.get('tractora'))
            ws.cell(row=fila, column=8, value=cam.get('remolque'))
            
            wb.save(self.excel_path)
            wb.close()
            
            drive_ok = self._sync_to_drive()
            
            mensaje = f"✅ *¡CAMIONERO MODIFICADO!*\n\n👤 {cam.get('nombre')}\n🚛 {cam.get('tractora')}\n\n"
            mensaje += "☁️ _Sincronizado con Drive_" if drive_ok else "⚠️ _Guardado local_"
            
            await update.message.reply_text(mensaje, parse_mode="Markdown", reply_markup=teclado_admin)
            
        except Exception as e:
            logger.error(f"[GESTIONES] Error: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=teclado_admin)
        
        context.user_data.clear()
        return ConversationHandler.END

    # ============================================================
    # UTILIDADES
    # ============================================================
    
    def _sync_to_drive(self):
        """Sincroniza con Google Drive"""
        try:
            if self.subir_drive:
                self.subir_drive()
                return True
        except Exception as e:
            logger.warning(f"[GESTIONES] Error subiendo a Drive: {e}")
        return False
