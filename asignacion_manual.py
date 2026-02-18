"""
ASIGNACIÓN MANUAL DE VIAJES v1.1
==================================
Panel interactivo para que el admin asigne viajes manualmente.
Conductores ordenados por cercanía GPS al punto de carga.

FLUJO:
1. Admin pulsa "📋 Asignar viaje" → ve lista de viajes sin conductor
2. Pincha en un viaje → ve conductores de la misma zona ORDENADOS POR CERCANÍA
3. Conductores libres (🟢) y ocupados (🔶) con distancia en km
4. Pincha en un conductor → confirmación
5. Confirma → asigna en BD + Excel + notifica al conductor

PRIORIDAD DE POSICIÓN DEL CONDUCTOR:
1. GPS real (Movildata → última posición de la tractora)
2. Última descarga asignada (si tiene viaje, estará cerca de ahí)
3. Ubicación base (campo 'ubicacion' en BD)

INTEGRACIÓN en bot_transporte.py:
    from asignacion_manual import AsignacionManual

    asignacion_manual = AsignacionManual(
        db_path=config.DB_PATH,
        excel_path=config.EXCEL_EMPRESA,
        on_excel_updated=subir_excel_a_drive if config.DRIVE_ENABLED else None,
        movildata_api=movildata_api,
    )
    asignacion_manual.registrar_handlers(app)
"""

import sqlite3
import logging
import math
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes,
    CallbackQueryHandler,
)

logger = logging.getLogger(__name__)

# Importar coordenadas del asignador (149 ubicaciones conocidas)
try:
    from asignador_viajes import COORDENADAS_LUGARES
except ImportError:
    COORDENADAS_LUGARES = {}

# Columna TRANSPORTISTA en Excel (openpyxl, 1-indexed)
COL_TRANSPORTISTA = 22  # Columna V

# Máximo de viajes/conductores por página
PAGE_SIZE_VIAJES = 8
PAGE_SIZE_CONDUCTORES = 10


class AsignacionManual:
    """
    Panel interactivo de asignación manual de viajes.
    """

    def __init__(
        self,
        db_path: str,
        excel_path: str = None,
        on_excel_updated: callable = None,
        admin_ids: list = None,
        movildata_api=None,
    ):
        self.db_path = db_path
        self.excel_path = excel_path
        self.on_excel_updated = on_excel_updated
        self.admin_ids = admin_ids or []
        self.movildata = movildata_api

        # Cache temporal de conductores por sesión de asignación
        # {chat_id: {viaje_id: [lista_conductores]}}
        self._cache_conductores = {}

        logger.info("[ASIGNACIÓN] Manual v1.1 inicializada (con ordenación por cercanía)")

    # ================================================================
    # REGISTRAR HANDLERS
    # ================================================================
    def registrar_handlers(self, app):
        """Registra los CallbackQueryHandlers en la app."""
        app.add_handler(CallbackQueryHandler(self._cb_listar_viajes, pattern=r"^asgn:list"))
        app.add_handler(CallbackQueryHandler(self._cb_ver_viaje, pattern=r"^asgn:v_"))
        app.add_handler(CallbackQueryHandler(self._cb_seleccionar_conductor, pattern=r"^asgn:c_"))
        app.add_handler(CallbackQueryHandler(self._cb_confirmar, pattern=r"^asgn:ok_"))
        app.add_handler(CallbackQueryHandler(self._cb_cancelar, pattern=r"^asgn:cancel$"))
        logger.info("[ASIGNACIÓN] Handlers registrados")

    # ================================================================
    # PUNTO DE ENTRADA (llamado desde bot_transporte.py)
    # ================================================================
    async def inicio(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Muestra la lista de viajes sin asignar. Se llama desde el handler de botón."""
        await self._mostrar_viajes_sin_asignar(update, context, editar=False)

    # ================================================================
    # CALLBACKS
    # ================================================================
    async def _cb_listar_viajes(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Callback: volver a la lista de viajes."""
        query = update.callback_query
        await query.answer()

        # Extraer página si viene en el callback
        data = query.data  # "asgn:list" o "asgn:list_2" (página 2)
        page = 0
        if "_" in data.split("list")[1] if "list" in data else "":
            try:
                page = int(data.split("_")[-1])
            except ValueError:
                page = 0

        await self._mostrar_viajes_sin_asignar(update, context, editar=True, page=page)

    async def _cb_ver_viaje(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Callback: admin pincha en un viaje → muestra conductores disponibles."""
        query = update.callback_query
        await query.answer()

        # Extraer viaje_id: "asgn:v_5317"
        try:
            viaje_id = int(query.data.split("_")[1])
        except (IndexError, ValueError):
            await query.edit_message_text("❌ Error: viaje no válido")
            return

        await self._mostrar_conductores_para_viaje(update, context, viaje_id)

    async def _cb_seleccionar_conductor(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Callback: admin pincha en un conductor → pedir confirmación."""
        query = update.callback_query
        await query.answer()

        # Extraer: "asgn:c_5317_3" (viaje_id, índice_conductor)
        try:
            parts = query.data.split("_")
            viaje_id = int(parts[1])
            idx_conductor = int(parts[2])
        except (IndexError, ValueError):
            await query.edit_message_text("❌ Error: datos no válidos")
            return

        await self._mostrar_confirmacion(update, context, viaje_id, idx_conductor)

    async def _cb_confirmar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Callback: confirmar asignación."""
        query = update.callback_query
        await query.answer()

        # "asgn:ok_5317_3"
        try:
            parts = query.data.split("_")
            viaje_id = int(parts[1])
            idx_conductor = int(parts[2])
        except (IndexError, ValueError):
            await query.edit_message_text("❌ Error: datos no válidos")
            return

        await self._ejecutar_asignacion(update, context, viaje_id, idx_conductor)

    async def _cb_cancelar(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Callback: cancelar y volver a la lista."""
        query = update.callback_query
        await query.answer()
        await self._mostrar_viajes_sin_asignar(update, context, editar=True)

    # ================================================================
    # PANTALLA 1: LISTA DE VIAJES SIN ASIGNAR
    # ================================================================
    async def _mostrar_viajes_sin_asignar(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE,
        editar: bool = False, page: int = 0
    ):
        viajes = self._obtener_viajes_sin_asignar()

        if not viajes:
            texto = "✅ No hay viajes pendientes de asignar."
            if editar:
                await update.callback_query.edit_message_text(texto)
            else:
                await update.message.reply_text(texto)
            return

        total = len(viajes)
        inicio = page * PAGE_SIZE_VIAJES
        fin = min(inicio + PAGE_SIZE_VIAJES, total)
        viajes_pagina = viajes[inicio:fin]

        texto = f"📦 *VIAJES SIN ASIGNAR* ({total})\n"
        texto += f"Página {page + 1}/{(total - 1) // PAGE_SIZE_VIAJES + 1}\n\n"
        texto += "Selecciona un viaje para asignar conductor:\n"

        botones = []
        for v in viajes_pagina:
            cliente = v['cliente'] or '?'
            carga = (v['lugar_carga'] or '?')[:12]
            descarga = (v['lugar_entrega'] or '?')[:12]
            precio = f"{v['precio']}€" if v.get('precio') else ''
            mercancia = (v['mercancia'] or '')[:8]

            label = f"📦 {cliente} | {carga}→{descarga} | {precio}"
            if mercancia:
                label += f" | {mercancia}"

            # Limitar a 64 bytes
            callback = f"asgn:v_{v['id']}"
            botones.append([InlineKeyboardButton(label[:60], callback_data=callback)])

        # Paginación
        nav = []
        if page > 0:
            nav.append(InlineKeyboardButton("◀️ Anterior", callback_data=f"asgn:list_{page - 1}"))
        if fin < total:
            nav.append(InlineKeyboardButton("Siguiente ▶️", callback_data=f"asgn:list_{page + 1}"))
        if nav:
            botones.append(nav)

        markup = InlineKeyboardMarkup(botones)

        if editar:
            await update.callback_query.edit_message_text(
                texto, reply_markup=markup, parse_mode='Markdown'
            )
        else:
            await update.message.reply_text(
                texto, reply_markup=markup, parse_mode='Markdown'
            )

    # ================================================================
    # PANTALLA 2: DETALLE VIAJE + CONDUCTORES DISPONIBLES
    # ================================================================
    async def _mostrar_conductores_para_viaje(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE,
        viaje_id: int
    ):
        query = update.callback_query
        chat_id = query.message.chat_id

        viaje = self._obtener_viaje(viaje_id)
        if not viaje:
            await query.edit_message_text("❌ Viaje no encontrado")
            return

        zona = viaje.get('zona', '')
        lugar_carga = viaje.get('lugar_carga', '')
        conductores = self._obtener_conductores_zona(zona, lugar_carga)

        # Guardar en cache para la confirmación
        if chat_id not in self._cache_conductores:
            self._cache_conductores[chat_id] = {}
        self._cache_conductores[chat_id][viaje_id] = conductores

        # Cabecera del viaje
        texto = f"📦 *VIAJE #{viaje_id}*\n"
        texto += f"━━━━━━━━━━━━━━━━━━━\n"
        texto += f"🏢 Cliente: *{viaje.get('cliente', 'N/A')}*\n"
        texto += f"📍 Carga: {viaje.get('lugar_carga', '?')}\n"
        texto += f"📍 Descarga: {viaje.get('lugar_entrega', '?')}\n"
        texto += f"📦 Mercancía: {viaje.get('mercancia', 'N/A')}\n"
        texto += f"📏 {viaje.get('km', '?')} km | 💰 {viaje.get('precio', '?')}€\n"
        texto += f"🗺️ Zona: {zona}\n"

        if viaje.get('observaciones'):
            texto += f"📝 {viaje['observaciones'][:80]}\n"

        texto += f"\n━━━━━━━━━━━━━━━━━━━\n"

        if not conductores:
            texto += f"❌ No hay conductores en zona *{zona}*"
            botones = [[InlineKeyboardButton("◀️ Volver", callback_data="asgn:list")]]
            await query.edit_message_text(texto, reply_markup=InlineKeyboardMarkup(botones), parse_mode='Markdown')
            return

        texto += f"👥 *CONDUCTORES {zona}* ({len(conductores)}):\n"
        texto += f"📍 Ordenados por cercanía a {viaje.get('lugar_carga', '?')}\n\n"

        botones = []
        for idx, c in enumerate(conductores):
            nombre = c['nombre']
            tractora = c.get('tractora', '?')
            ubicacion = (c.get('ubicacion', '') or '')[:10]
            viajes_asignados = c.get('_viajes_asignados', 0)
            ausente = c.get('absentismo', '') or ''
            distancia = c.get('_distancia_km')

            # Estado del conductor
            if ausente:
                emoji = "🚫"
                estado = "ABS"
            elif viajes_asignados > 0:
                emoji = "🔶"
                estado = f"{viajes_asignados}v"
            else:
                emoji = "🟢"
                estado = "Libre"

            # Distancia
            if distancia is not None:
                dist_txt = f"{distancia}km"
            else:
                dist_txt = "?km"

            label = f"{emoji} {nombre} | {tractora} | {dist_txt} | {estado}"
            callback = f"asgn:c_{viaje_id}_{idx}"
            botones.append([InlineKeyboardButton(label[:60], callback_data=callback)])

        botones.append([InlineKeyboardButton("◀️ Volver a viajes", callback_data="asgn:list")])

        await query.edit_message_text(
            texto, reply_markup=InlineKeyboardMarkup(botones), parse_mode='Markdown'
        )

    # ================================================================
    # PANTALLA 3: CONFIRMACIÓN
    # ================================================================
    async def _mostrar_confirmacion(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE,
        viaje_id: int, idx_conductor: int
    ):
        query = update.callback_query
        chat_id = query.message.chat_id

        viaje = self._obtener_viaje(viaje_id)
        conductores = self._cache_conductores.get(chat_id, {}).get(viaje_id, [])

        if not viaje or idx_conductor >= len(conductores):
            await query.edit_message_text("❌ Error: datos expirados. Vuelve a empezar.")
            return

        conductor = conductores[idx_conductor]
        viajes_asignados = conductor.get('_viajes_asignados', 0)

        texto = f"✅ *CONFIRMAR ASIGNACIÓN*\n"
        texto += f"━━━━━━━━━━━━━━━━━━━\n\n"
        texto += f"📦 *Viaje:* {viaje.get('cliente', '?')}\n"
        texto += f"   {viaje.get('lugar_carga', '?')} → {viaje.get('lugar_entrega', '?')}\n"
        texto += f"   {viaje.get('mercancia', '')} | {viaje.get('km', '?')} km | {viaje.get('precio', '?')}€\n\n"
        texto += f"👤 *Conductor:* {conductor['nombre']}\n"
        texto += f"   🚛 {conductor.get('tractora', '?')} | 📍 {conductor.get('ubicacion', '?')}\n"

        if viajes_asignados > 0:
            texto += f"\n⚠️ _Este conductor ya tiene {viajes_asignados} viaje(s) asignado(s)._\n"

        texto += f"\n¿Confirmar asignación?"

        botones = [
            [
                InlineKeyboardButton("✅ Confirmar", callback_data=f"asgn:ok_{viaje_id}_{idx_conductor}"),
                InlineKeyboardButton("❌ Cancelar", callback_data=f"asgn:v_{viaje_id}"),
            ],
            [InlineKeyboardButton("◀️ Volver a viajes", callback_data="asgn:list")],
        ]

        await query.edit_message_text(
            texto, reply_markup=InlineKeyboardMarkup(botones), parse_mode='Markdown'
        )

    # ================================================================
    # EJECUTAR ASIGNACIÓN
    # ================================================================
    async def _ejecutar_asignacion(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE,
        viaje_id: int, idx_conductor: int
    ):
        query = update.callback_query
        chat_id = query.message.chat_id

        viaje = self._obtener_viaje(viaje_id)
        conductores = self._cache_conductores.get(chat_id, {}).get(viaje_id, [])

        if not viaje or idx_conductor >= len(conductores):
            await query.edit_message_text("❌ Error: datos expirados. Vuelve a empezar.")
            return

        conductor = conductores[idx_conductor]
        nombre = conductor['nombre']
        tractora = conductor.get('tractora', '')

        # 1. Actualizar BD
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE viajes_empresa
                SET conductor_asignado = ?, tractora_asignada = ?
                WHERE id = ?
            """, (nombre, tractora, viaje_id))
            conn.commit()
            conn.close()
            logger.info(f"[ASIGNACIÓN] BD actualizada: viaje {viaje_id} → {nombre}")
        except Exception as e:
            logger.error(f"[ASIGNACIÓN] Error BD: {e}")
            await query.edit_message_text(f"❌ Error actualizando BD: {e}")
            return

        # 2. Actualizar Excel
        fila_excel = viaje.get('fila_excel')
        if self.excel_path and fila_excel is not None:
            self._actualizar_excel(fila_excel, nombre)

        # 3. Subir a Drive
        if self.on_excel_updated:
            try:
                self.on_excel_updated()
                logger.info("[ASIGNACIÓN] Excel subido a Drive")
            except Exception as e:
                logger.error(f"[ASIGNACIÓN] Error subiendo a Drive: {e}")

        # 4. Notificar al conductor por Telegram
        telegram_id = conductor.get('telegram_id')
        if telegram_id and context.bot:
            try:
                msg_conductor = (
                    f"🚛 *VIAJE ASIGNADO*\n\n"
                    f"🏢 Cliente: *{viaje.get('cliente', '?')}*\n"
                    f"📍 Carga: {viaje.get('lugar_carga', '?')}\n"
                    f"📍 Descarga: {viaje.get('lugar_entrega', '?')}\n"
                    f"📦 Mercancía: {viaje.get('mercancia', 'N/A')}\n"
                    f"📏 {viaje.get('km', '?')} km\n"
                )
                if viaje.get('observaciones'):
                    msg_conductor += f"📝 {viaje['observaciones'][:100]}\n"

                await context.bot.send_message(
                    chat_id=telegram_id,
                    text=msg_conductor,
                    parse_mode='Markdown'
                )
                logger.info(f"[ASIGNACIÓN] Conductor {nombre} notificado ({telegram_id})")
            except Exception as e:
                logger.error(f"[ASIGNACIÓN] Error notificando conductor: {e}")

        # 5. Mensaje de éxito al admin
        texto = (
            f"✅ *VIAJE ASIGNADO*\n\n"
            f"📦 {viaje.get('cliente', '?')}: "
            f"{viaje.get('lugar_carga', '?')} → {viaje.get('lugar_entrega', '?')}\n"
            f"👤 {nombre} ({tractora})\n\n"
        )
        if telegram_id:
            texto += f"📩 Conductor notificado por Telegram"
        else:
            texto += f"⚠️ Conductor sin Telegram vinculado (no notificado)"

        botones = [
            [InlineKeyboardButton("📦 Asignar más viajes", callback_data="asgn:list")],
        ]

        await query.edit_message_text(
            texto, reply_markup=InlineKeyboardMarkup(botones), parse_mode='Markdown'
        )

        # Limpiar cache
        if chat_id in self._cache_conductores:
            self._cache_conductores[chat_id].pop(viaje_id, None)

    # ================================================================
    # QUERIES BD
    # ================================================================
    def _obtener_viajes_sin_asignar(self) -> list:
        """Viajes sin conductor, ordenados por precio descendente."""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM viajes_empresa
                WHERE (conductor_asignado IS NULL OR conductor_asignado = '')
                  AND estado != 'completado'
                ORDER BY precio DESC
            """)
            resultado = [dict(r) for r in cursor.fetchall()]
            conn.close()
            return resultado
        except Exception as e:
            logger.error(f"[ASIGNACIÓN] Error obteniendo viajes: {e}")
            return []

    def _obtener_viaje(self, viaje_id: int) -> Optional[dict]:
        """Obtiene un viaje por ID."""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM viajes_empresa WHERE id = ?", (viaje_id,))
            row = cursor.fetchone()
            conn.close()
            return dict(row) if row else None
        except Exception as e:
            logger.error(f"[ASIGNACIÓN] Error obteniendo viaje {viaje_id}: {e}")
            return None

    def _obtener_conductores_zona(self, zona: str, lugar_carga: str = "") -> list:
        """
        Conductores de la misma zona, ordenados por cercanía al punto de carga.

        Prioridad de posición:
        1. GPS real via Movildata (tractora → última posición)
        2. Última descarga del viaje asignado (si tiene viaje, estará cerca de la descarga)
        3. Campo 'ubicacion' de la BD + COORDENADAS_LUGARES
        
        Orden final: ausentes al final, el resto por distancia ascendente.
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            # Conductores de la zona
            cursor.execute("""
                SELECT * FROM conductores_empresa
                WHERE zona = ?
                  AND nombre IS NOT NULL AND nombre != ''
                ORDER BY nombre
            """, (zona,))
            conductores = [dict(r) for r in cursor.fetchall()]

            # Contar viajes asignados por conductor + obtener última descarga
            cursor.execute("""
                SELECT conductor_asignado, COUNT(*) as n
                FROM viajes_empresa
                WHERE conductor_asignado IS NOT NULL AND conductor_asignado != ''
                  AND estado != 'completado'
                GROUP BY conductor_asignado
            """)
            viajes_por_conductor = {r['conductor_asignado']: r['n'] for r in cursor.fetchall()}

            # Última descarga de cada conductor (para estimar posición)
            cursor.execute("""
                SELECT conductor_asignado, lugar_entrega
                FROM viajes_empresa
                WHERE conductor_asignado IS NOT NULL AND conductor_asignado != ''
                  AND estado != 'completado'
                ORDER BY id DESC
            """)
            ultima_descarga = {}
            for r in cursor.fetchall():
                nombre = r['conductor_asignado']
                if nombre not in ultima_descarga:
                    ultima_descarga[nombre] = r['lugar_entrega']

            conn.close()

            # Coordenadas del punto de carga
            coords_carga = self._obtener_coordenadas(lugar_carga)

            # Enriquecer conductores con distancia
            for c in conductores:
                c['_viajes_asignados'] = viajes_por_conductor.get(c['nombre'], 0)
                c['_distancia_km'] = None
                c['_posicion_origen'] = None  # GPS, descarga o base

                if not coords_carga:
                    continue

                lat_c, lon_c = None, None

                # 1. GPS real via Movildata (puede fallar: API simulada, sin datos, etc.)
                tractora = c.get('tractora') or ''
                if self.movildata and tractora:
                    try:
                        pos = self.movildata.get_last_location_plate(tractora)
                        if pos and isinstance(pos, dict):
                            lat = pos.get('latitud')
                            lon = pos.get('longitud')
                            # Validar que son floats reales y no 0,0
                            if (lat and lon
                                    and isinstance(lat, (int, float))
                                    and isinstance(lon, (int, float))
                                    and not (lat == 0 and lon == 0)):
                                lat_c = float(lat)
                                lon_c = float(lon)
                                c['_posicion_origen'] = 'GPS'
                    except Exception as e:
                        logger.debug(f"[ASIGNACIÓN] GPS no disponible para {tractora}: {e}")

                # 2. Última descarga del viaje asignado
                if lat_c is None and c['nombre'] in ultima_descarga:
                    try:
                        coords_desc = self._obtener_coordenadas(ultima_descarga[c['nombre']])
                        if coords_desc:
                            lat_c, lon_c = coords_desc
                            c['_posicion_origen'] = 'descarga'
                    except Exception:
                        pass

                # 3. Ubicación base del conductor
                if lat_c is None:
                    try:
                        ubicacion = c.get('ubicacion') or ''
                        if ubicacion:
                            coords_ubi = self._obtener_coordenadas(ubicacion)
                            if coords_ubi:
                                lat_c, lon_c = coords_ubi
                                c['_posicion_origen'] = 'base'
                    except Exception:
                        pass

                # Calcular distancia (solo si tenemos ambas coordenadas válidas)
                if lat_c is not None and lon_c is not None:
                    try:
                        c['_distancia_km'] = self._calcular_distancia_km(
                            lat_c, lon_c, coords_carga[0], coords_carga[1]
                        )
                    except Exception:
                        c['_distancia_km'] = None

            # Ordenar: ausentes al final, resto por distancia (None = al final de los activos)
            def sort_key(c):
                ausente = 1 if (c.get('absentismo') or '') else 0
                dist = c.get('_distancia_km')
                if dist is None:
                    dist = 99999
                return (ausente, dist)

            conductores.sort(key=sort_key)
            return conductores

        except Exception as e:
            logger.error(f"[ASIGNACIÓN] Error obteniendo conductores zona {zona}: {e}")
            return []

    # ================================================================
    # UTILIDADES DE DISTANCIA
    # ================================================================
    @staticmethod
    def _obtener_coordenadas(lugar: str) -> Optional[tuple]:
        """Busca coordenadas de un lugar en COORDENADAS_LUGARES."""
        try:
            if not lugar or not isinstance(lugar, str):
                return None
            lugar_upper = lugar.upper().strip()
            if not lugar_upper:
                return None

            # Búsqueda exacta
            if lugar_upper in COORDENADAS_LUGARES:
                return COORDENADAS_LUGARES[lugar_upper]

            # Búsqueda parcial (ej: "CALAHORRA (LA RIOJA)" → match "CALAHORRA")
            for nombre, coords in COORDENADAS_LUGARES.items():
                if nombre in lugar_upper or lugar_upper in nombre:
                    return coords

            return None
        except Exception:
            return None

    @staticmethod
    def _calcular_distancia_km(lat1: float, lon1: float, lat2: float, lon2: float) -> Optional[int]:
        """Distancia Haversine en km (redondeada). None si falla."""
        try:
            R = 6371
            dlat = math.radians(lat2 - lat1)
            dlon = math.radians(lon2 - lon1)
            a = (math.sin(dlat / 2) ** 2 +
                 math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
                 math.sin(dlon / 2) ** 2)
            c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
            return round(R * c)
        except Exception:
            return None

    # ================================================================
    # ACTUALIZAR EXCEL
    # ================================================================
    def _actualizar_excel(self, fila_excel: int, nombre_conductor: str):
        """Escribe el nombre del conductor en la columna TRANSPORTISTA del Excel."""
        try:
            from openpyxl import load_workbook

            if not self.excel_path or not Path(self.excel_path).exists():
                logger.warning("[ASIGNACIÓN] Excel no encontrado")
                return

            wb = load_workbook(self.excel_path)
            ws = wb.active

            fila_openpyxl = fila_excel + 1  # 0-indexed → 1-indexed

            if fila_openpyxl > ws.max_row:
                logger.error(f"[ASIGNACIÓN] Fila {fila_openpyxl} fuera de rango")
                wb.close()
                return

            celda = ws.cell(row=fila_openpyxl, column=COL_TRANSPORTISTA)
            anterior = celda.value
            celda.value = nombre_conductor

            wb.save(self.excel_path)
            wb.close()

            logger.info(
                f"[ASIGNACIÓN] Excel fila {fila_openpyxl}: "
                f"TRANSPORTISTA = '{nombre_conductor}' (antes: '{anterior}')"
            )

        except Exception as e:
            logger.error(f"[ASIGNACIÓN] Error actualizando Excel: {e}")
