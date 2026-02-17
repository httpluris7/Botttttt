"""
LECTOR DE EMAILS DE VIAJES (v2.0 - INTEGRADO CON BOT)
=====================================================
Lee emails de un buzón común, interpreta los datos con GPT-4
y los añade al Excel de viajes.

INTEGRACIÓN CON BOT:
- Se importa desde bot_transporte.py
- Usa el drive_service del bot
- Notifica a admins por Telegram
- Se ejecuta cada X minutos con JobQueue

USO:
    from lector_emails_viajes import LectorEmailsViajes, crear_job_lector_emails
    
    # En main():
    lector = LectorEmailsViajes(config, drive_service)
    crear_job_lector_emails(app, lector, admin_ids)
"""

import imaplib
import email
from email.header import decode_header
import os
import json
import logging
from datetime import datetime
from typing import Optional, Dict, List
import re
import openpyxl
import sqlite3
from pathlib import Path

# OpenAI
try:
    from openai import OpenAI
except ImportError:
    OpenAI = None

# Google Drive
try:
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    import io
    DRIVE_AVAILABLE = True
except ImportError:
    DRIVE_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class LectorEmailsViajes:
    """
    Lee emails y extrae datos de viajes usando GPT-4.
    Diseñado para integrarse con bot_transporte.py
    """
    
    def __init__(
        self,
        email_user: str,
        email_password: str,
        openai_api_key: str,
        excel_path: str,
        db_path: str = None,
        drive_service=None,
        drive_excel_id: str = None,
        imap_server: str = "imap.gmail.com",
        imap_port: int = 993,
        confianza_minima: int = 70
    ):
        self.email_user = email_user
        self.email_password = email_password
        self.imap_server = imap_server
        self.imap_port = imap_port
        self.excel_path = excel_path
        self.db_path = db_path
        self.drive_service = drive_service
        self.drive_excel_id = drive_excel_id
        self.confianza_minima = confianza_minima
        
        # Verificar/crear columnas de fechas en BD
        if db_path and Path(db_path).exists():
            self._verificar_columnas_bd()
        
        # OpenAI client
        if openai_api_key and OpenAI:
            self.openai_client = OpenAI(api_key=openai_api_key)
        else:
            self.openai_client = None
            logger.warning("OpenAI no configurado - interpretación de emails desactivada")
        
        # Prompt mejorado para GPT-4
        self.prompt_sistema = """Eres un experto en logística de transporte que extrae datos de viajes desde emails.

ANALIZA el email y extrae los siguientes campos en formato JSON:

{
    "cliente": "nombre de la empresa cliente (buscar en 'Cliente:', 'Empresa:', firma del email, o dominio del remitente)",
    "num_pedido": "número de pedido (buscar en 'Pedido:', 'Nº Pedido:', 'OT-', 'Orden:')",
    "ref_cliente": "referencia del cliente (buscar en 'Ref:', 'Referencia:', 'Ref. Cliente:', 'Ref. interna:')",
    "lugar_carga": "ciudad de recogida en MAYÚSCULAS (buscar en 'Origen:', 'Carga:', 'Recogida:', 'Desde:')",
    "fecha_carga": "fecha de carga en formato DD/MM/YYYY",
    "hora_carga": "hora o rango de carga (ej: '08:00' o '08:00-14:00')",
    "lugar_descarga": "ciudad de entrega en MAYÚSCULAS (buscar en 'Destino:', 'Descarga:', 'Entrega:', 'Hasta:')",
    "fecha_descarga": "fecha de entrega en formato DD/MM/YYYY",
    "hora_descarga": "hora o rango de entrega",
    "mercancia": "tipo de mercancía (normalizar: SECO, REFRIGERADO +Xº, CONGELADO -Xº)",
    "num_pales": "número de palés (solo el número, sin texto)",
    "intercambio": "SI o NO (si hay intercambio de palés)",
    "observaciones": "notas adicionales importantes",
    "confianza": "número 1-100 indicando confianza en la extracción"
}

REGLAS IMPORTANTES:
1. Para CLIENTE: Si no hay campo explícito, buscar en:
   - Firma del email (ej: "Dpto. Logística HERO" → cliente = "HERO")
   - Dominio del remitente (ej: "pedidos@hero.es" → cliente = "HERO")
   - Nombre de la empresa en el asunto o cuerpo

2. Para REFERENCIA: Buscar cualquier código con formato:
   - "REF-2026/123", "V12345", "2026/123QWE", etc.
   - Si hay "Nº Pedido" Y "Referencia", usar la referencia

3. Para FECHAS: Convertir siempre a DD/MM/YYYY
   - "mañana" → calcular fecha
   - "18/02" → "18/02/2026"
   - "18 de febrero" → "18/02/2026"

4. Para HORAS: Mantener formato original si es rango
   - "entre 8 y 14" → "08:00-14:00"
   - "antes de las 6" → "06:00"

5. Si un campo NO está en el email, poner null

6. Si hay MÚLTIPLES viajes en un email, devolver un ARRAY

Responde SOLO con el JSON válido, sin explicaciones ni markdown."""

    def _verificar_columnas_bd(self):
        """Verifica y crea columnas de fechas si no existen"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Verificar columnas existentes
            cursor.execute("PRAGMA table_info(viajes_empresa)")
            columnas = [row[1] for row in cursor.fetchall()]
            
            # Añadir columnas si no existen
            nuevas_columnas = [
                ('fecha_carga', 'TEXT'),
                ('hora_carga', 'TEXT'),
                ('fecha_descarga', 'TEXT'),
                ('hora_descarga', 'TEXT'),
                ('email_origen', 'TEXT')
            ]
            
            for col_nombre, col_tipo in nuevas_columnas:
                if col_nombre not in columnas:
                    cursor.execute(f"ALTER TABLE viajes_empresa ADD COLUMN {col_nombre} {col_tipo}")
                    logger.info(f"[BD] Columna '{col_nombre}' añadida a viajes_empresa")
            
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"[BD] Error verificando columnas: {e}")
    
    def _actualizar_viaje_bd(self, viaje: Dict, fila_excel: int) -> bool:
        """Actualiza el viaje en la BD con fechas/horas"""
        logger.info(f"[BD] Intentando actualizar viaje fila {fila_excel}, db_path={self.db_path}")
        
        if not self.db_path or not Path(self.db_path).exists():
            logger.error(f"[BD] db_path no existe: {self.db_path}")
            return False
        
        try:
            logger.info(f"[BD] Conectando a {self.db_path}")
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Buscar viaje por fila_excel
            logger.info(f"[BD] UPDATE fila_excel={fila_excel}, fecha_carga={viaje.get('fecha_carga')}")
            cursor.execute("""
                UPDATE viajes_empresa 
                SET fecha_carga = ?,
                    hora_carga = ?,
                    fecha_descarga = ?,
                    hora_descarga = ?,
                    email_origen = ?
                WHERE fila_excel = ?
            """, (
                viaje.get('fecha_carga'),
                viaje.get('hora_carga'),
                viaje.get('fecha_descarga'),
                viaje.get('hora_descarga'),
                viaje.get('_email_asunto', '')[:100],
                fila_excel
            ))
            
            logger.info(f"[BD] UPDATE afectó {cursor.rowcount} filas")
            
            # Si no existe, insertar nuevo registro con datos básicos
            if cursor.rowcount == 0:
                logger.info(f"[BD] No existe, insertando nuevo registro")
                cursor.execute("""
                    INSERT INTO viajes_empresa (
                        cliente, num_pedido, ref_cliente, lugar_carga, lugar_entrega,
                        mercancia, intercambio, num_pales, fila_excel,
                        fecha_carga, hora_carga, fecha_descarga, hora_descarga, email_origen,
                        estado, fecha_sync
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pendiente', ?)
                """, (
                    viaje.get('cliente'),
                    viaje.get('num_pedido'),
                    viaje.get('ref_cliente'),
                    viaje.get('lugar_carga'),
                    viaje.get('lugar_descarga'),
                    viaje.get('mercancia'),
                    viaje.get('intercambio'),
                    viaje.get('num_pales'),
                    fila_excel,
                    viaje.get('fecha_carga'),
                    viaje.get('hora_carga'),
                    viaje.get('fecha_descarga'),
                    viaje.get('hora_descarga'),
                    viaje.get('_email_asunto', '')[:100],
                    datetime.now().isoformat()
                ))
            
            conn.commit()
            conn.close()
            logger.info(f"[BD] ✅ Viaje actualizado/insertado (fila {fila_excel})")
            return True
        except Exception as e:
            logger.error(f"[BD] Error actualizando viaje: {e}")
            return False

    def conectar(self) -> bool:
        """Conecta al servidor IMAP"""
        try:
            self.mail = imaplib.IMAP4_SSL(self.imap_server, self.imap_port)
            self.mail.login(self.email_user, self.email_password)
            logger.info(f"✅ [EMAIL] Conectado a {self.email_user}")
            return True
        except Exception as e:
            logger.error(f"❌ [EMAIL] Error conectando: {e}")
            return False
    
    def desconectar(self):
        """Cierra la conexión"""
        try:
            self.mail.logout()
        except:
            pass
    
    def _decodificar_header(self, header) -> str:
        """Decodifica header de email"""
        if not header:
            return ""
        decoded = decode_header(header)
        result = []
        for part, encoding in decoded:
            if isinstance(part, bytes):
                result.append(part.decode(encoding or 'utf-8', errors='replace'))
            else:
                result.append(part)
        return ' '.join(result)
    
    def _extraer_cuerpo(self, msg) -> str:
        """Extrae el cuerpo del email"""
        cuerpo = ""
        
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get("Content-Disposition"))
                
                if content_type == "text/plain" and "attachment" not in content_disposition:
                    try:
                        payload = part.get_payload(decode=True)
                        charset = part.get_content_charset() or 'utf-8'
                        cuerpo += payload.decode(charset, errors='replace')
                    except:
                        pass
        else:
            try:
                payload = msg.get_payload(decode=True)
                charset = msg.get_content_charset() or 'utf-8'
                cuerpo = payload.decode(charset, errors='replace')
            except:
                cuerpo = str(msg.get_payload())
        
        return cuerpo.strip()
    
    def leer_emails_no_leidos(self, carpeta: str = "INBOX", limit: int = 10) -> List[Dict]:
        """Lee emails no leídos"""
        emails = []
        
        try:
            self.mail.select(carpeta)
            status, mensajes = self.mail.search(None, 'UNSEEN')
            
            if status != 'OK':
                return emails
            
            ids_mensajes = mensajes[0].split()
            logger.info(f"📬 [EMAIL] {len(ids_mensajes)} emails no leídos")
            
            ids_mensajes = ids_mensajes[-limit:] if len(ids_mensajes) > limit else ids_mensajes
            
            for msg_id in ids_mensajes:
                try:
                    status, data = self.mail.fetch(msg_id, '(RFC822)')
                    if status != 'OK':
                        continue
                    
                    msg = email.message_from_bytes(data[0][1])
                    
                    email_data = {
                        'id': msg_id.decode(),
                        'de': self._decodificar_header(msg.get('From')),
                        'asunto': self._decodificar_header(msg.get('Subject')),
                        'fecha': msg.get('Date'),
                        'cuerpo': self._extraer_cuerpo(msg)
                    }
                    
                    emails.append(email_data)
                    logger.info(f"📧 [EMAIL] {email_data['asunto'][:50]}...")
                    
                except Exception as e:
                    logger.error(f"[EMAIL] Error procesando email {msg_id}: {e}")
            
        except Exception as e:
            logger.error(f"[EMAIL] Error leyendo emails: {e}")
        
        return emails
    
    def interpretar_email(self, email_data: Dict) -> Optional[List[Dict]]:
        """Usa GPT-4 para interpretar el email y extraer datos del viaje"""
        
        if not self.openai_client:
            logger.error("[EMAIL] OpenAI no configurado")
            return None
        
        # Extraer dominio del remitente para ayudar a identificar cliente
        remitente = email_data.get('de', '')
        dominio = ""
        if '@' in remitente:
            dominio = remitente.split('@')[1].split('>')[0].split('.')[0].upper()
        
        contenido = f"""
REMITENTE: {email_data.get('de', '')}
DOMINIO: {dominio}
ASUNTO: {email_data.get('asunto', '')}
FECHA DEL EMAIL: {email_data.get('fecha', '')}

CONTENIDO:
{email_data.get('cuerpo', '')}
"""
        
        try:
            response = self.openai_client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": self.prompt_sistema},
                    {"role": "user", "content": contenido}
                ],
                temperature=0.1,
                max_tokens=1500
            )
            
            respuesta = response.choices[0].message.content.strip()
            
            # Limpiar respuesta
            respuesta = re.sub(r'^```json\s*', '', respuesta)
            respuesta = re.sub(r'\s*```$', '', respuesta)
            
            datos = json.loads(respuesta)
            
            if isinstance(datos, dict):
                datos = [datos]
            
            logger.info(f"✅ [EMAIL] Interpretados {len(datos)} viaje(s)")
            return datos
            
        except json.JSONDecodeError as e:
            logger.error(f"[EMAIL] Error parseando JSON: {e}")
            return None
        except Exception as e:
            logger.error(f"[EMAIL] Error llamando a OpenAI: {e}")
            return None
    
    def _descargar_excel_de_drive(self) -> bool:
        """Descarga el Excel desde Drive"""
        if not self.drive_service or not DRIVE_AVAILABLE:
            return False
        
        if not self.drive_excel_id:
            return False
        
        try:
            logger.info(f"[DRIVE] Descargando Excel...")
            
            request = self.drive_service.files().get_media(fileId=self.drive_excel_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            
            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            fh.seek(0)
            with open(self.excel_path, 'wb') as f:
                f.write(fh.read())
            
            logger.info(f"[DRIVE] ✅ Excel descargado")
            return True
        except Exception as e:
            logger.error(f"[DRIVE] Error descargando: {e}")
            return False
    
    def _subir_excel_a_drive(self) -> bool:
        """Sube el Excel a Drive"""
        if not self.drive_service or not DRIVE_AVAILABLE:
            return False
        
        if not self.drive_excel_id:
            return False
        
        if not Path(self.excel_path).exists():
            return False
        
        try:
            logger.info(f"[DRIVE] Subiendo Excel...")
            
            media = MediaFileUpload(
                self.excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                resumable=True
            )
            
            self.drive_service.files().update(
                fileId=self.drive_excel_id,
                media_body=media
            ).execute()
            
            logger.info(f"[DRIVE] ✅ Excel subido")
            return True
        except Exception as e:
            logger.error(f"[DRIVE] Error subiendo: {e}")
            return False
    
    def añadir_viaje_excel(self, viaje: Dict) -> bool:
        """Añade un viaje al Excel y a la BD"""
        
        if not self.excel_path or not Path(self.excel_path).exists():
            logger.error(f"[EXCEL] No encontrado: {self.excel_path}")
            return False
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            # Buscar primera fila vacía
            fila_nueva = 3
            for fila in range(3, 500):
                if not ws.cell(row=fila, column=9).value:
                    fila_nueva = fila
                    break
            
            # Escribir datos (manejando None)
            cliente = viaje.get('cliente') or ''
            ws.cell(row=fila_nueva, column=9, value=cliente.upper() if cliente else '')
            
            ws.cell(row=fila_nueva, column=10, value=viaje.get('num_pedido'))
            ws.cell(row=fila_nueva, column=11, value=viaje.get('ref_cliente'))
            
            intercambio = viaje.get('intercambio') or 'NO'
            ws.cell(row=fila_nueva, column=12, value='SI' if intercambio.upper() in ['SI', 'SÍ', 'YES', 'S'] else 'NO')
            
            # Nº palés
            num_pales = viaje.get('num_pales')
            if num_pales:
                try:
                    ws.cell(row=fila_nueva, column=13, value=int(num_pales))
                except:
                    pass
            
            # Lugares
            lugar_carga = viaje.get('lugar_carga') or ''
            lugar_descarga = viaje.get('lugar_descarga') or ''
            ws.cell(row=fila_nueva, column=14, value=lugar_carga.upper() if lugar_carga else '')
            ws.cell(row=fila_nueva, column=17, value=lugar_descarga.upper() if lugar_descarga else '')
            
            # NO escribir fechas/horas en columnas Excel (se guardan en BD)
            
            # Mercancía
            mercancia = viaje.get('mercancia') or ''
            ws.cell(row=fila_nueva, column=20, value=mercancia.upper() if mercancia else '')
            
            # Observaciones (sin fechas/horas, van a la BD)
            obs_parts = []
            if viaje.get('observaciones'):
                obs_parts.append(viaje.get('observaciones'))
            
            ws.cell(row=fila_nueva, column=28, value=' | '.join(obs_parts) if obs_parts else '')
            
            wb.save(self.excel_path)
            wb.close()
            
            logger.info(f"✅ [EXCEL] Viaje añadido fila {fila_nueva}: {cliente} | {lugar_carga} → {lugar_descarga}")
            
            # Guardar fechas/horas en BD
            self._actualizar_viaje_bd(viaje, fila_nueva)
            
            return True
            
        except Exception as e:
            logger.error(f"[EXCEL] Error añadiendo viaje: {e}")
            return False
    
    def marcar_como_leido(self, email_id: str):
        """Marca un email como leído"""
        try:
            self.mail.store(email_id.encode(), '+FLAGS', '\\Seen')
        except Exception as e:
            logger.error(f"[EMAIL] Error marcando como leído: {e}")
    
    def mover_a_procesados(self, email_id: str, carpeta_destino: str = "Procesados"):
        """Mueve email a carpeta de procesados"""
        try:
            self.mail.create(carpeta_destino)
        except:
            pass
        
        try:
            self.mail.copy(email_id.encode(), carpeta_destino)
            self.mail.store(email_id.encode(), '+FLAGS', '\\Deleted')
            self.mail.expunge()
        except Exception as e:
            logger.error(f"[EMAIL] Error moviendo email: {e}")
    
    def procesar_emails(self) -> List[Dict]:
        """
        Proceso principal: lee emails, interpreta y añade viajes
        
        Returns:
            Lista de viajes procesados con éxito
        """
        viajes_procesados = []
        
        # 1. Descargar Excel de Drive
        if self.drive_service:
            self._descargar_excel_de_drive()
        
        # 2. Conectar al email
        if not self.conectar():
            return viajes_procesados
        
        try:
            emails = self.leer_emails_no_leidos()
            
            if not emails:
                self.desconectar()
                return viajes_procesados
            
            for email_data in emails:
                logger.info(f"📧 [EMAIL] Procesando: {email_data['asunto']}")
                
                # Interpretar con GPT-4
                viajes = self.interpretar_email(email_data)
                
                if not viajes:
                    logger.warning(f"[EMAIL] No se pudieron extraer viajes")
                    self.marcar_como_leido(email_data['id'])
                    continue
                
                for viaje in viajes:
                    confianza = viaje.get('confianza', 0)
                    
                    viaje['_email_id'] = email_data['id']
                    viaje['_email_asunto'] = email_data['asunto']
                    viaje['_email_de'] = email_data['de']
                    viaje['_procesado_en'] = datetime.now().isoformat()
                    
                    # Solo añadir si confianza suficiente
                    if confianza >= self.confianza_minima:
                        if self.añadir_viaje_excel(viaje):
                            viaje['_añadido'] = True
                            viajes_procesados.append(viaje)
                        else:
                            viaje['_añadido'] = False
                    else:
                        viaje['_añadido'] = False
                        viaje['_motivo'] = f"Confianza {confianza}% < {self.confianza_minima}%"
                        logger.warning(f"[EMAIL] Viaje descartado: {viaje['_motivo']}")
                
                # Marcar email como procesado
                self.marcar_como_leido(email_data['id'])
                self.mover_a_procesados(email_data['id'])
            
            # 3. Subir Excel a Drive si hubo cambios
            if viajes_procesados and self.drive_service:
                self._subir_excel_a_drive()
            
        finally:
            self.desconectar()
        
        logger.info(f"📊 [EMAIL] Total viajes añadidos: {len(viajes_procesados)}")
        return viajes_procesados
    
    def generar_mensaje_notificacion(self, viaje: Dict) -> str:
        """Genera mensaje de notificación para Telegram"""
        confianza = viaje.get('confianza', 0)
        emoji = "✅" if confianza >= 80 else "⚠️" if confianza >= 60 else "❌"
        
        cliente = viaje.get('cliente') or 'Sin cliente'
        pedido = viaje.get('num_pedido') or '-'
        ref = viaje.get('ref_cliente') or '-'
        
        lugar_carga = viaje.get('lugar_carga') or '?'
        fecha_carga = viaje.get('fecha_carga') or ''
        hora_carga = viaje.get('hora_carga') or ''
        
        lugar_descarga = viaje.get('lugar_descarga') or '?'
        fecha_descarga = viaje.get('fecha_descarga') or ''
        hora_descarga = viaje.get('hora_descarga') or ''
        
        mercancia = viaje.get('mercancia') or '-'
        pales = viaje.get('num_pales') or '-'
        intercambio = viaje.get('intercambio') or 'NO'
        
        mensaje = f"""📧 *VIAJE RECIBIDO POR EMAIL*

{emoji} Confianza: {confianza}%

🏢 *Cliente:* {cliente}
📋 *Pedido:* {pedido}
🏷️ *Referencia:* {ref}

📥 *CARGA:*
   📍 {lugar_carga}
   📅 {fecha_carga} {hora_carga}

📤 *DESCARGA:*
   📍 {lugar_descarga}
   📅 {fecha_descarga} {hora_descarga}

📦 *Mercancía:* {mercancia}
🔢 *Palés:* {pales}
🔄 *Intercambio:* {intercambio}

✅ _Añadido al Excel automáticamente_"""
        
        return mensaje


# ============================================================
# INTEGRACIÓN CON BOT
# ============================================================

async def job_lector_emails(context):
    """Job que se ejecuta periódicamente para leer emails"""
    lector = context.job.data.get('lector')
    admin_ids = context.job.data.get('admin_ids', [])
    
    if not lector:
        return
    
    try:
        viajes = lector.procesar_emails()
        
        # Notificar a admins por cada viaje añadido
        for viaje in viajes:
            if viaje.get('_añadido'):
                mensaje = lector.generar_mensaje_notificacion(viaje)
                
                for admin_id in admin_ids:
                    try:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text=mensaje,
                            parse_mode="Markdown"
                        )
                    except Exception as e:
                        logger.error(f"[EMAIL] Error notificando a admin {admin_id}: {e}")
    
    except Exception as e:
        logger.error(f"[EMAIL] Error en job_lector_emails: {e}")


def crear_job_lector_emails(app, lector: LectorEmailsViajes, admin_ids: list, intervalo_segundos: int = 300):
    """
    Crea el job periódico para leer emails
    
    Args:
        app: Application de telegram
        lector: Instancia de LectorEmailsViajes
        admin_ids: Lista de IDs de admins para notificar
        intervalo_segundos: Intervalo entre lecturas (default 5 min)
    """
    if not app.job_queue:
        logger.error("[EMAIL] JobQueue no disponible")
        return
    
    app.job_queue.run_repeating(
        job_lector_emails,
        interval=intervalo_segundos,
        first=60,  # Primera ejecución en 1 minuto
        data={'lector': lector, 'admin_ids': admin_ids}
    )
    
    logger.info(f"✅ [EMAIL] Job lector emails configurado (cada {intervalo_segundos//60} min)")


# ============================================================
# EJECUCIÓN STANDALONE (para pruebas)
# ============================================================

def main():
    """Ejecuta el lector una vez (para pruebas)"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Lector de emails de viajes')
    parser.add_argument('--config', help='Ruta al archivo de configuración JSON')
    parser.add_argument('--verbose', '-v', action='store_true', help='Modo verbose')
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Cargar configuración
    if not args.config or not Path(args.config).exists():
        print("❌ Usa: python lector_emails_viajes.py --config config_emails.json")
        return
    
    with open(args.config, 'r') as f:
        config = json.load(f)
    
    lector = LectorEmailsViajes(
        email_user=config.get('email_user'),
        email_password=config.get('email_password'),
        openai_api_key=config.get('openai_api_key'),
        excel_path=config.get('excel_path'),
        drive_service=None,  # Sin Drive en modo standalone
        drive_excel_id=config.get('drive_excel_id'),
        imap_server=config.get('imap_server', 'imap.gmail.com'),
        imap_port=config.get('imap_port', 993),
        confianza_minima=config.get('confianza_minima', 70)
    )
    
    viajes = lector.procesar_emails()
    
    print(f"\n{'='*50}")
    print(f"📊 RESUMEN")
    print(f"{'='*50}")
    print(f"Viajes procesados: {len(viajes)}")
    
    for v in viajes:
        print(f"\n{lector.generar_mensaje_notificacion(v)}")


if __name__ == "__main__":
    main()
