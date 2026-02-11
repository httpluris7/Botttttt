"""
CIERRE DE DÍA v2.0
===================
Gestiona el cierre diario creando un nuevo Excel para el día siguiente.

IMPORTANTE: 
- NO modifica el Excel original (PRUEBO.xlsx)
- CREA un nuevo archivo RUTAS_DD-MM-YYYY.xlsx

Estructura del Excel:
- Fila 1: Cabeceras de sección (ZONA NORTE, FECHA)
- Fila 2: Cabeceras de columnas
- Fila 3+: Datos

Columnas:
- 1-8: Datos conductor (Zona, Ubicación, H.LL, H.SA, Transportista, Absentismo, Tractora, Remolque)
- 9-28: Datos viaje (Cliente, Pedido, Ref, Intercambio, Palés, Carga, Horas, Descarga, Horas, etc.)

Proceso de cierre:
1. Analiza Excel actual
2. Identifica conductores que terminaron (hora salida descarga registrada)
3. Identifica viajes pendientes (sin hora salida descarga)
4. CREA nuevo Excel RUTAS_DD-MM-YYYY.xlsx (no modifica original)
5. Exporta conductores con ubicación actualizada
6. Exporta viajes pendientes
"""

import os
import sys
import logging
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURACIÓN DE COLUMNAS (estructura real del Excel)
# ============================================================

# Columnas de conductores (parte izquierda)
COL_CONDUCTOR = {
    'zona': 1,            # A - 24/45
    'ubicacion': 2,       # B - UBICACIÓN
    'hora_llegada': 3,    # C - H.LL
    'hora_salida': 4,     # D - H.SA
    'nombre': 5,          # E - TRANSPORTISTA (nombre del conductor)
    'absentismo': 6,      # F - ABSENTISMO
    'tractora': 7,        # G - TRACTORA
    'remolque': 8,        # H - REMOLQUE
}

# Columnas de viajes (parte derecha)
COL_VIAJE = {
    'cliente': 9,              # I - CLIENTE
    'num_pedido': 10,          # J - Nº PEDIDO
    'ref_cliente': 11,         # K - REF. CLIENTE
    'intercambio': 12,         # L - INTERCAMBIO
    'num_pales': 13,           # M - Nº PALÉS
    'lugar_carga': 14,         # N - LUGAR DE CARGA
    'hora_llegada_carga': 15,  # O - HORA LLEGADA (carga)
    'hora_salida_carga': 16,   # P - HORA SALIDA (carga)
    'lugar_descarga': 17,      # Q - LUGAR DE ENTREGA
    'hora_llegada_descarga': 18,  # R - HORA LLEGADA (descarga)
    'hora_salida_descarga': 19,   # S - HORA SALIDA (descarga) ← CLAVE
    'mercancia': 20,           # T - MERCANCIA
    'cargador': 21,            # U - CARGADOR
    'transportista': 22,       # V - TRANSPORTISTA (del viaje)
    'precio_cobro': 23,        # W - PRECIO COBRO
    'km': 24,                  # X - KM
    'eur_km': 25,              # Y - EUR/KM
    'pago_cargador': 26,       # Z - PAGO CARGADOR
    'num_exp': 27,             # AA - Nº EXP.
    'observaciones': 28,       # AB - OBSERVACIONES
}

# Filas especiales
FILA_SECCIONES = 1      # Fila con "ZONA NORTE", "FECHA", etc.
FILA_CABECERAS = 2      # Fila con nombres de columnas
FILA_INICIO_DATOS = 3   # Primera fila de datos


@dataclass
class ConductorParaExportar:
    """Datos de un conductor para exportar al nuevo día"""
    fila_original: int
    zona: str
    ubicacion: str
    nombre: str
    absentismo: str
    tractora: str
    remolque: str


@dataclass
class ViajePendiente:
    """Datos de un viaje pendiente para exportar al nuevo día"""
    fila_original: int
    datos: Dict  # Todos los datos del viaje


class CierreDia:
    """
    Gestiona el cierre de día y creación del Excel nuevo.
    
    IMPORTANTE: No modifica el Excel original, solo crea uno nuevo.
    """
    
    def __init__(self, excel_path: str, db_path: str, 
                 directorio_excels: str = None,
                 subir_drive_func=None,
                 subir_archivo_nuevo_func=None):
        """
        Args:
            excel_path: Ruta al Excel actual (PRUEBO.xlsx)
            db_path: Ruta a la base de datos
            directorio_excels: Directorio donde guardar los Excels nuevos
            subir_drive_func: Función para subir PRUEBO.xlsx a Drive
            subir_archivo_nuevo_func: Función para subir archivos nuevos a Drive
        """
        self.excel_path = excel_path
        self.db_path = db_path
        self.directorio = directorio_excels or str(Path(excel_path).parent)
        self.subir_drive = subir_drive_func
        self.subir_archivo_nuevo = subir_archivo_nuevo_func
        
        # Archivo que guarda el Excel activo
        self.archivo_activo_path = os.path.join(self.directorio, "excel_activo.txt")
        
        logger.info("[CIERRE] Módulo de cierre de día v2.0 inicializado")
    
    # ============================================================
    # UTILIDADES
    # ============================================================
    
    def obtener_excel_activo(self) -> str:
        """Obtiene el nombre del Excel activo actual"""
        return os.path.basename(self.excel_path)
    
    def generar_nombre_excel(self, fecha: datetime = None) -> str:
        """Genera el nombre del Excel para una fecha"""
        if fecha is None:
            fecha = datetime.now()
        return f"RUTAS_{fecha.strftime('%d-%m-%Y')}.xlsx"
    
    def _valor_celda(self, ws, fila: int, columna: int) -> str:
        """Obtiene el valor de una celda como string limpio"""
        valor = ws.cell(row=fila, column=columna).value
        if valor is None:
            return ''
        return str(valor).strip()
    
    def _tiene_hora_salida_descarga(self, ws, fila: int) -> bool:
        """Verifica si la fila tiene hora de salida de descarga"""
        valor = self._valor_celda(ws, fila, COL_VIAJE['hora_salida_descarga'])
        return bool(valor and valor not in ['', 'None', 'VACIO'])
    
    # ============================================================
    # ANÁLISIS DEL EXCEL ACTUAL
    # ============================================================
    
    def analizar_excel_actual(self) -> Dict:
        """
        Analiza el Excel actual para identificar:
        - Conductores que terminaron (tienen hora salida descarga)
        - Conductores disponibles (sin viaje asignado)
        - Viajes pendientes (sin hora salida descarga)
        
        Returns:
            Dict con listas de conductores y viajes
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            conductores_terminaron = []      # Terminaron viaje
            conductores_disponibles = []     # Sin viaje asignado
            viajes_pendientes = []
            viajes_completados = []
            
            conductores_vistos = set()  # Para evitar duplicados
            
            # Recorrer filas de datos (desde fila 3)
            for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
                nombre_conductor = self._valor_celda(ws, fila, COL_CONDUCTOR['nombre'])
                cliente = self._valor_celda(ws, fila, COL_VIAJE['cliente'])
                
                # Saltar filas vacías
                if not nombre_conductor and not cliente:
                    continue
                
                # Saltar filas que son cabeceras repetidas (secciones intermedias)
                if nombre_conductor.upper() in ['TRANSPORTISTA', 'NOMBRE', 'CONDUCTOR']:
                    continue
                if cliente.upper() in ['CLIENTE', 'CLIENTES']:
                    continue
                
                # Verificar si tiene hora salida descarga
                tiene_hora_salida = self._tiene_hora_salida_descarga(ws, fila)
                
                # Procesar conductor (si hay nombre y no lo hemos visto)
                if nombre_conductor and nombre_conductor.upper() not in conductores_vistos:
                    conductores_vistos.add(nombre_conductor.upper())
                    
                    ubicacion = self._valor_celda(ws, fila, COL_CONDUCTOR['ubicacion'])
                    
                    # Si terminó viaje, actualizar ubicación a lugar de descarga
                    if tiene_hora_salida:
                        lugar_descarga = self._valor_celda(ws, fila, COL_VIAJE['lugar_descarga'])
                        if lugar_descarga:
                            ubicacion = lugar_descarga
                    
                    conductor = ConductorParaExportar(
                        fila_original=fila,
                        zona=self._valor_celda(ws, fila, COL_CONDUCTOR['zona']),
                        ubicacion=ubicacion,
                        nombre=nombre_conductor,
                        absentismo=self._valor_celda(ws, fila, COL_CONDUCTOR['absentismo']),
                        tractora=self._valor_celda(ws, fila, COL_CONDUCTOR['tractora']),
                        remolque=self._valor_celda(ws, fila, COL_CONDUCTOR['remolque']),
                    )
                    
                    if tiene_hora_salida:
                        conductores_terminaron.append(conductor)
                    elif not cliente:
                        # Conductor sin viaje asignado
                        conductores_disponibles.append(conductor)
                    # Si tiene viaje pendiente, se procesará con el viaje
                
                # Procesar viaje (si hay cliente)
                if cliente:
                    viaje_datos = {}
                    for campo, columna in COL_VIAJE.items():
                        viaje_datos[campo] = self._valor_celda(ws, fila, columna)
                    
                    # Añadir datos del conductor asociado
                    viaje_datos['conductor_nombre'] = nombre_conductor
                    viaje_datos['conductor_zona'] = self._valor_celda(ws, fila, COL_CONDUCTOR['zona'])
                    viaje_datos['conductor_ubicacion'] = self._valor_celda(ws, fila, COL_CONDUCTOR['ubicacion'])
                    viaje_datos['conductor_absentismo'] = self._valor_celda(ws, fila, COL_CONDUCTOR['absentismo'])
                    viaje_datos['conductor_tractora'] = self._valor_celda(ws, fila, COL_CONDUCTOR['tractora'])
                    viaje_datos['conductor_remolque'] = self._valor_celda(ws, fila, COL_CONDUCTOR['remolque'])
                    
                    viaje = ViajePendiente(
                        fila_original=fila,
                        datos=viaje_datos
                    )
                    
                    if tiene_hora_salida:
                        viajes_completados.append(viaje)
                    else:
                        viajes_pendientes.append(viaje)
            
            wb.close()
            
            resultado = {
                'conductores_terminaron': conductores_terminaron,
                'conductores_disponibles': conductores_disponibles,
                'viajes_pendientes': viajes_pendientes,
                'viajes_completados': viajes_completados,
                'total_conductores': len(conductores_vistos),
            }
            
            logger.info(f"[CIERRE] Análisis: {len(conductores_terminaron)} terminaron, "
                       f"{len(conductores_disponibles)} disponibles, "
                       f"{len(viajes_pendientes)} viajes pendientes, "
                       f"{len(viajes_completados)} viajes completados")
            
            return resultado
            
        except Exception as e:
            logger.error(f"[CIERRE] Error analizando Excel: {e}")
            import traceback
            traceback.print_exc()
            return {
                'conductores_terminaron': [],
                'conductores_disponibles': [],
                'viajes_pendientes': [],
                'viajes_completados': [],
                'total_conductores': 0,
                'error': str(e)
            }
    
    # ============================================================
    # CREAR NUEVO EXCEL (SIN MODIFICAR EL ORIGINAL)
    # ============================================================
    
    def crear_excel_nuevo(self, fecha: datetime = None) -> Tuple[str, str]:
        """
        CREA un nuevo Excel para el día siguiente.
        NO modifica el Excel original (PRUEBO.xlsx).
        
        Args:
            fecha: Fecha para el nuevo Excel (default: hoy)
        
        Returns:
            Tuple (nombre_archivo, ruta_completa)
        """
        if fecha is None:
            fecha = datetime.now()
        
        nombre_nuevo = self.generar_nombre_excel(fecha)
        ruta_nueva = os.path.join(self.directorio, nombre_nuevo)
        
        try:
            # Cargar Excel actual como plantilla (solo lectura)
            wb_original = load_workbook(self.excel_path)
            ws_original = wb_original.active
            
            # Analizar datos actuales
            analisis = self.analizar_excel_actual()
            conductores_terminaron = analisis['conductores_terminaron']
            conductores_disponibles = analisis['conductores_disponibles']
            viajes_pendientes = analisis['viajes_pendientes']
            
            # Crear nuevo workbook
            wb_nuevo = Workbook()
            ws_nuevo = wb_nuevo.active
            ws_nuevo.title = "Rutas"
            
            # Copiar fila 1 (secciones: ZONA NORTE, FECHA, etc.)
            for col in range(1, ws_original.max_column + 1):
                valor = ws_original.cell(row=FILA_SECCIONES, column=col).value
                ws_nuevo.cell(row=FILA_SECCIONES, column=col, value=valor)
                # Copiar formato
                celda_orig = ws_original.cell(row=FILA_SECCIONES, column=col)
                celda_nueva = ws_nuevo.cell(row=FILA_SECCIONES, column=col)
                if celda_orig.font:
                    celda_nueva.font = Font(
                        bold=celda_orig.font.bold,
                        size=celda_orig.font.size
                    )
            
            # Copiar fila 2 (cabeceras)
            for col in range(1, ws_original.max_column + 1):
                valor = ws_original.cell(row=FILA_CABECERAS, column=col).value
                ws_nuevo.cell(row=FILA_CABECERAS, column=col, value=valor)
                # Copiar formato
                celda_orig = ws_original.cell(row=FILA_CABECERAS, column=col)
                celda_nueva = ws_nuevo.cell(row=FILA_CABECERAS, column=col)
                if celda_orig.font:
                    celda_nueva.font = Font(
                        bold=celda_orig.font.bold,
                        size=celda_orig.font.size
                    )
            
            # Copiar ancho de columnas
            for col in range(1, ws_original.max_column + 1):
                letra = get_column_letter(col)
                if ws_original.column_dimensions[letra].width:
                    ws_nuevo.column_dimensions[letra].width = ws_original.column_dimensions[letra].width
            
            fila_actual = FILA_INICIO_DATOS
            conductores_escritos = set()
            
            # 1. Escribir conductores que TERMINARON (ubicación actualizada, H.LL/H.SA vacías)
            for conductor in conductores_terminaron:
                if conductor.nombre.upper() in conductores_escritos:
                    continue
                conductores_escritos.add(conductor.nombre.upper())
                
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['zona'], value=conductor.zona)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['ubicacion'], value=conductor.ubicacion)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['hora_llegada'], value="VACIO")
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['hora_salida'], value="VACIO")
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['nombre'], value=conductor.nombre)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['absentismo'], value=conductor.absentismo)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['tractora'], value=conductor.tractora)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['remolque'], value=conductor.remolque)
                fila_actual += 1
            
            # 2. Escribir conductores DISPONIBLES (sin viaje, H.LL/H.SA vacías)
            for conductor in conductores_disponibles:
                if conductor.nombre.upper() in conductores_escritos:
                    continue
                conductores_escritos.add(conductor.nombre.upper())
                
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['zona'], value=conductor.zona)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['ubicacion'], value=conductor.ubicacion)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['hora_llegada'], value="VACIO")
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['hora_salida'], value="VACIO")
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['nombre'], value=conductor.nombre)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['absentismo'], value=conductor.absentismo)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['tractora'], value=conductor.tractora)
                ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['remolque'], value=conductor.remolque)
                fila_actual += 1
            
            # 3. Escribir VIAJES PENDIENTES con sus conductores
            for viaje in viajes_pendientes:
                nombre_conductor = viaje.datos.get('conductor_nombre', '')
                
                # Escribir datos del conductor (si tiene y no está duplicado)
                if nombre_conductor and nombre_conductor.upper() not in conductores_escritos:
                    conductores_escritos.add(nombre_conductor.upper())
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['zona'], 
                                 value=viaje.datos.get('conductor_zona', ''))
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['ubicacion'], 
                                 value=viaje.datos.get('conductor_ubicacion', ''))
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['hora_llegada'], value='')
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['hora_salida'], value='')
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['nombre'], value=nombre_conductor)
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['absentismo'], 
                                 value=viaje.datos.get('conductor_absentismo', ''))
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['tractora'], 
                                 value=viaje.datos.get('conductor_tractora', ''))
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['remolque'], 
                                 value=viaje.datos.get('conductor_remolque', ''))
                elif nombre_conductor:
                    # Conductor ya escrito, solo poner el nombre
                    ws_nuevo.cell(row=fila_actual, column=COL_CONDUCTOR['nombre'], value=nombre_conductor)
                
                # Escribir datos del viaje (sin horas de registro)
                for campo, columna in COL_VIAJE.items():
                    if campo.startswith('conductor_'):
                        continue  # Saltar campos de conductor
                    
                    valor = viaje.datos.get(campo, '')
                    
                    # Limpiar horas de registro para el nuevo día
                    if campo in ['hora_llegada_carga', 'hora_salida_carga', 
                                'hora_llegada_descarga', 'hora_salida_descarga']:
                        valor = ''
                    
                    ws_nuevo.cell(row=fila_actual, column=columna, value=valor)
                
                fila_actual += 1
            
            # Guardar nuevo Excel
            wb_nuevo.save(ruta_nueva)
            wb_nuevo.close()
            wb_original.close()
            
            logger.info(f"[CIERRE] ✅ Nuevo Excel CREADO: {nombre_nuevo}")
            logger.info(f"[CIERRE]    - Ruta: {ruta_nueva}")
            logger.info(f"[CIERRE]    - {len(conductores_escritos)} conductores")
            logger.info(f"[CIERRE]    - {len(viajes_pendientes)} viajes pendientes")
            
            return nombre_nuevo, ruta_nueva
            
        except Exception as e:
            logger.error(f"[CIERRE] Error creando Excel nuevo: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    # ============================================================
    # EJECUTAR CIERRE COMPLETO
    # ============================================================
    
    def ejecutar_cierre(self, fecha_nuevo: datetime = None) -> Dict:
        """
        Ejecuta el cierre de día completo.
        
        IMPORTANTE: Solo CREA un nuevo Excel, NO modifica el original.
        
        Args:
            fecha_nuevo: Fecha para el nuevo Excel (default: hoy)
        
        Returns:
            Dict con resultado del cierre
        """
        resultado = {
            'exito': False,
            'excel_original': os.path.basename(self.excel_path),
            'excel_nuevo': None,
            'ruta_excel_nuevo': None,
            'conductores_exportados': 0,
            'viajes_pendientes': 0,
            'viajes_completados': 0,
            'drive_subido': False,
            'errores': []
        }
        
        try:
            logger.info("[CIERRE] ═══════════════════════════════════════")
            logger.info("[CIERRE] INICIANDO CIERRE DE DÍA")
            logger.info(f"[CIERRE] Excel original: {self.excel_path}")
            logger.info("[CIERRE] ═══════════════════════════════════════")
            
            # Paso 1: Analizar Excel actual
            logger.info("[CIERRE] Paso 1: Analizando Excel actual...")
            analisis = self.analizar_excel_actual()
            
            if 'error' in analisis:
                resultado['errores'].append(f"Error analizando: {analisis['error']}")
                return resultado
            
            total_conductores = len(analisis['conductores_terminaron']) + len(analisis['conductores_disponibles'])
            resultado['conductores_exportados'] = total_conductores
            resultado['viajes_pendientes'] = len(analisis['viajes_pendientes'])
            resultado['viajes_completados'] = len(analisis['viajes_completados'])
            
            # Paso 2: Crear nuevo Excel (SIN modificar el original)
            logger.info("[CIERRE] Paso 2: Creando nuevo Excel...")
            nombre_nuevo, ruta_nueva = self.crear_excel_nuevo(fecha_nuevo)
            resultado['excel_nuevo'] = nombre_nuevo
            resultado['ruta_excel_nuevo'] = ruta_nueva
            
            # Paso 3: Subir el nuevo Excel a Drive
            if self.subir_archivo_nuevo:
                logger.info(f"[CIERRE] Paso 3: Subiendo {nombre_nuevo} a Drive...")
                try:
                    if self.subir_archivo_nuevo(ruta_nueva, nombre_nuevo):
                        resultado['drive_subido'] = True
                        logger.info(f"[CIERRE] ✅ {nombre_nuevo} subido a Drive")
                    else:
                        resultado['errores'].append("No se pudo subir a Drive")
                except Exception as e:
                    resultado['errores'].append(f"Error subiendo a Drive: {e}")
            
            resultado['exito'] = True
            
            logger.info("[CIERRE] ═══════════════════════════════════════")
            logger.info("[CIERRE] ✅ CIERRE DE DÍA COMPLETADO")
            logger.info(f"[CIERRE]    Excel CREADO: {nombre_nuevo}")
            logger.info(f"[CIERRE]    Excel original SIN MODIFICAR: {os.path.basename(self.excel_path)}")
            logger.info(f"[CIERRE]    Conductores: {resultado['conductores_exportados']}")
            logger.info(f"[CIERRE]    Viajes pendientes: {resultado['viajes_pendientes']}")
            logger.info(f"[CIERRE]    Viajes completados (archivados): {resultado['viajes_completados']}")
            logger.info("[CIERRE] ═══════════════════════════════════════")
            
            return resultado
            
        except Exception as e:
            logger.error(f"[CIERRE] ❌ Error en cierre: {e}")
            import traceback
            traceback.print_exc()
            resultado['errores'].append(str(e))
            return resultado
    
    # ============================================================
    # VERIFICACIONES
    # ============================================================
    
    def verificar_cierre_seguro(self) -> Dict:
        """
        Verifica si es seguro hacer el cierre.
        
        Returns:
            Dict con info de seguridad
        """
        analisis = self.analizar_excel_actual()
        
        # Verificar si ya existe Excel del día
        nombre_hoy = self.generar_nombre_excel()
        ruta_hoy = os.path.join(self.directorio, nombre_hoy)
        excel_hoy_existe = os.path.exists(ruta_hoy)
        
        total_conductores = len(analisis.get('conductores_terminaron', [])) + \
                           len(analisis.get('conductores_disponibles', []))
        
        return {
            'seguro': not excel_hoy_existe,
            'excel_hoy_existe': excel_hoy_existe,
            'nombre_excel_hoy': nombre_hoy,
            'conductores_terminaron': len(analisis.get('conductores_terminaron', [])),
            'conductores_disponibles': len(analisis.get('conductores_disponibles', [])),
            'viajes_pendientes': len(analisis.get('viajes_pendientes', [])),
            'viajes_completados': len(analisis.get('viajes_completados', [])),
            'advertencia': f"Ya existe {nombre_hoy}" if excel_hoy_existe else None
        }
    
    def listar_excels_historicos(self, limite: int = 7) -> List[Dict]:
        """Lista los Excels históricos disponibles"""
        excels = []
        
        for archivo in os.listdir(self.directorio):
            if archivo.startswith("RUTAS_") and archivo.endswith(".xlsx"):
                ruta = os.path.join(self.directorio, archivo)
                fecha_mod = datetime.fromtimestamp(os.path.getmtime(ruta))
                
                excels.append({
                    'nombre': archivo,
                    'ruta': ruta,
                    'fecha_modificacion': fecha_mod,
                    'tamaño': os.path.getsize(ruta)
                })
        
        # Ordenar por fecha descendente
        excels.sort(key=lambda x: x['fecha_modificacion'], reverse=True)
        
        return excels[:limite]


# ============================================================
# FUNCIÓN PARA INTEGRAR EN BOT
# ============================================================

def crear_cierre_dia(excel_path: str, db_path: str, subir_drive_func=None, subir_archivo_nuevo_func=None):
    """
    Crea una instancia del módulo de cierre de día.
    """
    return CierreDia(
        excel_path=excel_path,
        db_path=db_path,
        subir_drive_func=subir_drive_func,
        subir_archivo_nuevo_func=subir_archivo_nuevo_func
    )


# ============================================================
# SCRIPT PARA CRON
# ============================================================

if __name__ == "__main__":
    """
    Uso con cron:
    0 8 * * * cd /root/bot-transporte/Botttttt && python3 cierre_dia.py >> /var/log/cierre_dia.log 2>&1
    """
    from dotenv import load_dotenv
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    load_dotenv()
    
    EXCEL_EMPRESA = os.getenv("EXCEL_EMPRESA", "PRUEBO.xlsx")
    DB_PATH = os.getenv("DB_PATH", "logistica.db")
    
    cierre = CierreDia(
        excel_path=EXCEL_EMPRESA,
        db_path=DB_PATH
    )
    
    # Verificar
    verificacion = cierre.verificar_cierre_seguro()
    
    if not verificacion['seguro']:
        logger.warning(f"[CIERRE] ⚠️ {verificacion['advertencia']}")
        if len(sys.argv) <= 1 or sys.argv[1] != "--forzar":
            logger.warning("[CIERRE] Use --forzar para ignorar.")
            sys.exit(1)
    
    # Ejecutar
    resultado = cierre.ejecutar_cierre()
    
    if resultado['exito']:
        print(f"✅ Excel creado: {resultado['excel_nuevo']}")
        sys.exit(0)
    else:
        print(f"❌ Errores: {resultado['errores']}")
        sys.exit(1)
