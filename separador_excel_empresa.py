"""
SEPARADOR DE EXCEL DE EMPRESA v2.1
===================================
Lee PRUEBO.xlsx (formato de la empresa) y lo separa en tablas internas.

CAMBIOS v2.1:
- FIX Bug #1: Preserva telegram_id y telefono durante sync
- FIX Bug #2: Preserva estado de viajes durante sync

Los trabajadores siguen usando su Excel como siempre.
El sistema extrae los datos automáticamente en segundo plano.

Tablas generadas:
- conductores_empresa: Todos los conductores únicos
- viajes_empresa: Todos los viajes
- vehiculos_empresa: Tractoras y remolques
"""

import sqlite3
import pandas as pd
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from pathlib import Path
import hashlib

logger = logging.getLogger(__name__)


class SeparadorExcelEmpresa:
    """
    Separa el Excel de la empresa en tablas internas para el bot.
    Transparente para los trabajadores - ellos no notan nada.
    """
    
    # Palabras que indican que es un header, no datos
    PALABRAS_HEADER = ['TRANSPORTISTA', 'CLIENTE', 'FECHA', 'UBICACIÓN', 'TRACTORA']
    
    # Palabras que indican secciones especiales (no son conductores/viajes)
    SECCIONES_IGNORAR = ['CARGADORES', 'VENTAS', 'BITEMPERATURA', 'COLCHONETAS', 
                         'TALLER', 'PROVISIONAL', 'DEFINITIVO', 'DISCO LIMPIO']
    
    def __init__(self, db_path: str = "logistica.db"):
        self.db_path = db_path
        self.ultimo_hash = None
        self._crear_tablas()
    
    def _crear_tablas(self):
        """Crea las tablas internas si no existen"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Tabla de conductores extraídos del Excel de empresa
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS conductores_empresa (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                tractora TEXT,
                remolque TEXT,
                ubicacion TEXT,
                hora_llegada TEXT,
                hora_salida TEXT,
                absentismo TEXT,
                zona TEXT,
                fila_excel INTEGER,
                fecha_sync TEXT,
                telegram_id INTEGER,
                telefono TEXT,
                UNIQUE(nombre, tractora)
            )
        """)
        
        # Tabla de viajes extraídos del Excel de empresa (con direcciones)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS viajes_empresa (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente TEXT,
                num_pedido TEXT,
                ref_cliente TEXT,
                lugar_carga TEXT,
                direccion_carga TEXT,
                lugar_entrega TEXT,
                direccion_descarga TEXT,
                mercancia TEXT,
                precio REAL,
                km INTEGER,
                eur_km REAL,
                intercambio TEXT,
                num_pales INTEGER,
                observaciones TEXT,
                zona TEXT,
                fila_excel INTEGER,
                fecha_sync TEXT,
                conductor_asignado TEXT,
                tractora_asignada TEXT,
                estado TEXT DEFAULT 'pendiente'
            )
        """)
        
        # Tabla de vehículos (tractoras + remolques únicos)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS vehiculos_empresa (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                matricula TEXT UNIQUE,
                tipo TEXT,  -- TRACTORA o REMOLQUE
                conductor_habitual TEXT,
                fecha_sync TEXT
            )
        """)
        
        # Tabla de log de sincronización
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sync_excel_empresa (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                archivo TEXT,
                conductores_encontrados INTEGER,
                viajes_encontrados INTEGER,
                vehiculos_encontrados INTEGER,
                hash_archivo TEXT,
                estado TEXT
            )
        """)
        
        conn.commit()
        conn.close()
        logger.info("[SEPARADOR] Tablas internas creadas/verificadas")
    
    def _es_header_o_ignorar(self, valor: str) -> bool:
        """Detecta si un valor es un header o sección a ignorar"""
        if not valor:
            return True
        
        valor_upper = str(valor).upper().strip()
        
        # Es un header
        for palabra in self.PALABRAS_HEADER:
            if palabra in valor_upper:
                return True
        
        # Es una sección a ignorar
        for seccion in self.SECCIONES_IGNORAR:
            if seccion in valor_upper:
                return True
        
        return False
    
    def _detectar_zona(self, df: pd.DataFrame, fila: int) -> str:
        """Detecta la zona mirando hacia arriba desde una fila"""
        for i in range(fila, -1, -1):
            valor = df.iloc[i, 0]
            if pd.notna(valor):
                valor_str = str(valor).upper()
                if 'ZONA' in valor_str:
                    return valor_str
                if 'FECHA' in valor_str:
                    continue
        return "SIN ZONA"
    
    def _calcular_hash_archivo(self, ruta: str) -> str:
        """Calcula hash del archivo para detectar cambios"""
        with open(ruta, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()
    
    def extraer_conductores(self, df: pd.DataFrame) -> List[Dict]:
        """Extrae todos los conductores únicos del Excel"""
        conductores = []
        conductores_vistos = set()
        
        for idx in range(2, len(df)):
            nombre = df.iloc[idx, 4]  # Columna TRANSPORTISTA
            
            if pd.isna(nombre) or not str(nombre).strip():
                continue
            
            nombre_str = str(nombre).strip()
            
            if self._es_header_o_ignorar(nombre_str):
                continue
            
            tractora = str(df.iloc[idx, 6]).strip() if pd.notna(df.iloc[idx, 6]) else ""
            
            clave = f"{nombre_str}|{tractora}"
            if clave in conductores_vistos:
                continue
            conductores_vistos.add(clave)
            
            conductor = {
                "nombre": nombre_str,
                "tractora": tractora,
                "remolque": str(df.iloc[idx, 7]).strip() if pd.notna(df.iloc[idx, 7]) else "",
                "ubicacion": str(df.iloc[idx, 1]).strip() if pd.notna(df.iloc[idx, 1]) else "",
                "hora_llegada": str(df.iloc[idx, 2]).strip() if pd.notna(df.iloc[idx, 2]) else "",
                "hora_salida": str(df.iloc[idx, 3]).strip() if pd.notna(df.iloc[idx, 3]) else "",
                "absentismo": str(df.iloc[idx, 5]).strip() if pd.notna(df.iloc[idx, 5]) else "",
                "zona": self._detectar_zona(df, idx),
                "fila_excel": idx
            }
            conductores.append(conductor)
        
        logger.info(f"[SEPARADOR] Conductores extraídos: {len(conductores)}")
        return conductores
    
    def extraer_viajes(self, df: pd.DataFrame) -> List[Dict]:
        """Extrae todos los viajes del Excel (con direcciones)"""
        viajes = []
        
        for idx in range(2, len(df)):
            cliente = df.iloc[idx, 8]
            
            if pd.isna(cliente) or not str(cliente).strip():
                continue
            
            cliente_str = str(cliente).strip()
            
            if self._es_header_o_ignorar(cliente_str):
                continue
            
            lugar_carga = str(df.iloc[idx, 13]).strip() if pd.notna(df.iloc[idx, 13]) else ""
            lugar_entrega = str(df.iloc[idx, 16]).strip() if pd.notna(df.iloc[idx, 16]) else ""
            
            if not lugar_carga and not lugar_entrega:
                continue
            
            direccion_carga = ""
            direccion_descarga = ""
            
            if len(df.columns) > 28 and pd.notna(df.iloc[idx, 28]):
                direccion_carga = str(df.iloc[idx, 28]).strip()
            
            if len(df.columns) > 29 and pd.notna(df.iloc[idx, 29]):
                direccion_descarga = str(df.iloc[idx, 29]).strip()
            
            try:
                precio = float(df.iloc[idx, 22]) if pd.notna(df.iloc[idx, 22]) else 0
            except:
                precio = 0
            
            try:
                km = int(df.iloc[idx, 23]) if pd.notna(df.iloc[idx, 23]) else 0
            except:
                km = 0
            
            try:
                eur_km = float(df.iloc[idx, 24]) if pd.notna(df.iloc[idx, 24]) else 0
            except:
                eur_km = 0
            
            try:
                num_pales = int(df.iloc[idx, 12]) if pd.notna(df.iloc[idx, 12]) else 0
            except:
                num_pales = 0
            
            viaje = {
                "cliente": cliente_str,
                "num_pedido": str(df.iloc[idx, 9]).strip() if pd.notna(df.iloc[idx, 9]) else "",
                "ref_cliente": str(df.iloc[idx, 10]).strip() if pd.notna(df.iloc[idx, 10]) else "",
                "lugar_carga": lugar_carga,
                "direccion_carga": direccion_carga,
                "lugar_entrega": lugar_entrega,
                "direccion_descarga": direccion_descarga,
                "mercancia": str(df.iloc[idx, 19]).strip() if pd.notna(df.iloc[idx, 19]) else "",
                "precio": precio,
                "km": km,
                "eur_km": eur_km,
                "intercambio": str(df.iloc[idx, 11]).strip() if pd.notna(df.iloc[idx, 11]) else "",
                "num_pales": num_pales,
                "observaciones": str(df.iloc[idx, 27]).strip() if pd.notna(df.iloc[idx, 27]) else "",
                "zona": self._detectar_zona(df, idx),
                "fila_excel": idx,
                "conductor_asignado": str(df.iloc[idx, 4]).strip() if pd.notna(df.iloc[idx, 4]) else "",
                "tractora_asignada": str(df.iloc[idx, 6]).strip() if pd.notna(df.iloc[idx, 6]) else ""
            }
            viajes.append(viaje)
        
        logger.info(f"[SEPARADOR] Viajes extraídos: {len(viajes)}")
        return viajes
    
    def extraer_vehiculos(self, conductores: List[Dict]) -> List[Dict]:
        """Extrae vehículos únicos (tractoras y remolques)"""
        vehiculos = []
        matriculas_vistas = set()
        
        for c in conductores:
            if c["tractora"] and c["tractora"] not in matriculas_vistas:
                if not self._es_header_o_ignorar(c["tractora"]):
                    matriculas_vistas.add(c["tractora"])
                    vehiculos.append({
                        "matricula": c["tractora"],
                        "tipo": "TRACTORA",
                        "conductor_habitual": c["nombre"]
                    })
            
            if c["remolque"] and c["remolque"] not in matriculas_vistas:
                if not self._es_header_o_ignorar(c["remolque"]):
                    matriculas_vistas.add(c["remolque"])
                    vehiculos.append({
                        "matricula": c["remolque"],
                        "tipo": "REMOLQUE",
                        "conductor_habitual": c["nombre"]
                    })
        
        logger.info(f"[SEPARADOR] Vehículos extraídos: {len(vehiculos)}")
        return vehiculos
    
    def _preservar_datos_conductores(self, cursor) -> Dict[str, Dict]:
        """
        BUGFIX #1: Guarda telegram_id y telefono antes de borrar.
        Returns: Dict {nombre_upper: {'telegram_id': X, 'telefono': Y}}
        """
        datos_preservados = {}
        cursor.execute("""
            SELECT nombre, telegram_id, telefono 
            FROM conductores_empresa 
            WHERE telegram_id IS NOT NULL OR telefono IS NOT NULL
        """)
        for row in cursor.fetchall():
            nombre = row[0].upper().strip() if row[0] else ""
            if nombre:
                datos_preservados[nombre] = {
                    'telegram_id': row[1],
                    'telefono': row[2]
                }
        logger.info(f"[SEPARADOR] Preservados datos de {len(datos_preservados)} conductores")
        return datos_preservados
    
    def _preservar_estados_viajes(self, cursor) -> Dict[str, str]:
        """
        BUGFIX #2: Guarda estado de viajes antes de borrar.
        Returns: Dict {clave_viaje: estado}
        """
        estados_preservados = {}
        cursor.execute("""
            SELECT cliente, lugar_carga, lugar_entrega, conductor_asignado, estado 
            FROM viajes_empresa 
            WHERE estado IS NOT NULL AND estado != 'pendiente'
        """)
        for row in cursor.fetchall():
            cliente = (row[0] or "").upper().strip()
            carga = (row[1] or "").upper().strip()
            entrega = (row[2] or "").upper().strip()
            conductor = (row[3] or "").upper().strip()
            clave = f"{cliente}|{carga}|{entrega}|{conductor}"
            estados_preservados[clave] = row[4]
        logger.info(f"[SEPARADOR] Preservados estados de {len(estados_preservados)} viajes")
        return estados_preservados
    
    def _restaurar_datos_conductores(self, cursor, datos_preservados: Dict[str, Dict]):
        """BUGFIX #1: Restaura telegram_id y telefono después de insertar"""
        restaurados = 0
        for nombre_upper, datos in datos_preservados.items():
            cursor.execute("""
                UPDATE conductores_empresa 
                SET telegram_id = ?, telefono = ?
                WHERE UPPER(TRIM(nombre)) = ?
            """, (datos['telegram_id'], datos['telefono'], nombre_upper))
            if cursor.rowcount > 0:
                restaurados += 1
        logger.info(f"[SEPARADOR] Restaurados datos de {restaurados} conductores")
    
    def _restaurar_estados_viajes(self, cursor, estados_preservados: Dict[str, str]):
        """BUGFIX #2: Restaura estados de viajes después de insertar"""
        restaurados = 0
        for clave, estado in estados_preservados.items():
            partes = clave.split("|")
            if len(partes) == 4:
                cliente, carga, entrega, conductor = partes
                cursor.execute("""
                    UPDATE viajes_empresa 
                    SET estado = ?
                    WHERE UPPER(TRIM(cliente)) = ? 
                    AND UPPER(TRIM(lugar_carga)) = ?
                    AND UPPER(TRIM(lugar_entrega)) = ?
                    AND UPPER(TRIM(conductor_asignado)) = ?
                """, (estado, cliente, carga, entrega, conductor))
                if cursor.rowcount > 0:
                    restaurados += 1
        logger.info(f"[SEPARADOR] Restaurados estados de {restaurados} viajes")
    
    def guardar_en_bd(self, conductores: List[Dict], viajes: List[Dict], 
                      vehiculos: List[Dict], archivo: str, hash_archivo: str):
        """Guarda los datos extraídos en la base de datos (preservando telegram_id y estados)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        fecha_sync = datetime.now().isoformat()
        
        try:
            # BUGFIX #1 y #2: Preservar datos antes de borrar
            datos_conductores = self._preservar_datos_conductores(cursor)
            estados_viajes = self._preservar_estados_viajes(cursor)
            
            # Limpiar tablas anteriores
            cursor.execute("DELETE FROM conductores_empresa")
            cursor.execute("DELETE FROM viajes_empresa")
            cursor.execute("DELETE FROM vehiculos_empresa")
            
            # Insertar conductores
            for c in conductores:
                cursor.execute("""
                    INSERT INTO conductores_empresa
                    (nombre, tractora, remolque, ubicacion, hora_llegada, hora_salida,
                     absentismo, zona, fila_excel, fecha_sync)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (c["nombre"], c["tractora"], c["remolque"], c["ubicacion"],
                      c["hora_llegada"], c["hora_salida"], c["absentismo"],
                      c["zona"], c["fila_excel"], fecha_sync))
            
            # Insertar viajes
            for v in viajes:
                cursor.execute("""
                    INSERT INTO viajes_empresa
                    (cliente, num_pedido, ref_cliente, lugar_carga, direccion_carga,
                     lugar_entrega, direccion_descarga, mercancia, precio, km, eur_km,
                     intercambio, num_pales, observaciones, zona, fila_excel, fecha_sync,
                     conductor_asignado, tractora_asignada)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (v["cliente"], v["num_pedido"], v["ref_cliente"], 
                      v["lugar_carga"], v["direccion_carga"],
                      v["lugar_entrega"], v["direccion_descarga"],
                      v["mercancia"], v["precio"], v["km"], v["eur_km"],
                      v["intercambio"], v["num_pales"], v["observaciones"],
                      v["zona"], v["fila_excel"], fecha_sync,
                      v["conductor_asignado"], v["tractora_asignada"]))
            
            # Insertar vehículos
            for veh in vehiculos:
                cursor.execute("""
                    INSERT OR REPLACE INTO vehiculos_empresa
                    (matricula, tipo, conductor_habitual, fecha_sync)
                    VALUES (?, ?, ?, ?)
                """, (veh["matricula"], veh["tipo"], veh["conductor_habitual"], fecha_sync))
            
            # BUGFIX #1 y #2: Restaurar datos preservados
            self._restaurar_datos_conductores(cursor, datos_conductores)
            self._restaurar_estados_viajes(cursor, estados_viajes)
            
            # Log de sincronización
            cursor.execute("""
                INSERT INTO sync_excel_empresa
                (fecha, archivo, conductores_encontrados, viajes_encontrados, 
                 vehiculos_encontrados, hash_archivo, estado)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (fecha_sync, archivo, len(conductores), len(viajes), 
                  len(vehiculos), hash_archivo, "exitoso"))
            
            conn.commit()
            logger.info(f"[SEPARADOR] Datos guardados: {len(conductores)} conductores, "
                       f"{len(viajes)} viajes, {len(vehiculos)} vehículos")
            
        except Exception as e:
            conn.rollback()
            logger.error(f"[SEPARADOR] Error guardando: {e}")
            raise
        finally:
            conn.close()
    
    def sincronizar_desde_archivo(self, ruta_excel: str, forzar: bool = False) -> Dict:
        """Sincroniza el Excel de la empresa a las tablas internas."""
        logger.info(f"[SEPARADOR] Iniciando sincronización: {ruta_excel}")
        
        if not Path(ruta_excel).exists():
            logger.error(f"[SEPARADOR] Archivo no encontrado: {ruta_excel}")
            return {"exito": False, "error": "Archivo no encontrado"}
        
        hash_actual = self._calcular_hash_archivo(ruta_excel)
        
        if not forzar and hash_actual == self.ultimo_hash:
            logger.debug("[SEPARADOR] Sin cambios en el archivo")
            return {"exito": True, "cambios": False}
        
        try:
            df = pd.read_excel(ruta_excel, header=None)
            logger.info(f"[SEPARADOR] Excel leído: {len(df)} filas, {len(df.columns)} columnas")
            
            conductores = self.extraer_conductores(df)
            viajes = self.extraer_viajes(df)
            vehiculos = self.extraer_vehiculos(conductores)
            
            self.guardar_en_bd(conductores, viajes, vehiculos, ruta_excel, hash_actual)
            
            self.ultimo_hash = hash_actual
            
            return {
                "exito": True,
                "cambios": True,
                "conductores": len(conductores),
                "viajes": len(viajes),
                "vehiculos": len(vehiculos)
            }
            
        except Exception as e:
            logger.error(f"[SEPARADOR] Error: {e}", exc_info=True)
            return {"exito": False, "error": str(e)}
    
    def vincular_conductor_telegram(self, nombre: str, telegram_id: int, 
                                    telefono: str = None) -> bool:
        """Vincula un conductor con su Telegram ID"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                UPDATE conductores_empresa 
                SET telegram_id = ?, telefono = ?
                WHERE nombre LIKE ?
            """, (telegram_id, telefono, f"%{nombre}%"))
            
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error vinculando: {e}")
            return False
        finally:
            conn.close()
    
    def obtener_conductor_por_telegram(self, telegram_id: int) -> Optional[Dict]:
        """Obtiene datos de un conductor por su Telegram ID"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT nombre, tractora, remolque, ubicacion, zona, telefono
            FROM conductores_empresa
            WHERE telegram_id = ?
        """, (telegram_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return {
                "nombre": row[0],
                "tractora": row[1],
                "remolque": row[2],
                "ubicacion": row[3],
                "zona": row[4],
                "telefono": row[5]
            }
        return None
    
    def obtener_viajes_conductor(self, nombre: str) -> List[Dict]:
        """Obtiene los viajes asignados a un conductor"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT * FROM viajes_empresa
            WHERE conductor_asignado LIKE ?
            ORDER BY fila_excel
        """, (f"%{nombre}%",))
        
        viajes = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return viajes
    
    def obtener_resumen(self) -> Dict:
        """Obtiene un resumen de los datos extraídos"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM conductores_empresa")
        total_conductores = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM viajes_empresa")
        total_viajes = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM vehiculos_empresa")
        total_vehiculos = cursor.fetchone()[0]
        
        cursor.execute("SELECT DISTINCT zona FROM conductores_empresa WHERE zona != ''")
        zonas = [row[0] for row in cursor.fetchall()]
        
        cursor.execute("SELECT fecha FROM sync_excel_empresa ORDER BY id DESC LIMIT 1")
        row = cursor.fetchone()
        ultima_sync = row[0] if row else "Nunca"
        
        conn.close()
        
        return {
            "conductores": total_conductores,
            "viajes": total_viajes,
            "vehiculos": total_vehiculos,
            "zonas": zonas,
            "ultima_sincronizacion": ultima_sync
        }
    
    def actualizar_transportista_excel(self, ruta_excel: str, fila_excel: int, 
                                        nombre_conductor: str) -> bool:
        """Actualiza la columna TRANSPORTISTA en el Excel."""
        try:
            from openpyxl import load_workbook
            
            if not Path(ruta_excel).exists():
                logger.error(f"[SEPARADOR] Excel no encontrado: {ruta_excel}")
                return False
            
            wb = load_workbook(ruta_excel)
            ws = wb.active
            
            fila_openpyxl = fila_excel + 1
            columna_transportista = 22
            
            if fila_openpyxl > ws.max_row:
                logger.error(f"[SEPARADOR] Fila {fila_openpyxl} fuera de rango")
                return False
            
            celda = ws.cell(row=fila_openpyxl, column=columna_transportista)
            valor_anterior = celda.value
            celda.value = nombre_conductor
            
            wb.save(ruta_excel)
            wb.close()
            
            logger.info(f"[SEPARADOR] ✅ Excel actualizado: Fila {fila_openpyxl}, "
                       f"Col V = '{nombre_conductor}' (antes: '{valor_anterior}')")
            return True
            
        except Exception as e:
            logger.error(f"[SEPARADOR] Error actualizando Excel: {e}")
            return False
    
    def actualizar_asignacion_viaje(self, viaje_id: int, nombre_conductor: str,
                                     ruta_excel: str) -> Dict:
        """Actualiza la asignación de un viaje tanto en BD como en Excel."""
        resultado = {
            "exito": False,
            "bd_actualizada": False,
            "excel_actualizado": False,
            "fila_excel": None
        }
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT fila_excel FROM viajes_empresa WHERE id = ?", (viaje_id,))
            row = cursor.fetchone()
            
            if not row:
                logger.error(f"[SEPARADOR] Viaje {viaje_id} no encontrado")
                conn.close()
                return resultado
            
            fila_excel = row[0]
            resultado["fila_excel"] = fila_excel
            
            cursor.execute("""
                UPDATE viajes_empresa 
                SET conductor_asignado = ?
                WHERE id = ?
            """, (nombre_conductor, viaje_id))
            conn.commit()
            conn.close()
            resultado["bd_actualizada"] = True
            
            if fila_excel and ruta_excel:
                resultado["excel_actualizado"] = self.actualizar_transportista_excel(
                    ruta_excel, fila_excel, nombre_conductor
                )
            
            resultado["exito"] = resultado["bd_actualizada"] and resultado["excel_actualizado"]
            return resultado
            
        except Exception as e:
            logger.error(f"[SEPARADOR] Error en actualizar_asignacion_viaje: {e}")
            return resultado
