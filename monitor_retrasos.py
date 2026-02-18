"""
MONITOR DE RETRASOS v1.0
=========================
Detecta retrasos en llegadas de carga/descarga y alerta al admin.

LÓGICA:
- Lee hora_carga / hora_descarga de viajes_empresa (formato "09:00-11:00")
- Comprueba si el conductor ha registrado llegada en Excel (col O / col R)
- Si ha pasado el 25% de la franja sin registro → alerta admin
- Evita alertas duplicadas (memoria por viaje+tipo)

INTEGRACIÓN:
    from monitor_retrasos import MonitorRetrasos

    monitor = MonitorRetrasos(db_path, excel_path, bot, admin_ids)
    # En job_queue:
    app.job_queue.run_repeating(monitor.verificar_retrasos, interval=300, first=60)
"""

import sqlite3
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Set, Tuple

logger = logging.getLogger(__name__)

# Columnas Excel (openpyxl, 1-indexed)
COL_CARGA_LLEGADA = 15     # Col O
COL_DESCARGA_LLEGADA = 18  # Col R


class MonitorRetrasos:
    """
    Monitoriza franjas horarias de viajes y alerta si no hay registro de llegada
    cuando ha transcurrido el 25% de la franja.
    """

    def __init__(self, db_path: str, excel_path: str, bot=None, admin_ids: list = None):
        self.db_path = db_path
        self.excel_path = excel_path
        self.bot = bot
        self.admin_ids = admin_ids or []

        # Set de alertas ya enviadas: "viaje_id:tipo" (ej: "5317:carga")
        self._alertas_enviadas: Set[str] = set()

        # Umbral: porcentaje de franja transcurrido para alertar (0.25 = 25%)
        self.umbral = 0.25

        logger.info(f"[MONITOR] Inicializado v1.0 | umbral={int(self.umbral*100)}% | admins={len(self.admin_ids)}")

    # ----------------------------------------------------------------
    # MÉTODO PRINCIPAL (se llama desde job_queue)
    # ----------------------------------------------------------------
    async def verificar_retrasos(self, context=None):
        """
        Comprueba todos los viajes activos y alerta si hay retrasos.
        Compatible con job_queue.run_repeating().
        """
        try:
            viajes = self._obtener_viajes_activos()
            if not viajes:
                return

            ahora = datetime.now()
            alertas = []

            for viaje in viajes:
                viaje_id = viaje['id']
                conductor = viaje['conductor_asignado'] or 'Sin asignar'
                fila_excel = viaje['fila_excel']
                cliente = viaje['cliente'] or ''
                lugar_carga = viaje['lugar_carga'] or ''
                lugar_descarga = viaje['lugar_entrega'] or ''

                # --- Comprobar CARGA ---
                hora_carga = viaje.get('hora_carga') or ''
                if hora_carga and '-' in hora_carga:
                    clave_carga = f"{viaje_id}:carga"
                    if clave_carga not in self._alertas_enviadas:
                        llegada_carga = self._leer_celda_excel(fila_excel, COL_CARGA_LLEGADA)
                        if not llegada_carga:
                            fecha_carga = viaje.get('fecha_carga') or ''
                            retraso = self._calcular_retraso(hora_carga, fecha_carga, ahora)
                            if retraso:
                                alertas.append({
                                    'viaje_id': viaje_id,
                                    'tipo': 'CARGA',
                                    'conductor': conductor,
                                    'cliente': cliente,
                                    'lugar': lugar_carga,
                                    'franja': hora_carga,
                                    'fecha': fecha_carga,
                                    'minutos_retraso': retraso['minutos_pasados'],
                                    'minutos_franja': retraso['minutos_franja'],
                                    'porcentaje': retraso['porcentaje'],
                                    'clave': clave_carga,
                                })

                # --- Comprobar DESCARGA ---
                hora_descarga = viaje.get('hora_descarga') or ''
                if hora_descarga and '-' in hora_descarga:
                    clave_descarga = f"{viaje_id}:descarga"
                    if clave_descarga not in self._alertas_enviadas:
                        llegada_descarga = self._leer_celda_excel(fila_excel, COL_DESCARGA_LLEGADA)
                        if not llegada_descarga:
                            fecha_descarga = viaje.get('fecha_descarga') or ''
                            retraso = self._calcular_retraso(hora_descarga, fecha_descarga, ahora)
                            if retraso:
                                alertas.append({
                                    'viaje_id': viaje_id,
                                    'tipo': 'DESCARGA',
                                    'conductor': conductor,
                                    'cliente': cliente,
                                    'lugar': lugar_descarga,
                                    'franja': hora_descarga,
                                    'fecha': fecha_descarga,
                                    'minutos_retraso': retraso['minutos_pasados'],
                                    'minutos_franja': retraso['minutos_franja'],
                                    'porcentaje': retraso['porcentaje'],
                                    'clave': clave_descarga,
                                })

            # Enviar alertas
            if alertas:
                await self._enviar_alertas(alertas, context)

        except Exception as e:
            logger.error(f"[MONITOR] Error verificando retrasos: {e}", exc_info=True)

    # ----------------------------------------------------------------
    # LECTURA BD
    # ----------------------------------------------------------------
    def _obtener_viajes_activos(self) -> list:
        """Obtiene viajes con conductor asignado y estado pendiente/en_curso."""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, cliente, num_pedido, lugar_carga, lugar_entrega,
                       conductor_asignado, fila_excel, estado,
                       hora_carga, hora_descarga, fecha_carga, fecha_descarga
                FROM viajes_empresa
                WHERE conductor_asignado IS NOT NULL
                  AND conductor_asignado != ''
                  AND estado IN ('pendiente', 'en_curso', 'asignado')
                  AND (hora_carga IS NOT NULL OR hora_descarga IS NOT NULL)
            """)

            viajes = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return viajes

        except Exception as e:
            logger.error(f"[MONITOR] Error leyendo BD: {e}")
            return []

    # ----------------------------------------------------------------
    # LECTURA EXCEL
    # ----------------------------------------------------------------
    def _leer_celda_excel(self, fila_excel: int, columna: int) -> Optional[str]:
        """
        Lee una celda del Excel.
        fila_excel: índice 0-based (se suma 1 para openpyxl).
        columna: índice 1-based (openpyxl).
        """
        try:
            from openpyxl import load_workbook

            if not Path(self.excel_path).exists():
                return None

            wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            ws = wb.active

            fila_openpyxl = fila_excel + 1

            if fila_openpyxl > ws.max_row:
                wb.close()
                return None

            valor = ws.cell(row=fila_openpyxl, column=columna).value
            wb.close()

            if valor and str(valor).strip():
                return str(valor).strip()

            return None

        except Exception as e:
            logger.error(f"[MONITOR] Error leyendo Excel fila={fila_excel} col={columna}: {e}")
            return None

    # ----------------------------------------------------------------
    # CÁLCULO DE RETRASO
    # ----------------------------------------------------------------
    def _calcular_retraso(self, franja: str, fecha_str: str, ahora: datetime) -> Optional[dict]:
        """
        Calcula si ha pasado el umbral (25%) de la franja horaria.

        Args:
            franja: "09:00-11:00" o "08:00-14:00"
            fecha_str: "18/02/2026" o "" (si vacío, asume hoy)
            ahora: datetime actual

        Returns:
            dict con info de retraso si se supera el umbral, None si no.
        """
        try:
            partes = franja.split('-')
            if len(partes) != 2:
                return None

            hora_inicio_str = partes[0].strip()
            hora_fin_str = partes[1].strip()

            hora_inicio = datetime.strptime(hora_inicio_str, "%H:%M")
            hora_fin = datetime.strptime(hora_fin_str, "%H:%M")

            # Determinar la fecha del viaje
            if fecha_str and fecha_str.strip():
                try:
                    fecha_viaje = datetime.strptime(fecha_str.strip(), "%d/%m/%Y").date()
                except ValueError:
                    # Intentar formato alternativo
                    try:
                        fecha_viaje = datetime.strptime(fecha_str.strip(), "%Y-%m-%d").date()
                    except ValueError:
                        fecha_viaje = ahora.date()
            else:
                fecha_viaje = ahora.date()

            # Solo comprobar viajes de hoy
            if fecha_viaje != ahora.date():
                return None

            # Construir datetimes completos
            dt_inicio = ahora.replace(
                hour=hora_inicio.hour, minute=hora_inicio.minute,
                second=0, microsecond=0
            )
            dt_fin = ahora.replace(
                hour=hora_fin.hour, minute=hora_fin.minute,
                second=0, microsecond=0
            )

            # Manejar franjas que cruzan medianoche (ej: 22:00-06:00)
            if dt_fin <= dt_inicio:
                dt_fin += timedelta(days=1)

            minutos_franja = (dt_fin - dt_inicio).total_seconds() / 60
            if minutos_franja <= 0:
                return None

            # Calcular punto de alerta (25% de la franja)
            minutos_umbral = minutos_franja * self.umbral
            dt_umbral = dt_inicio + timedelta(minutes=minutos_umbral)

            # ¿Ya pasó el umbral?
            if ahora < dt_umbral:
                return None

            # ¿Ya pasó toda la franja? (también alertar, es peor)
            minutos_pasados = (ahora - dt_inicio).total_seconds() / 60

            porcentaje = min(minutos_pasados / minutos_franja * 100, 100)

            return {
                'minutos_pasados': round(minutos_pasados),
                'minutos_franja': round(minutos_franja),
                'porcentaje': round(porcentaje),
            }

        except Exception as e:
            logger.error(f"[MONITOR] Error calculando retraso '{franja}': {e}")
            return None

    # ----------------------------------------------------------------
    # ENVÍO DE ALERTAS
    # ----------------------------------------------------------------
    async def _enviar_alertas(self, alertas: list, context=None):
        """Envía alertas agrupadas a los admins."""
        if not self.bot and not context:
            logger.warning("[MONITOR] No hay bot/context para enviar alertas")
            return

        bot = context.bot if context else self.bot

        # Agrupar por conductor para no spamear
        por_conductor = {}
        for alerta in alertas:
            conductor = alerta['conductor']
            if conductor not in por_conductor:
                por_conductor[conductor] = []
            por_conductor[conductor].append(alerta)

        for conductor, lista in por_conductor.items():
            lineas = [f"⚠️ *ALERTA RETRASO* — {conductor}\n"]

            for a in lista:
                emoji = "📦" if a['tipo'] == 'CARGA' else "📤"
                estado_barra = self._barra_progreso(a['porcentaje'])

                lineas.append(
                    f"{emoji} *{a['tipo']}* en {a['lugar']}\n"
                    f"   Cliente: {a['cliente']}\n"
                    f"   Franja: {a['franja']} ({a['fecha'] or 'hoy'})\n"
                    f"   {estado_barra} {a['porcentaje']}% transcurrido "
                    f"({a['minutos_retraso']}min de {a['minutos_franja']}min)\n"
                    f"   ❌ Sin registro de llegada\n"
                )

                # Marcar como alertado
                self._alertas_enviadas.add(a['clave'])

            mensaje = "\n".join(lineas)

            for admin_id in self.admin_ids:
                try:
                    await bot.send_message(
                        chat_id=admin_id,
                        text=mensaje,
                        parse_mode='Markdown'
                    )
                    logger.info(f"[MONITOR] Alerta enviada a admin {admin_id}: {conductor}")
                except Exception as e:
                    logger.error(f"[MONITOR] Error enviando alerta a {admin_id}: {e}")

    # ----------------------------------------------------------------
    # UTILIDADES
    # ----------------------------------------------------------------
    @staticmethod
    def _barra_progreso(porcentaje: int) -> str:
        """Genera barra visual de progreso."""
        bloques = 10
        llenos = max(1, round(porcentaje / 100 * bloques))
        vacios = bloques - llenos
        return f"[{'█' * llenos}{'░' * vacios}]"

    def resetear_alertas_diarias(self):
        """
        Limpia alertas enviadas. Llamar al inicio de cada día
        (por ejemplo desde cierre_dia o un job a las 00:00).
        """
        cantidad = len(self._alertas_enviadas)
        self._alertas_enviadas.clear()
        logger.info(f"[MONITOR] Reset diario: {cantidad} alertas limpiadas")

    def get_estado(self) -> dict:
        """Devuelve estado actual del monitor para diagnóstico."""
        return {
            'alertas_enviadas': len(self._alertas_enviadas),
            'detalle': list(self._alertas_enviadas),
            'umbral': f"{int(self.umbral * 100)}%",
            'admin_ids': self.admin_ids,
        }
